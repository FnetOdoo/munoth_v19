import io
import base64
import logging
from odoo import models, fields, api

_logger = logging.getLogger(__name__)


class ResPartner(models.Model):
    _inherit = "res.partner"

    # ── Scheduled Action entry point ──────────────────────────
    @api.model
    def action_export_contacts_xlsx(self):
        """Call this method from Scheduled Action."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            _logger.error("openpyxl not installed. Run: pip install openpyxl")
            return

        contacts = self.env["res.partner"].search_read(
            [["active", "=", True]],
            fields=[
                "id", "name", "email", "phone", "pan_no",
                "parent_id", "street", "city", "country_id", "is_company"
            ],
            order="name asc",
            limit=5000,
        )

        # ── Build XLSX ────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Odoo Contacts"

        header_font  = Font(bold=True, color="FFFFFF", size=11)
        header_fill  = PatternFill("solid", fgColor="4B3FBD")
        center_align = Alignment(horizontal="center", vertical="center")

        headers    = ["Odoo ID", "Name", "Email", "Phone", "pan_no",
                      "Company", "Street", "City", "Country", "Type"]
        col_widths = [10, 30, 32, 18, 18, 28, 28, 18, 18, 12]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell            = ws.cell(row=1, column=col_idx, value=header)
            cell.font       = header_font
            cell.fill       = header_fill
            cell.alignment  = center_align
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 22

        for row_idx, c in enumerate(contacts, start=2):
            parent       = c.get("parent_id")
            company      = parent[1] if parent else ""
            country      = c.get("country_id")
            country_name = country[1] if country else ""
            contact_type = "Company" if c.get("is_company") else "Individual"

            row_data = [
                c.get("id", ""),
                c.get("name", "")    or "",
                c.get("email", "")   or "",
                c.get("phone", "")   or "",
                c.get("mobile", "")  or "",
                company,
                c.get("street", "")  or "",
                c.get("city", "")    or "",
                country_name,
                contact_type,
            ]

            fill_color = "F3F2FF" if row_idx % 2 == 0 else "FFFFFF"
            row_fill   = PatternFill("solid", fgColor=fill_color)

            for col_idx, value in enumerate(row_data, start=1):
                cell           = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill      = row_fill
                cell.alignment = Alignment(vertical="center")

        ws.freeze_panes = "A2"

        # ── Save file to /tmp ─────────────────────────────────
        file_name = "/home/hari45/odoo_contacts.xlsx"
        wb.save(file_name)
        _logger.info("✅ Contacts exported to %s (%d records)", file_name, len(contacts))




class DummyPartner(models.Model):
    _name        = "dummy.partner"
    _description = "Dummy Partner from XLSX"

    name    = fields.Char(string="Name")
    email   = fields.Char(string="Email")
    phone   = fields.Char(string="Phone")
    pan_no  = fields.Char(string="PAN No")

    @api.model
    def action_import_from_remote_xlsx(self):
        """
        1. Connect to remote server via IP (SFTP)
        2. Download XLSX file
        3. Read rows
        4. Create dummy.partner records on THIS Odoo server
        """
        try:
            import paramiko
            import openpyxl
            import io
        except ImportError:
            _logger.error("Run: pip install paramiko openpyxl")
            return

        # ── Remote Server Config ──────────────────────────────
        REMOTE_HOST     = "192.168.105.129"        # ← remote server IP
        REMOTE_PORT     = 22                      # ← SSH port (default 22)
        REMOTE_USER     = "sasidharan"                # ← remote server username
        REMOTE_PASSWORD = "sasi@123"         # ← remote server password
        REMOTE_PATH     = "/home/sasidharan/Downloads/odoo_contacts.xlsx"  # ← XLSX path on remote

        # ── Step 1: Connect to remote server via SFTP ─────────
        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(
                hostname = REMOTE_HOST,
                port     = REMOTE_PORT,
                username = REMOTE_USER,
                password = REMOTE_PASSWORD,
                timeout  = 10,
            )
            _logger.info("✅ Connected to remote server: %s", REMOTE_HOST)
        except Exception as e:
            _logger.error("❌ Cannot connect to remote server: %s", str(e))
            return

        # ── Step 2: Download XLSX into memory ─────────────────
        try:
            sftp       = ssh.open_sftp()
            file_obj   = io.BytesIO()
            sftp.getfo(REMOTE_PATH, file_obj)
            file_obj.seek(0)
            sftp.close()
            ssh.close()
            _logger.info("✅ XLSX downloaded from: %s", REMOTE_PATH)
        except Exception as e:
            _logger.error("❌ Cannot read XLSX from remote: %s", str(e))
            ssh.close()
            return

        # ── Step 3: Read XLSX rows ─────────────────────────────
        try:
            wb   = openpyxl.load_workbook(file_obj)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            _logger.error("❌ Cannot parse XLSX: %s", str(e))
            return

        if not rows:
            _logger.warning("XLSX file is empty.")
            return

        # Skip header row
        data_rows = rows[1:]

        # ── Step 4: Create records on THIS Odoo server ────────
        created  = 0
        skipped  = 0

        for row in data_rows:
            try:
                name   = row[1] or ""
                email  = row[2] or ""
                phone  = row[3] or ""
                pan_no = row[4] or ""
            except IndexError:
                continue

            if not name:
                continue

            # Skip if already exists
            existing = self.env["dummy.partner"].search(
                [["email", "=", email]], limit=1
            )
            if existing:
                skipped += 1
                continue

            self.env["dummy.partner"].create({
                "name":   name,
                "email":  email,
                "phone":  phone,
                "pan_no": pan_no,
            })
            created += 1

        _logger.info(
            "✅ Import done — Created: %d | Skipped (duplicate): %d",
            created, skipped
        )

    # @api.model
    # def action_export_to_google_sheet(self):
    #     """
    #     Fetch first 10 contacts from Odoo
    #     and write them to Google Sheet.
    #     """
    #     try:
    #         import gspread
    #         from google.oauth2.service_account import Credentials
    #     except ImportError:
    #         _logger.error("Run: pip install gspread google-auth")
    #         return
    #
    #     # ── Google Sheet Config ───────────────────────────────
    #     CREDS_FILE = "/home/hari45/service_account.json"
    #     SHEET_ID = "1b7QfVPA2MAlyouhGPzhzo74wSyZ9sTb_RyWw5gUkvEI"
    #     SHEET_NAME = "Sheet1"
    #
    #     # ── Step 1: Fetch only 10 contacts from Odoo ─────────
    #     contacts = self.env["res.partner"].search_read(
    #         [["active", "=", True]],
    #         fields=["id", "name", "email", "phone", "pan_no"],
    #         order="name asc",
    #         limit=10,  # ← only 10 records
    #     )
    #
    #     if not contacts:
    #         _logger.warning("No contacts found in Odoo.")
    #         return
    #
    #     _logger.info("✅ Fetched %d contacts from Odoo", len(contacts))
    #
    #     # ── Step 2: Connect to Google Sheet ──────────────────
    #     try:
    #         scopes = [
    #             "https://www.googleapis.com/auth/spreadsheets",
    #             "https://www.googleapis.com/auth/drive",
    #         ]
    #         creds = Credentials.from_service_account_file(
    #             CREDS_FILE, scopes=scopes
    #         )
    #         gc = gspread.authorize(creds)
    #         sheet = gc.open_by_key(SHEET_ID).worksheet(SHEET_NAME)
    #         _logger.info("✅ Connected to Google Sheet")
    #     except FileNotFoundError:
    #         _logger.error("❌ service_account.json not found at: %s", CREDS_FILE)
    #         return
    #     except Exception as e:
    #         _logger.error("❌ Google Sheet connection failed: %s", str(e))
    #         return
    #
    #     # ── Step 3: Build rows ────────────────────────────────
    #     header = ["Odoo ID", "Name", "Email", "Phone", "PAN No"]
    #     rows = [header]
    #
    #     for c in contacts:
    #         rows.append([
    #             c.get("id", ""),
    #             c.get("name", "") or "",
    #             c.get("email", "") or "",
    #             c.get("phone", "") or "",
    #             c.get("pan_no", "") or "",
    #         ])
    #
    #     # ── Step 4: Clear sheet and write ────────────────────
    #     sheet.clear()
    #     sheet.update("A1", rows)
    #
    #     _logger.info("✅ %d contacts written to Google Sheet", len(contacts))