from odoo import models, fields, api, _
import base64
# from xlrd import open_workbook
import io
from openpyxl import load_workbook
from odoo.exceptions import UserError, ValidationError


class ProductTemplate(models.Model):
    _inherit = 'product.template'

    mrp_ok = fields.Boolean(string='Mrp OK')
    enable_inspection = fields.Boolean('Enable Inspection')
    default_inspection_template = fields.Many2one('inspection.template', string='Inspection Template')


class Product(models.Model):
    _inherit = 'product.product'

    def get_available_quantity(self, location_id, lot_id=None):
        """
        Get the available quantity of the product in a specific location.
        :param location_id: ID of the location to check.
        :return: Available quantity of the product in the specified location.
        """
        domain = [('product_id', '=', self.id),('location_id', '=', location_id.id)]
        if lot_id:
            domain.append(('lot_id', '=', lot_id.id))
        return sum(self.env['stock.quant'].search(domain).mapped('quantity'))


class ProductionBreakdown(models.Model):
    _name = 'production.breakdown'
    _description = "Production Breakdown"

    time = fields.Datetime(default=fields.Datetime.now)
    start_time = fields.Datetime(default=fields.Datetime.now)
    remarks = fields.Text(compute='_compute_production_reason')
    reason = fields.Text()
    type = fields.Selection([
        ('RESTART', 'RESTART'),
        ('HOLD', 'HOLD')
    ])
    code = fields.Selection([('restart','RE'),('hold','H')])
    root_cause = fields.Char()
    action_taken = fields.Char()

    @api.depends('type', 'reason')
    def _compute_production_reason(self):
        for rec in self:
            if rec.reason:
                rec.remarks = f"{rec.type} - {rec.reason}"
            else:
                rec.remarks = ''

    def action_done_breakdown(self):
        return True


class StockPickingType(models.Model):
    _inherit = 'stock.picking.type'

    code = fields.Selection(selection_add=[
        ('mrp_operation', 'Manufacturing')
    ], ondelete={'mrp_operation': 'cascade'})


class ProductSerialNumber(models.Model):
    _name = 'product.serial.number'
    _description = 'Product Serial Number'

    name = fields.Char('Lot/Serial Number', required=True)
    product_id = fields.Many2one('product.product', required=True)
    product_uom_id = fields.Many2one(
        'uom.uom', 'Unit of Measure',
        domain="[('relative_uom_id', '=', product_uom_category_id)]")
    product_uom_category_id = fields.Many2one(related='product_id.uom_id.relative_uom_id')
    product_qty = fields.Float(default=1.0)
    lot_id = fields.Many2one('stock.lot', compute='_compute_lot_id', store=True)
    is_available = fields.Boolean('Is Available', compute='_compute_lot_id')
    tray_id = fields.Many2one('product.tray')

    # @api.constrains('name')
    # def _check_special_characters(self):
    #     for record in self:
    #         if any(char in "!@#%^&*()[]{};:,.<>?/\|`~-_+=" for char in record.name):
    #             raise ValidationError("Special characters are not allowed in Serial Number field.")

    @api.depends('name', 'product_id')
    def _compute_lot_id(self):
        for rec in self:
            rec.lot_id = False
            rec.is_available = False
            lot_id = self.env['stock.lot'].search([('product_id', '=', rec.product_id.id), ('name', '=', rec.name)])
            if lot_id:
                rec.lot_id = lot_id.id
                rec.is_available = True

    # @api.depends('name', 'product_id')
    # def _compute_lot_id(self):
    #     if not self:
    #         return
    #
    #     product_ids = tuple(self.mapped('product_id').ids)
    #     lot_names = tuple(self.mapped('name'))
    #
    #     if not product_ids or not lot_names:
    #         return
    #
    #     query = """
    #             SELECT id, name FROM stock_production_lot
    #             WHERE product_id IN %s AND name IN %s
    #         """
    #
    #     self.env.cr.execute(query, (product_ids, lot_names))
    #     lot_records = self.env.cr.fetchall()
    #
    #     lot_map = {lot[1]: lot[0] for lot in lot_records}
    #
    #     for rec in self:
    #         lot_id = lot_map.get(rec.name)
    #         rec.lot_id = lot_id or False
    #         rec.is_available = bool(lot_id)

    # def clean_unused_serial_number(self):
    #     numbers = self.env['product.serial.number'].search([])
    #

class MaterialLine(models.Model):
    _name = 'material.line'
    _description = 'Material Line'

    product_id = fields.Many2one('product.product', required=True)
    product_uom_category_id = fields.Many2one(related='product_id.uom_id.relative_uom_id')
    product_uom_id = fields.Many2one(
        'uom.uom', 'Unit of Measure',
        domain="[('relative_uom_id', '=', product_uom_category_id)]")
    name = fields.Char(string="Description")
    product_qty = fields.Float(default=1.0, string="Required Qty")
    component_ids = fields.One2many('manufacturing.component', 'material_id', string="Components Lines")
    has_tracking = fields.Selection(related='product_id.tracking', readonly=True)
    active_model = fields.Char("Active Model")
    active_id = fields.Char("Active ID")
    qty_done = fields.Float("Qty Done", compute='_compute_qty_done')
    location_src_id = fields.Many2one('stock.location')
    location_dest_id = fields.Many2one('stock.location')
    out_file_name = fields.Char("Serial File")
    out_file = fields.Binary("Serials")
    production_plan_id = fields.Many2one('production.plan', string="Production Plan")

    def compute_mode(self):
        for rec in self:
            rec.active_model = ''
            rec.active_id = ''
            # if rec.injection_id:



    @api.depends('component_ids')
    def _compute_qty_done(self):
        for rec in self:
            rec.qty_done = 0
            if sum(rec.component_ids.mapped('available_qty')) > rec.product_qty:
                rec.qty_done = rec.product_qty
            elif sum(rec.component_ids.mapped('available_qty')) <= rec.product_qty:
                rec.qty_done = sum(rec.component_ids.mapped('available_qty'))

    def action_show_details(self):
        if self.manufacturing_process_id.operation_id.allow_lot_create and not self.manufacturing_process_id.lot_ids:
            raise UserError("This is lot enabled product.Please upload the Lot/Serial Number in the Serial Number tab.")
        if self.manufacturing_process_id.operation_id.allow_lot_create and self.manufacturing_process_id.lot_ids:
            for lot in self.manufacturing_process_id.lot_ids:
               if not lot.lot_id and not lot.is_available:
                   raise UserError("This is lot enabled product. Please Create a Lot")
        view_id = self.env.ref('fnet_mrp.view_material_line_form', False)
        return {
            'name': _('Import Material'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'material.line',
            'views': [(view_id.id, 'form')],
            'view_id': view_id.id,
            'target': 'new',
            'res_id': self.id,
            'context': {},
        }

    def action_download_sample(self):
        return {
            "type": "ir.actions.act_url",
            "url": '/fnet_mrp/static/Sample_Serials.xlsx',
            "target": "new",
        }

    def action_upload_serial(self):
        print("\n========== action_upload_serial START ==========")
        print("Product:", self.product_id.display_name)
        print("Tracking:", self.product_id.tracking)
        print("Qty Done:", self.qty_done)
        print("Product Qty:", self.product_qty)

        if self.qty_done >= self.product_qty:
            print("ERROR: Already reserved")
            raise UserError(_("Already reserved."))

        record = self.manufacturing_process_id
        print("Manufacturing Process:", record.id if record else False)

        child_records = []
        serials = []

        if self.product_id.tracking == 'none':
            print("Tracking = none")

            child_records.append((0, 0, {
                'product_id': self.product_id.id,
                'product_uom_id': self.product_uom_id.id,
                'product_qty': self.product_qty - sum(self.component_ids.mapped('product_qty')),
                'check_available': True,
                'material_id': self.id,
                'location_src_id': self.location_src_id.id,
                'location_dest_id': self.location_dest_id.id,
            }))

            print("Added Non-Tracked Product Child Record")

        else:
            print("Tracking Enabled")

            if not self.out_file and not self.production_plan_id:
                print("ERROR: No upload file and no production plan selected")
                raise UserError(
                    _("This is lot enabled product. Please upload the serial number updated file.")
                )

            if self.out_file:
                print("Processing Uploaded Excel File")

                file_data = base64.b64decode(self.out_file)
                wb = load_workbook(filename=io.BytesIO(file_data))
                sheet = wb.active

                print("Sheet Name:", sheet.title)
                print("Total Rows:", sheet.max_row)

                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    row_values = list(row)

                    print(f"\nRow {row_idx}: {row_values}")

                    if not row_values[0]:
                        print("Skipped Empty Row")
                        continue

                    serial_no = str(row_values[0])

                    print("Searching Lot:", serial_no)

                    lot_id = self.env['stock.lot'].search(
                        [('name', '=', serial_no), ('product_id', '=', self.product_id.id)]
                    )

                    print("Lot Found:", lot_id.ids)

                    if not lot_id:
                        print("ERROR: Lot Not Found ->", serial_no)
                        raise UserError(_("Serial number %s is not available." % serial_no))

                    available_qty = self.product_id.get_available_quantity(
                        self.location_src_id,
                        lot_id
                    )

                    print("Available Qty:", available_qty)

                    if not available_qty:
                        print("ERROR: Lot Exists but Quantity Not Available")
                        raise UserError(
                            _("Lot %s is not available in the location %s." %
                              (serial_no, self.location_src_id.name))
                        )

                    print("Adding Child Record For:", serial_no)

                    child_records.append((0, 0, {
                        'product_id': self.product_id.id,
                        'product_uom_id': self.product_uom_id.id,
                        'name': self.product_id.name,
                        'lot_id': lot_id.id,
                        'product_qty': 1,
                        'check_available': True,
                        'material_id': self.id,
                        'location_src_id': self.location_src_id.id,
                        'location_dest_id': self.location_dest_id.id,
                    }))

                    if self.product_id.id == record.product_id.id:
                        print("Adding Serial To Manufacturing Process:", serial_no)

                        serials.append((0, 0, {
                            'product_id': self.product_id.id,
                            'name': lot_id.name,
                            'product_uom_id': self.product_uom_id.id,
                            'lot_id': lot_id.id,
                        }))
            elif record.before_manufacturing_process_id.is_split_process:
                print("Loading  Split Process")
                lot_ids = record.lot_ids

                print("Lots Found Count:", len(lot_ids))
                print("Lots Found:", lot_ids.ids)

                if not lot_ids:
                    print("ERROR: No Lots Found")
                    raise UserError(
                        _("No lots available for the plan %s at location %s" %
                          (self.production_plan_id.name,
                           self.location_src_id.name))
                    )

                for lot in lot_ids:
                    print("Processing Lot:", lot.id, lot.name)

                    child_records.append((0, 0, {
                        'product_id': self.product_id.id,
                        'name': self.product_id.name,
                        'product_uom_id': self.product_uom_id.id,
                        'lot_id': lot.lot_id.id,  # <-- FIXED: use the actual stock.lot id
                        'product_qty': 1,
                        'check_available': True,
                        'material_id': self.id,
                        'location_src_id': self.location_src_id.id,
                        'location_dest_id': self.location_dest_id.id,
                    }))

                    if self.product_id.id == record.product_id.id:
                        print("Adding Serial:", lot.name)

                        serials.append((0, 0, {
                            'product_id': self.product_id.id,
                            'name': lot.name,
                            'product_uom_id': self.product_uom_id.id,
                        }))

            else:
                print("Loading Lots From Production Plan")

                print("Production Plan ID:", self.production_plan_id.id)
                print("Production Plan Name:", self.production_plan_id.name)
                print("Location:", self.location_src_id.display_name)

                lot_ids = self.env['stock.lot'].search(
                    [
                        ('product_id', '=', self.product_id.id),
                        ('production_plan_id', '=', self.production_plan_id.id),
                        ('final_location_id', '=', self.location_src_id.id)
                    ]
                )

                print("Lots Found Count:", len(lot_ids))
                print("Lots Found:", lot_ids.ids)

                if not lot_ids:
                    print("ERROR: No Lots Found")
                    raise UserError(
                        _("No lots available for the plan %s at location %s" %
                          (self.production_plan_id.name,
                           self.location_src_id.name))
                    )

                for lot in lot_ids:
                    print("Processing Lot:", lot.id, lot.name)

                    child_records.append((0, 0, {
                        'product_id': self.product_id.id,
                        'name': self.product_id.name,
                        'product_uom_id': self.product_uom_id.id,
                        'lot_id': lot.id,
                        'product_qty': 1,
                        'check_available': True,
                        'material_id': self.id,
                        'location_src_id': self.location_src_id.id,
                        'location_dest_id': self.location_dest_id.id,
                    }))

                    if self.product_id.id == record.product_id.id:
                        print("Adding Serial:", lot.name)

                        serials.append((0, 0, {
                            'product_id': self.product_id.id,
                            'name': lot.name,
                            'product_uom_id': self.product_uom_id.id,
                        }))

        print("\n========== FINAL RESULT ==========")
        print("Child Records Count:", len(child_records))
        print("Serial Records Count:", len(serials))

        if record:
            print("Writing Component Records To Manufacturing Process:", record.id)

            existing_component_ids = record.component_ids.ids
            keep_commands = [(4, cid) for cid in existing_component_ids]
            record.component_ids = keep_commands + child_records

            print("Current Component Count After Write:", len(record.component_ids))

            # Uncomment if needed
            # record.lot_ids = serials
            # print("Lot IDs Written:", len(serials))

        print("========== action_upload_serial END ==========\n")


class ManufacturingComponents(models.Model):
    _name = 'manufacturing.component'
    _description = 'Manufacturing Component'

    material_id = fields.Many2one('material.line', string="Component")
    product_id = fields.Many2one('product.product', required=True)
    product_uom_id = fields.Many2one(
        'uom.uom', 'Unit of Measure',
        domain="[('relative_uom_id', '=', product_uom_category_id)]")
    product_uom_category_id = fields.Many2one(related='product_id.uom_id.relative_uom_id')
    name = fields.Char(string="Description")
    product_qty = fields.Float(default=1.0)
    company_id = fields.Many2one(
        'res.company', 'Company', index=True,
        default=lambda self: self.env.company)
    location_src_id = fields.Many2one('stock.location')
    location_dest_id = fields.Many2one('stock.location')
    has_tracking = fields.Selection(related='product_id.tracking', readonly=True)
    lot_ids = fields.Many2many('stock.lot', string='Lot/Serial')
    lot_id = fields.Many2one('stock.lot', domain="[('product_id', '=', product_id), ('company_id', '=', company_id), ('final_location_id', '=', location_src_id)]", check_company=True)
    lot_number = fields.Char()
    consumed = fields.Float()
    available_qty = fields.Float(store=True)
    output_same = fields.Boolean()
    check_available = fields.Boolean(copy=False)
    stage_id = fields.Many2one('mrp.stage')
    estimation_id = fields.Many2one('mrp.estimation')
    plan_qty = fields.Float()
    planned_qty = fields.Float()
    stage_require_qty = fields.Float()
    stage = fields.Char()
    delivery_lead_days = fields.Float()
    remaining_qty = fields.Float(compute='_compute_remaining_qty')
    tray_id = fields.Many2one('product.tray')
    serial_number_id = fields.Many2one('product.serial.number')
    state = fields.Selection([
        ('draft', 'Draft'),
        ('confirmed', 'Confirmed'),
        ('progress', 'In Progress'),
        ('hold', 'Hold'),
        ('done', 'Done'),
        ('close', 'Closed'),
        ('cancel', 'Cancelled')], string='State',
        copy=False, index=True, default='draft',
        store=True)
    stock_availability = fields.Selection([
        ('ok', 'OK'),
        ('not_ok', 'Not Ok'),
        ('not_checked', 'Not Checked')
    ], compute='_compute_stock_available')
    estimation_stage_id = fields.Many2one('estimation.stage.line')

    @api.onchange('lot_id')
    def _onchange_lot_id(self):
        if self.lot_id:
            self.lot_number = self.lot_id.name

    @api.depends('product_qty', 'available_qty')
    def _compute_remaining_qty(self):
        for rec in self:
            rec.remaining_qty = 0
            if rec.product_qty < rec.available_qty:
                rec.remaining_qty = 0
            if rec.product_qty > rec.available_qty:
                rec.remaining_qty = rec.product_qty - rec.available_qty

    @api.onchange('product_id')
    def onchange_product_id_uom_update(self):
        if self.product_id:
            self.product_uom_id = self.product_id.uom_id.id

    @api.depends('state', 'available_qty', 'planned_qty', 'delivery_lead_days', 'check_available')
    def _compute_stock_available(self):
        for rec in self:
            rec.stock_availability = 'not_checked'
            if rec.state == 'draft' and rec.check_available != False:
                if rec.available_qty >= rec.product_qty:
                    rec.stock_availability = 'ok'
                else:
                    rec.stock_availability = 'not_ok'
                    if rec.product_id.variant_seller_ids:
                        rec.delivery_lead_days = sum(rec.product_id.variant_seller_ids.mapped('delay'))
            else:
                rec.stock_availability = 'not_checked'

    @api.onchange('product_id')
    def _onchange_product_id(self):
        if self.product_id:
            self.name = self.product_id.name

    def remove_manufacturing_component_line(self):
        for rec in self:
            if rec.serial_number_id and rec.tray_id:
                rec.serial_number_id.tray_id = rec.tray_id.id
            rec.unlink()

    def action_import_input(self):
        return {
            'name': _('Import Material'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'input.import.wizard',
            'target': 'new',
            'context': {
                'default_product_id': self.product_id.id,
                'product_uom_id': self.product_uom_id.id
            },
        }

    def unlink(self):
        for rec in self:
            if rec.state in ['close', 'done']:
                raise UserError(_("Cannot removed"))
        return super(ManufacturingComponents, self).unlink()


class ManufacturingFinishedLines(models.Model):
    _name = 'mrp.finished.line'
    _description = 'Manufacturing Finished Lines'

    product_id = fields.Many2one('product.product', required=True)
    lot_id = fields.Many2one('stock.lot')
    product_uom_id = fields.Many2one(
        'uom.uom', 'Unit of Measure',
        domain="[('relative_uom_id', '=', product_uom_category_id)]")
    product_uom_category_id = fields.Many2one(related='product_id.uom_id.relative_uom_id')
    product_qty = fields.Float(default=1.0)
    tray_id = fields.Many2one('product.tray')
    # qr_code_printing_id = fields.Many2one('qr.code.printing', string='QR Code Printing')
    # cell_drying_id = fields.Many2one('cell.drying', string="Manufacturing")

    manufacturing_process_id = fields.Many2one('manufacturing.process', string="Manufacturing Process")


class ProductTry(models.Model):
    _name = 'product.tray'
    _description = "Product Tray"
    _rec_name = 'number'
    _order = "id desc"

    product_id = fields.Many2one('product.product')
    number = fields.Char()
    lot_ids = fields.One2many('product.serial.number', 'tray_id')
    tray_type_id = fields.Many2one('tray.type')
    max_count = fields.Integer(related='tray_type_id.no_of_product_occupied', store=True)

    @api.onchange('lot_ids')
    def _onchange_of_lot_ids(self):
        if self.lot_ids:
            total_line = len(self.lot_ids)
            if total_line > self.max_count:
                raise UserError(_('Maximum capacity of the tray has been reached.'))

    def open_serial_numbers(self):
        return {
            'name': _('Serial Numbers'),
            'type': 'ir.actions.act_window',
            'view_mode': 'list,form',
            'res_model': 'product.serial.number',
            'domain': [('tray_id', '=', self.id)]
        }


class TrayType(models.Model):
    _name = 'tray.type'
    _description = 'Tray Type'

    name = fields.Char()
    no_of_product_occupied = fields.Integer()


class AmbientCondition(models.Model):
    _name = 'ambient.condition'
    _description = 'Ambient Condition'
    _order = "id desc"

    temperature = fields.Float()
    humidity = fields.Float()
    date = fields.Datetime()


class EquipmentDetails(models.Model):
    _name = 'equipment.details'
    _description = 'Equipment Details'
    _order = "id desc"

    accumulate_meters_slitted = fields.Char()
    lower_knife_speed_ratio = fields.Char()
    slitting_speed = fields.Char()
    upper_knife_speed_ratio = fields.Char()
    equipment_status = fields.Char()
    cutter_number = fields.Char()
    date = fields.Datetime()


class StockMoveLine(models.Model):
    _inherit = 'stock.move.line'

    tray_id = fields.Many2one('product.tray')


class StockProductionLot(models.Model):
    _inherit = 'stock.lot'

    product_tracking = fields.Selection(related='product_id.tracking')
    final_location_id = fields.Many2one('stock.location', compute='_compute_lot_current_location', store=True)
    location_id = fields.Many2one('stock.location')

    @api.onchange('final_location_id')
    def _onchange_of_final_location_id(self):
        if self.final_location_id:
            self.location_id = self.final_location_id.id

    @api.depends('quant_ids', 'product_tracking', 'product_id')
    def _compute_lot_current_location(self):
        for rec in self:
            if rec.quant_ids:
                if rec.product_tracking in ['lot', 'serial']:
                    quantity_ids = rec.quant_ids.filtered(lambda x: x.available_quantity == 1)
                    if quantity_ids:
                        limit = 1
                        for quantity in quantity_ids:
                            if limit == 1:
                                rec.final_location_id = quantity.location_id.id
                    else:
                        rec.final_location_id =False
                else:
                    rec.final_location_id = False
            else:
                rec.final_location_id = False


class TypeChangeDuration(models.Model):
    _name = 'type.change.duration'
    _description = 'Type Change Duration'
    _rec_name = 'type'

    type = fields.Selection([
        ('anode_slitting', 'Anode Slitting '),
        ('cathode_slitting', 'Cathode Slitting '),
        ('anode_drying', 'Anode Drying'),
        ('cathode_drying', 'Cathode Drying'),
        ('diaphragm_drying', 'Dia Drying'),
        ('anode_electrode_making', 'Anode Electrode Making'),
        ('cathode_electrode_making', 'Cathode Electrode Making'),
        ('winding', 'Winding'),
        ('hot_press_jelly', 'Hot Press Jelly'),
        ('assembly', 'Assembly'),
        ('qr_code_print', 'QR Code Printing'),
        ('cell_drying', 'Cell Drying'),
        ('injection', 'Injection'),
        ('high_temperature', 'High Temperature'),
        ('cell_clamp_baking', 'Cell Clamp Baking'),
        ('aged_formation_cell', 'Aged Formation Cell'),
        ('degas', 'Degas'),
        ('dsf', 'Double side Folding'),
        ('pad_printing', 'Pad Printing'),
        ('capacity_test', 'Capacity Test'),
        ('voltage_test', 'Voltage Test'),
        ('aged_formation_cell_2', 'Aged Formation Cell 2'),
        ('voltage_test_2', 'Voltage Test 2'),
        ('packing', 'Packing')
    ])
    duration = fields.Float()


class InspectionTeam(models.Model):
    _name = 'inspection.team'

    name = fields.Char("Team Name")
    user_ids = fields.Many2many('res.users', 'rel_inspection_user', string="Responsible")
    approve_user_ids = fields.Many2many('res.users', 'rel_inspection_approve', string="Approval Responsible")


class Company(models.Model):
    _inherit = 'res.company'

    ccr_doc_no = fields.Char(string='CCR Doc No')
    ccr_rev_no = fields.Char(string='CCR Rev No')
    ccr_rev_date = fields.Date(string='CCR Rev Date')

    pqc_doc_no=fields.Char(string='PQC Doc No')
    pqc_rev_no = fields.Char(string='PQC Rev No')
    pqc_rev_date = fields.Date(string='PQC Rev Date')

    in_ins_doc_no = fields.Char(string='Inward Doc No')
    in_ins_rev_no = fields.Char(string='Inward Rev No')
    in_ins_rev_date = fields.Date(string='Inward Rev Date')

    out_ins_doc_no = fields.Char(string='Outward Doc No')
    out_ins_rev_no = fields.Char(string='Outward Rev No')
    out_ins_rev_date = fields.Date(string='Outward Rev Date')

    audit_ins_doc_no = fields.Char(string='Outward Doc No')
    audit_ins_rev_no = fields.Char(string='Outward Rev No')
    audit_ins_rev_date = fields.Date(string='Outward Rev Date')

    dn_doc_no = fields.Char(string='DN Doc No')
    dn_rev_no = fields.Char(string='DN Rev No')
    dn_rev_date = fields.Date(string='DN Rev Date')


class ResConfigSettings(models.TransientModel):
    _inherit = 'res.config.settings'

    ccr_doc_no = fields.Char(string='CCR Doc No',related='company_id.ccr_doc_no', readonly=False)
    ccr_rev_no = fields.Char(string='CCR Rev No',related='company_id.ccr_rev_no', readonly=False)
    ccr_rev_date = fields.Date(string='CCR Rev Date',related='company_id.ccr_rev_date', readonly=False)

    pqc_doc_no = fields.Char(string='PQC Doc No',related='company_id.pqc_doc_no', readonly=False)
    pqc_rev_no = fields.Char(string='PQC Rev No',related='company_id.pqc_rev_no', readonly=False)
    pqc_rev_date = fields.Date(string='PQC Rev Date',related='company_id.pqc_rev_date', readonly=False)

    in_ins_doc_no = fields.Char(string='Inward Doc No', related='company_id.in_ins_doc_no', readonly=False)
    in_ins_rev_no = fields.Char(string='Inward Rev No', related='company_id.in_ins_rev_no', readonly=False)
    in_ins_rev_date = fields.Date(string='Inward Rev Date', related='company_id.in_ins_rev_date', readonly=False)

    out_ins_doc_no = fields.Char(string='Outward Doc No', related='company_id.out_ins_doc_no', readonly=False)
    out_ins_rev_no = fields.Char(string='Outward Rev No', related='company_id.out_ins_rev_no', readonly=False)
    out_ins_rev_date = fields.Date(string='Outward Rev Date', related='company_id.out_ins_rev_date', readonly=False)

    audit_ins_doc_no = fields.Char(string='Outward Doc No', related='company_id.audit_ins_doc_no', readonly=False)
    audit_ins_rev_no = fields.Char(string='Outward Rev No', related='company_id.audit_ins_rev_no', readonly=False)
    audit_ins_rev_date = fields.Date(string='Outward Rev Date', related='company_id.audit_ins_rev_date', readonly=False)

    dn_doc_no = fields.Char(string='DN Doc No', related='company_id.dn_doc_no', readonly=False)
    dn_rev_no = fields.Char(string='DN Rev No', related='company_id.dn_rev_no', readonly=False)
    dn_rev_date = fields.Date(string='DN Rev Date', related='company_id.dn_rev_date', readonly=False)


class ManufacturingBatch(models.Model):
    _name = 'manufacturing.batch'
    _description = 'Manufacturing Batch'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'id desc'

    name = fields.Char()