from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import timedelta
import base64
import logging
# from xlrd import open_workbook
import io
from openpyxl import load_workbook
_logger = logging.getLogger(__name__)

class MachineData(models.Model):
    _inherit = 'machine.parameter'

    packing_id = fields.Many2one('package.move')
    packing_value_condition = fields.Selection([
        ('good', 'Good'),
        ('fail', 'Fail')])


class Utilities(models.Model):
    _inherit = 'utility.parameter'

    packing_id = fields.Many2one('package.move')
    packing_utility_condition = fields.Selection([
        ('good', 'Good'),
        ('fail', 'Fail')
    ], compute='_compute_packing_utility')

    def _compute_packing_utility(self):
        for rec in self:
            rec.packing_utility_condition = 'fail'
            if rec.packing_id:
                if rec.packing_id.operation_id.min_humidity <= rec.humidity <= rec.packing_id.operation_id.max_humidity and rec.packing_id.operation_id.min_temperature <= rec.temperature <= rec.packing_id.operation_id.max_temperature:
                    rec.packing_utility_condition = 'good'
                else:
                    rec.packing_utility_condition = 'fail'


class QualityDetails(models.Model):
    _inherit = 'quality.parameter'

    packing_id = fields.Many2one('package.move')


class StockProductionLOt(models.Model):
    _inherit = 'stock.lot'

    packing_id = fields.Many2one('package.move')


class StockMove(models.Model):
    _inherit = 'stock.move'

    packing_id = fields.Many2one('package.move')
    operation_id = fields.Many2one('manufacturing.operation')


class ProductSerialNumber(models.Model):
    _inherit = 'product.serial.number'

    packing_id = fields.Many2one('package.move')


class StockMoveLIne(models.Model):
    _inherit = 'stock.move.line'

    packing_id = fields.Many2one('package.move')
    out_packing_id = fields.Many2one('package.move')


class ManufacturingComponents(models.Model):
    _inherit = 'manufacturing.component'

    packing_id = fields.Many2one('package.move')

    @api.onchange('product_id')
    def _onchange_packing_product_id(self):
        if self.packing_id:
            self.location_src_id = self.packing_id.location_src_id.id
            self.location_dest_id = self.packing_id.location_dest_id.id


class MaterialLine(models.Model):
    _inherit = 'material.line'

    packing_id = fields.Many2one('package.move')


class PadPrinting(models.Model):
    _name = 'package.move'
    _description = 'Package'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = "id desc"

    out_file_name = fields.Char("Serial File")
    out_file = fields.Binary("Serials")
    done_lot = fields.Boolean('')
    serial_created = fields.Boolean('')

    @api.model
    def default_get(self, fields):
        defaults = super(PadPrinting, self).default_get(fields)
        production_location = self.env['stock.location'].search([('usage', '=', 'production')], limit=1)
        defaults['production_location_id'] = production_location.id
        return defaults

    def action_upload_serial(self):
        if not self.out_file:
            raise UserError("Please upload the serial number updated file.")
        file_data = base64.b64decode(self.out_file)
        wb = load_workbook(filename=io.BytesIO(file_data))
        sheet = wb.active
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            row_values = list(row)
            self.env['product.serial.number'].create({
                'product_id': self.product_id.id,
                'name': str(row_values[0]),
                'product_uom_id': self.product_uom_id.id,
                'capacity_id': self.capacity_id.id
            })
        self.serial_created = True
        self.lot_creation()

    def action_download_sample(self):
        return {
            "type": "ir.actions.act_url",
            "url": '/fnet_mrp/static/Sample Serials.xlsx',
            "target": "new",
        }

    def _get_default_product_uom_id(self):
        return self.env['uom.uom'].search([], limit=1, order='id').id

    type = fields.Selection([
        ('anode_slitting', 'Anode Slitting '),
        ('cathode_slitting', 'Cathode Slitting '),
        ('anode_drying', 'Anode Drying'),
        ('cathode_drying', 'Cathode Drying'),
        ('diaphragm_drying', 'Diaphragm Drying'),
        ('anode_electrode_making', 'Anode Electrode Making'),
        ('cathode_electrode_making', 'Cathode Electrode Making'),
        ('winding', 'Winding'),
        ('hot_press_jelly', 'Hot Press Jelly'),
        ('assembly', 'Assembly'),
        ('qr_code_print', 'QR Code Printing'),
        ('cell_drying', 'Cell Drying'),
        ('injection', 'Injection'),
        ('high_temperature', 'High Temperature'),
        ('cell_baking_formation', 'Cell Formation'),
        ('aged_formation_cell', 'Aged Formation Cell'),
        ('degas', 'Degas'),
        ('dsf', 'Double side Folding'),
        ('pad_printing', 'Pad Printing'),
        ('capacity_test', 'Capacity Test'),
        ('voltage_test', 'Capacity Test'),
        ('packing', 'Packing'),
    ], default='packing')
    company_id = fields.Many2one(
        'res.company', 'Company', index=True,
        default=lambda self: self.env.company)
    location_src_id = fields.Many2one('stock.location')
    location_dest_id = fields.Many2one('stock.location')
    production_location_id = fields.Many2one('stock.location')
    lot_ids = fields.One2many('product.serial.number', 'packing_id', string='Serial Number')

    name = fields.Char(
        'Reference', copy=False, readonly=True, default=lambda x: _('New'))
    product_id = fields.Many2one(
        'product.product', 'Product',
        domain="""[
               ('type', 'in', ['combo', 'consu']),
               '|',
                   ('company_id', '=', False),
                   ('company_id', '=', company_id)
           ]
           """,
        readonly=True, check_company=True,)
    bom_id = fields.Many2one(
        'manufacturing.bom', 'Bill of Material')
    component_ids = fields.One2many(
        'manufacturing.component', 'packing_id', 'Components',
        copy=False)
    state = fields.Selection([
        ('draft', 'Draft'),
        ('confirmed', 'Confirmed'),
        ('progress', 'In Progress'),
        ('hold', 'Hold'),
        ('done', 'Done'),
        ('close', 'Closed'),
        ('cancel', 'Cancelled')], string='State',
        copy=False, index=True, default='draft',
        store=True, tracking=True,
        help=" * Draft: The MO is not confirmed yet.\n"
             " * Confirmed: The MO is confirmed, the stock rules and the reordering of the components are trigerred.\n"
             " * In Progress: The production has started (on the MO or on the WO).\n"
             " * To Close: The production is done, the MO has to be closed.\n"
             " * Done: The MO is closed, the stock moves are posted. \n"
             " * Cancelled: The MO has been cancelled, can't be confirmed anymore.")
    product_tracking = fields.Selection(related='product_id.tracking')
    product_tmpl_id = fields.Many2one('product.template', 'Product Template', related='product_id.product_tmpl_id')
    product_qty = fields.Float(
        'Quantity To Produce',
        default=1.0, digits='Product Unit of Measure',
        readonly=True, required=True, tracking=True,
       )
    product_uom_id = fields.Many2one(
        'uom.uom', 'Product Unit of Measure', default=_get_default_product_uom_id,
        readonly=True, required=True,
       domain="[('relative_uom_id', '=', product_uom_category_id)]")

    qty_producing = fields.Float(string="Quantity Producing", digits='Product Unit of Measure', copy=False)
    product_uom_category_id = fields.Many2one(related='product_id.uom_id.relative_uom_id')
    product_uom_qty = fields.Float(string='Total Quantity', store=True)
    user_id = fields.Many2one(
        'res.users', 'Responsible', default=lambda self: self.env.user,
        states={'done': [('readonly', True)], 'cancel': [('readonly', True)]})
    qty_produced = fields.Float(string="Quantity Produced")
    operation_id = fields.Many2one('manufacturing.operation')
    production_plan_id = fields.Many2one('production.plan')
    product_model_id = fields.Many2one('product.model')

    machine_data_ids = fields.One2many(
        'machine.parameter', 'packing_id', 'Machine Data',
        copy=False, states={'done': [('readonly', True)], 'cancel': [('readonly', True)]})
    utility_ids = fields.One2many(
        'utility.parameter', 'packing_id', 'Utility Parameter',
        copy=False, states={'done': [('readonly', True)], 'cancel': [('readonly', True)]})
    quality_ids = fields.One2many(
        'quality.parameter', 'packing_id', 'Quality',
        copy=False, states={'done': [('readonly', True)], 'cancel': [('readonly', True)]})
    machine_id = fields.Many2one('manufacturing.machine')
    start_time = fields.Datetime()
    end_time = fields.Datetime()
    product_tray_id = fields.Many2one('product.tray')
    breakdown_ids = fields.One2many('production.breakdown', 'packing_id')
    voltage_test_id = fields.Many2one('voltage.test')
    voltage_test_count = fields.Integer(compute='_compute_voltage_test_count')
    finished_move_ids = fields.One2many('stock.move.line', 'out_packing_id')
    input_material_lines = fields.One2many('material.line', 'packing_id')
    remaining_hours = fields.Float("Remaining Time", compute='_get_remaining_time')
    maintenance_user = fields.Char()
    capacity_id = fields.Many2one('capacity.test')
    choose_stage = fields.Selection(related='production_plan_id.choose_stage', string='Choose Stage')

    @api.model
    def default_get(self, fields):
        defaults = super(PadPrinting, self).default_get(fields)
        company_id = self.env.user.company_id or self.env['res.company'].search([])[0]
        defaults['maintenance_user'] = company_id.maintenance_user_id
        return defaults

    def request_breakdown(self):
        for rec in self:
            mail_values = {
                'email_from': self.user_id.login,
                'email_to': rec.maintenance_user,
                'body_html': """Dear Sir <br/> 
                                    The machine has been broken down. Would you please repair it<br/>
                                                Thanks & Regards.""",
                'subject': 'Machine Break down : %s' % (self.machine_id.name),
            }
            self.env['mail.mail'].sudo().create(mail_values).send()

    def _get_remaining_time(self):
        for rec in self:
            hr_remain = 0
            if rec.start_time and rec.operation_id.process_duration:
                end = rec.start_time + timedelta(hours=rec.operation_id.process_duration)
                hr_remain = ((end - fields.Datetime.now()).seconds / 3600)
            rec.remaining_hours = hr_remain

    def action_remove_line(self):
        self.component_ids.unlink()

    def _compute_voltage_test_count(self):
        for rec in self:
            rec.voltage_test_count = self.env['voltage.test'].search_count([('id', '=', rec.voltage_test_id.id)])

    def show_voltage_test_record(self):
        records = self.env['voltage.test'].search([('id', '=', self.voltage_test_id.id)])
        if len(records) == 1:
            return {
                'name': _('Voltage Test 2'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'voltage.test',
                'res_id': records.id,  # Assuming records is a single record
            }
        else:
            return {
                'name': _('Voltage Test'),
                'type': 'ir.actions.act_window',
                'view_mode': 'list,form',
                'res_model': 'Voltage.test',
                'domain': [('id', '=', self.voltage_test_id.id)]
            }

    @api.onchange('state')
    def _onchange_of_state(self):
        if self.state and self.component_ids:
            for line in self.component_ids:
                line.state = self.state

    @api.onchange('product_tray_id')
    def _onchange_product_tray_id(self):
        if self.product_tray_id:
            child_records = []
            lines = self.env['product.serial.number'].search([('tray_id', '=', self.product_tray_id.id)])
            for line in lines:
                child_records.append((0, 0, {
                    'product_id': self.product_tray_id.product_id.id,
                    'name': self.product_tray_id.product_id.name,
                    'lot_number': line.name,
                    'location_src_id': self.location_src_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'product_qty': 1.0,
                    'output_same': True,
                    'tray_id': self.product_tray_id.id,
                }))
            self.component_ids = child_records

    @api.onchange('production_plan_id')
    def _onchange_of_product_plan_id(self):
        if self.production_plan_id:
            production_operation = self.env['production.operation'].search(
                [('operation_type', '=', self.type), ('production_plan_id', '=', self.production_plan_id.id)], limit=1)
            self.operation_id = production_operation.operation_id.id
            self.product_model_id = self.production_plan_id.model_id.id
            self.bom_id = self.operation_id.bom_id.id
            if not self.bom_id:
                self.component_ids = False
            self._onchange_of_operation()

    @api.onchange('type')
    def _onchange_of_operation_type(self):
        if self.type:
            machine = self.env['manufacturing.machine'].search([('type', '=', self.type)], limit=1)
            if machine:
                self.machine_id = machine.id

    @api.onchange('operation_id')
    def _onchange_of_operation(self):
        if self.operation_id:
            production_location = self.env['stock.location'].search([('usage', '=', 'production')], limit=1)
            self.product_id = self.operation_id.product_id.id
            self.location_src_id = self.operation_id.location_src_id.id
            self.location_dest_id = self.operation_id.location_dest_id.id
            self.production_location_id = production_location.id
            self.bom_id = self.operation_id.bom_id.id
        else:
            self.bom_id = False
            self.component_ids =False
        if self.component_ids:
            for line in self.component_ids:
                line.location_src_id = self.location_src_id.id
                line.location_dest_id = self.production_location_id.id
        self.bom_id = self.operation_id.bom_id.id
        self._onchange_bom_id()
        self._onchange_of_operation_type()

    @api.onchange('product_id')
    def _onchange_product_id(self):
        if self._origin:
            if self.product_id:
                self.bom_id = False
                self.component_ids = False

    @api.onchange('bom_id', 'product_qty')
    def _onchange_bom_id(self):
        self.input_material_lines = False
        self.component_ids = False
        if self.bom_id:
            child_records = []
            for line in self.bom_id.bom_line_ids:
                child_records.append((0, 0, {
                    'product_id': line.product_id.id,
                    'product_uom_category_id': line.product_uom_category_id.id,
                    'product_uom_id': line.product_uom_id.id,
                    'location_src_id': self.location_src_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'name': line.product_id.name,
                    'product_qty': (line.product_qty / self.bom_id.product_qty) * self.product_qty
                }))
            self.input_material_lines = child_records

    @api.onchange('location_src_id')
    def onchange_location_src_id(self):
        for rec in self.input_material_lines:
            rec.location_src_id = self.location_src_id.id

    def get_serial_numbers(self):
        available_inputs = self.component_ids.filtered(lambda x: x.product_id.id == self.product_id.id and x.lot_id)
        child_records = []
        for line in available_inputs:
            child_records.append((0, 0, {
                'product_id': line.product_id.id,
                'product_uom_id': line.product_uom_id.id,
                'name': line.lot_id.name,
            }))
        self.lot_ids = child_records

    def clear_serial_numbers(self):
        self.lot_ids = False

    def check_available_stock(self):
        for line in self.component_ids:
            available_qty = line.product_id.get_available_quantity(line.location_src_id, line.lot_id)
            qty = line.product_uom_id._compute_quantity(line.product_qty, line.product_id.uom_id)
            line.available_qty = available_qty
            if qty > available_qty:
                raise UserError(
                    _("Required quantity is not available in the stock for %s. Please check on %s" % (
                        line.product_id.name, self.location_src_id.name)))
            line.check_available = True
        for line in self.input_material_lines:
            if line.product_qty != line.qty_done:
                raise UserError(_("Required materials and reserved materials are not same."))

    def action_start(self):
        for rec in self:
            # if not self.done_lot:
            #     raise UserError("Fill Upload Serial Number")
            rec.state = 'progress'
            self.start_time = fields.Datetime.now()

    def lot_creation(self):
        if self.product_id:
            self.product_id.write({
                'tracking': 'serial',
            })
        for lot in self.lot_ids:
            if lot.lot_id:
                available_qty = self.env['stock.quant']._get_available_quantity(self.product_id,
                                                                                self.production_location_id,
                                                                                lot_id=lot.lot_id)
                if available_qty == 0:
                    _logger.error(_("Lot not available in %s" % self.production_location_id.name))
                    # raise UserError(_("Lot not available in %s" % self.production_location_id.name))
        if self.product_id.tracking != 'none':
            for serial in self.capacity_id.lot_ids:
                if not serial.lot_id:
                    if not self.operation_id.allow_lot_create:
                        _logger.error(
                            _("The lot %s is not available and The current operation is not allowed to create new lot number.\n Please enable lot creation or check the inventory." % serial.name))
                        # raise UserError(_("The lot %s is not available and The current operation is not allowed to create new lot number.\n Please enable lot creation or check the inventory." %serial.name))
                    lot_id = self.env['stock.lot'].create({
                        'name': serial.name,
                        'ref': serial.name,
                        'product_id': self.product_id.id,
                        'company_id': self.env.company.id,
                        'production_plan_id': self.production_plan_id.id,
                        'final_location_id' : self.location_src_id.id
                    })
                    update_stock = self.env['stock.quant'].sudo().create({
                        'product_id': self.product_id.id,
                        'location_id': self.location_src_id.id,
                        'lot_id': lot_id.id,
                        'inventory_quantity': 1.0,
                    })
                    update_stock.action_apply_inventory()
        self.done_lot = True

    def action_done_production(self):
        # if not self.lot_ids:
        #     raise UserError("Please Create Lot")
        for rec in self:
            rec.check_available_stock()
            stock_moves = []
            for line in rec.component_ids:
                # stock_move = self.env['stock.move'].create({
                #     'name': rec.name,
                #     'product_id': line.product_id.id,
                #     'product_uom': line.product_uom_id.id,
                #     'quantity_done': line.product_qty,
                #     'location_id': line.location_src_id.id,
                #     'location_dest_id': rec.production_location_id.id,
                #     'state': 'draft',
                #     'packing_id': rec.id,
                # })
                # stock_move._action_assign()
                # existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
                # if not existing_move_lines:
                #     self.env['stock.move.line'].create({
                #         'product_id': line.product_id.id,
                #         'product_uom_id': line.product_id.uom_id.id,
                #         'location_id': line.location_src_id.id,
                #         'location_dest_id': rec.production_location_id.id,
                #         'company_id': rec.company_id.id,
                #         'packing_id': rec.id,
                #         'lot_id': line.lot_id.id,
                #         'move_id': stock_move.id,
                #         'state': 'draft',
                #     })
                # for move_line in existing_move_lines:
                #     move_line.write({
                #         'packing_id': rec.id,
                #         'lot_id': line.lot_id.id or False
                #     })
                # stock_move._action_confirm()  # Confirm the move
                # stock_move._action_done()
                stock_move_vals = {
                    'inventory_name': rec.name,
                    'product_id': line.product_id.id,
                    'product_uom': line.product_uom_id.id,
                    'product_uom_qty': line.product_qty,
                    'location_id': line.location_src_id.id,
                    'location_dest_id': rec.production_location_id.id,
                    'packing_id': rec.id,
                    'move_line_ids': [(0, 0, {
                        'product_id': line.product_id.id,
                        'product_uom_id': line.product_id.uom_id.id,
                        'quantity': line.product_qty,
                        'location_id': line.location_src_id.id,
                        'location_dest_id': rec.production_location_id.id,
                        'company_id': rec.company_id.id,
                        'packing_id': rec.id,
                        'lot_id': line.lot_id.id if line.lot_id else False,
                    })],
                }
                stock_moves.append(stock_move_vals)
            # Batch create stock moves
            created_stock_moves = self.env['stock.move'].create(stock_moves)
            # Confirm and assign
            created_stock_moves._action_confirm()
            created_stock_moves._action_assign()
            # Write lot + qty + custom fields onto assigned move lines
            for move, line in zip(created_stock_moves, rec.component_ids):
                if move.move_line_ids:
                    move.move_line_ids.write({
                        'quantity': line.product_qty,
                        'lot_id': line.lot_id.id if line.lot_id else False,
                        'packing_id': rec.id,
                    })
                else:
                    self.env['stock.move.line'].create({
                        'move_id': move.id,
                        'product_id': line.product_id.id,
                        'product_uom_id': line.product_uom_id.id,
                        'quantity': line.product_qty,
                        'location_id': line.location_src_id.id,
                        'location_dest_id': rec.production_location_id.id,
                        'company_id': rec.company_id.id,
                        'packing_id': rec.id,
                        'lot_id': line.lot_id.id if line.lot_id else False,
                    })
                move.move_line_ids.picked = True
            created_stock_moves._action_done()
        if self.product_id.tracking != 'none' and len(self.lot_ids) != self.product_qty:
            raise UserError(_("Quantity to produce is %s but you have entered %s serial numbers." % (
                self.product_qty, len(self.lot_ids))))
        for lot in self.lot_ids:
            if lot.lot_id:
                available_qty = self.env['stock.quant']._get_available_quantity(self.product_id,
                                                                                self.production_location_id,
                                                                                lot_id=lot.lot_id)
                if available_qty == 0:
                    raise UserError(_("Lot not available in %s" % self.production_location_id.name))
        if self.product_id.tracking == 'none':
            stock_move = self.env['stock.move'].create({
                'inventory_name': self.name,
                'product_id': self.product_id.id,
                'product_uom': self.product_uom_id.id,
                'product_uom_qty': self.product_qty,
                'location_id': self.production_location_id.id,
                'location_dest_id': self.location_dest_id.id,
                'packing_id': self.id,
            })
            stock_move._action_confirm()
            stock_move._action_assign()
            existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
            if not existing_move_lines:
                self.env['stock.move.line'].create({
                    'move_id': stock_move.id,
                    'product_id': self.product_id.id,
                    'product_uom_id': self.product_uom_id.id,
                    'quantity': self.product_qty,
                    'location_id': self.production_location_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'company_id': self.company_id.id,
                    'packing_id': self.id,
                    'out_packing_id': self.id,
                })
                existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
            existing_move_lines.write({
                'packing_id': self.id,
                'out_packing_id': self.id,
            })
            stock_move.move_line_ids.picked = True
        else:
            for serial in self.lot_ids:
                lot_id = serial.lot_id
                if not serial.lot_id:
                    if not self.operation_id.allow_lot_create:
                        raise UserError(
                            _("The lot %s is not available and The current operation is not allowed to create new lot number.\n Please enable lot creation or check the inventory."))
                    lot_id = self.env['stock.lot'].create({
                        'name': serial.name,
                        'ref': serial.name,
                        'product_id': self.product_id.id,
                        'company_id': self.env.company.id,
                    })
                    update_stock = self.env['stock.quant'].sudo().create({
                        'product_id': self.product_id.id,
                        'location_id': self.production_location_id.id,
                        'lot_id': lot_id.id,
                        'inventory_quantity': 1.0,
                    })
                    update_stock.action_apply_inventory()
                stock_move = self.env['stock.move'].create({
                    'inventory_name': self.name,
                    'product_id': self.product_id.id,
                    'product_uom': self.product_uom_id.id,
                    'product_uom_qty': 1,
                    'location_id': self.production_location_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'packing_id': self.id,
                })
                stock_move._action_confirm()
                stock_move._action_assign()
                existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
                if not existing_move_lines:
                    self.env['stock.move.line'].create({
                        'move_id': stock_move.id,
                        'product_id': self.product_id.id,
                        'product_uom_id': self.product_uom_id.id,
                        'quantity': 1,
                        'location_id': self.production_location_id.id,
                        'location_dest_id': self.location_dest_id.id,
                        'company_id': self.company_id.id,
                        'lot_id': lot_id.id,
                        'packing_id': self.id,
                        'out_packing_id': self.id,
                    })
                    existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
                existing_move_lines.write({
                    'packing_id': self.id,
                    'out_packing_id': self.id,
                    'lot_id': lot_id.id
                })
                for move_line in existing_move_lines:
                    move_line.write({
                        'packing_id': self.id,
                        'out_packing_id': self.id,
                        'lot_id': lot_id.id
                    })
                stock_move.move_line_ids.picked = True
        self.state = 'done'
        self.end_time = fields.Datetime.now()

    def action_close(self):
        # search_out_tray = self.finished_move_ids.filtered(lambda x: x.lot_id and not x.out_tray_id)
        # if search_out_tray:
        #     raise UserError(_('Please fill the out tray number'))
        for move_line in self.finished_move_ids:
            child_records = []
            if move_line.out_tray_id and move_line.lot_id:
                child_records.append((0, 0, {
                    'product_id': move_line.product_id.id,
                    'product_uom_id': move_line.product_uom_id.id,
                    'lot_id': move_line.lot_id.id,
                    'name': move_line.lot_id.name,
                    }))
            move_line.out_tray_id.lot_ids = child_records
        stock_moves = self.finished_move_ids.mapped('move_id')
        stock_moves._action_done()
        # for stock_move in stock_moves:
        #     # stock_move._action_confirm()  # Confirm the move
        #     stock_move._action_done()
        self.state = 'close'

    def action_view_lots(self):
        lot_ids = []
        production_plans = self.env['production.plan'].search([('state', '=', 'in_production')])
        lot_ids += self.env['stock.lot'].search([('production_plan_id', 'in', production_plans.ids), ('final_location_id', '=', self.location_src_id.id)]).mapped('id')
        return {
            'name': _('Available Lots'),
            'type': 'ir.actions.act_window',
            'view_mode': 'list,form',
            'res_model': 'stock.lot',
            'domain': [('id', 'in', lot_ids), ('location_id.usage', '!=', 'inventory')],
            'context': {'group_by': ['production_plan_id', 'final_location_id']},
        }

    def action_view_product_move(self):
        return {
            'res_model': 'stock.move.line',
            'type': 'ir.actions.act_window',
            'name': _("Stock Move"),
            'domain': [('packing_id', '=', self.id)],
            'view_mode': 'list,form',
        }

    def open_cell_formation_processing(self):
        return {
            'name': _('Cell Formation'),
            'type': 'ir.actions.act_window',
            'view_mode': 'list,form',
            'res_model': 'cell.clamp.baking',
            'domain': [('id', '=', self.formation_id.id)]
        }

    @api.model_create_multi
    def create(self, vals_list):
        seq = self.env['ir.sequence']
        for vals in vals_list:
            if not vals.get('name') or vals.get('name') == _('New'):
                vals['name'] = seq.next_by_code('packing.move') or _('New')
        return super().create(vals_list)

    # def create_scrap_product(self):
    #     for rec in self:
    #         scrap_location_id = self.env['stock.location'].search([('usage', '=', 'inventory')], limit=1)
    #         return {
    #             'name': _('Scrap Production?'),
    #             'type': 'ir.actions.act_window',
    #             'view_mode': 'form',
    #             'res_model': 'manufacturing.scrap',
    #             'target': 'new',
    #             'context': {
    #                 'default_product_id': rec.product_id.id,
    #                 'default_location_src_id': rec.location_dest_id.id,
    #                 'default_scrap_location_id': scrap_location_id.id,
    #                 'default_packing_id': rec.id,
    #             },
    #         }

    def create_scrap_product(self):
        for rec in self:
            location_reject_id = self.env['stock.location'].search([('usage', '=', 'inventory')], limit=1)
            if rec.operation_id.location_reject_id:
                location_reject_id = rec.operation_id.location_reject_id
            lot_ids = self.finished_move_ids.filtered(lambda x: x.lot_id).mapped('lot_id')
            return {
                'name': _('Scrap Production?'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'manufacturing.scrap',
                'target': 'new',
                'context': {
                    'default_mrp_lot_ids': lot_ids.ids,
                    'default_name': rec.name,
                    'default_location_reject_id': location_reject_id.id,
                    'default_production_plan_id': rec.production_plan_id.id or False,
                    'default_operation_id': rec.operation_id.id,
                    'default_product_id': rec.product_id.id,
                    'default_product_uom_id': rec.product_uom_id.id,
                    'default_location_src_id': rec.location_dest_id.id,
                    'default_packing_id': rec.id,
                },
            }

    def action_break_production(self):
        for rec in self:
            user = self.env.user
            employee = self.env['hr.employee'].search([('user_id', '=', user.id)], limit=1)
            view = self.env.ref('fnet_mrp.production_breakdown_form_view2')
            production = self.env['maintenance.equipment.category'].search([])[0]
            category_id = production.id if production else False
            self.env['maintenance.request'].create({
                'name': self.machine_id.name,
                'employee_id': employee.id,
                'request_date': fields.Date.today(),
                'maintenance_type': 'breakdown',
                'schedule_date': fields.Date.today(),
                'category_id': category_id,
            })
            return {
                'name': _('Production  Breakdown'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'views': [(view.id, 'form')],
                'view_id': view.id,
                'res_model': 'production.breakdown',
                'target': 'new',
                'context': {
                    'default_packing_id': rec.id,
                    'default_type': 'HOLD',
                    'default_code': 'hold',
                },
            }

    def action_restart_production(self):
        for rec in self:
            if any(not break_down.end_time and break_down.type == 'HOLD' for break_down in rec.breakdown_ids):
                raise UserError("Kindly enter the endtime")
            view = self.env.ref('fnet_mrp.production_breakdown_form_view2')
            return {
                'name': _('Production  Breakdown'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'views': [(view.id, 'form')],
                'view_id': view.id,
                'res_model': 'production.breakdown',
                'target': 'new',
                'context': {
                    'default_packing_id': rec.id,
                    'default_type': 'RESTART',
                    'default_code': 'restart',
                },
            }

    def action_update_lot_product(self):
        for rec in self:
            return {
                'name': _('Update Lot'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'tray.product.lot',
                'target': 'new',
                'context': {
                    'default_packing_id': rec.id,
                    'default_dest_location_id': rec.location_dest_id.id,
                    'default_location_src_id':rec.location_src_id.id
                },
            }

class ManufacturingScrap(models.Model):
    _inherit = 'manufacturing.scrap'
    _description = "Manufacturing Scrp"

    packing_id = fields.Many2one('package.move')

    def create_scrap_product(self):
        res = super(ManufacturingScrap, self).create_scrap_product()
        picking_type_id = self.env['stock.picking.type'].search([('code', '=', 'mrp_operation')], limit=1)
        for rec in self:
            if rec.packing_id:
                available_quant = self.env['stock.quant'].sudo().search([
                    ('product_id', '=', rec.product_id.id),
                    ('location_id', '=', rec.location_src_id.id), ('lot_id', '=', rec.lot_id.id)
                ], limit=1)

                if not available_quant or sum(available_quant.mapped('quantity')) < rec.scrap_qty:
                    raise UserError(f"Insufficient stock for product: {rec.product_id.name} "
                                    f"at location: {rec.location_src_id.complete_name}.\n"
                                    f"Available Quantity: {sum(available_quant.mapped('quantity')) if sum(available_quant.mapped('quantity')) else 0}\n"
                                    f"Required Quantity: {rec.scrap_qty}")
                picking = self.env['stock.picking'].create({
                    'picking_type_id': picking_type_id.id,
                    'location_id': rec.location_src_id.id,
                    'location_dest_id': rec.scrap_location_id.id,
                    'company_id': rec.env.company.id,
                    'move_type': 'direct',
                    'origin': rec.packing_id.name,
                })
                stock_move = self.env['stock.move'].create({
                    'inventory_name': rec.packing_id.name,
                    'product_id': rec.product_id.id,
                    'product_uom_qty': rec.scrap_qty,
                    'product_uom': rec.product_uom_id.id,
                    'location_id': rec.location_src_id.id,
                    'location_dest_id': rec.scrap_location_id.id,
                    'picking_id': picking.id,
                    'company_id': rec.env.company.id,
                    'packing_id': rec.packing_id.id,
                    # 'production_id': self.id,
                    'state': 'draft',  # Set the state to 'waiting' to be done.
                    'origin': rec.packing_id.name,
                    'manufacturing_scrap_id': rec.id,
                })

                stock_move_line = self.env['stock.move.line'].create({
                    'product_id': rec.product_id.id,
                    # 'product_uom_qty': 1.0,
                    'quantity': rec.scrap_qty,
                    'product_uom_id': rec.product_uom_id.id,
                    'location_id': rec.location_src_id.id,
                    'location_dest_id': rec.scrap_location_id.id,
                    'company_id': rec.company_id.id,
                    'packing_id': rec.packing_id.id,
                    'lot_id': rec.lot_id.id,
                    'picking_id': picking.id,
                    'move_id': stock_move.id,
                    'manufacturing_scrap_id': rec.id,
                    # 'production_id': self.id,
                    'state': 'draft',
                })
                picking.action_confirm()
                picking.button_validate()
            rec.state = 'done'
        return res


class ProductionBreakdown(models.Model):
    _inherit = 'production.breakdown'

    packing_id = fields.Many2one('package.move')
    end_time = fields.Datetime(string = 'End Time')


    def action_done_breakdown(self):
        for rec in self:
            if rec.packing_id:
                if rec.type == 'HOLD':
                    rec.packing_id.state = 'hold'
                else:
                    rec.packing_id.state = 'progress'

        return super(ProductionBreakdown, self).action_done_breakdown()