from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import timedelta, datetime
import base64
# from xlrd import open_workbook
import io
from openpyxl import load_workbook

class StockMoveLIne(models.Model):
    _inherit = 'stock.move.line'

    powerbank_id = fields.Many2one('mrp.powerbank')
    out_powerbank_id = fields.Many2one('mrp.powerbank')


class StockMove(models.Model):
    _inherit = 'stock.move'

    powerbank_id = fields.Many2one('mrp.powerbank')


class ManufacturingComponents(models.Model):
    _inherit = 'manufacturing.component'

    powerbank_id = fields.Many2one('mrp.powerbank')


class MaterialLine(models.Model):
    _inherit = 'material.line'

    powerbank_id = fields.Many2one('mrp.powerbank')


class ProductSerialNumber(models.Model):
    _inherit = 'product.serial.number'

    powerbank_id = fields.Many2one('mrp.powerbank')


class PowerBank(models.Model):
    _name = 'mrp.powerbank'
    _description = 'Power Bank'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = "id desc"

    @api.model
    def default_get(self, fields):
        defaults = super(PowerBank, self).default_get(fields)
        defaults['production_location_id'] = self.env['stock.location'].search([('usage', '=', 'production')], limit=1).id
        return defaults

    def _get_default_product_uom_id(self):
        return self.env['uom.uom'].search([], limit=1, order='id').id

    name = fields.Char(
        'Reference', copy=False, readonly=True, default=lambda x: _('New'))
    product_id = fields.Many2one(
        'product.product', 'Product',
        domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]", check_company=True)
    bom_id = fields.Many2one(
        'manufacturing.bom', 'Bill of Material')
    type = fields.Selection([('powerbank', 'Power Bank')], default='powerbank', string="Type")
    company_id = fields.Many2one('res.company', 'Company', index=True, default=lambda self: self.env.company)
    location_src_id = fields.Many2one('stock.location')
    location_dest_id = fields.Many2one('stock.location')
    production_location_id = fields.Many2one('stock.location')
    state = fields.Selection([
        ('draft', 'Draft'),
        ('progress', 'In Progress'),
        ('done', 'Done'),
        ('close', 'Closed'),
        ('cancel', 'Cancelled')], string='State',
        copy=False, index=True, default='draft',
        store=True, tracking=True)
    product_tracking = fields.Selection(related='product_id.tracking')
    product_tmpl_id = fields.Many2one('product.template', 'Product Template', related='product_id.product_tmpl_id')
    product_qty = fields.Float(
        'Quantity To Produce',
        default=1.0, digits='Product Unit of Measure',
        required=True, tracking=True,
        )
    product_uom_id = fields.Many2one(
        'uom.uom', 'Product Unit of Measure', default=_get_default_product_uom_id,
        readonly=True, required=True,
         domain="[('relative_uom_id', '=', product_uom_category_id)]")
    product_uom_category_id = fields.Many2one(related='product_id.uom_id.relative_uom_id')
    operation_id = fields.Many2one('manufacturing.operation')
    input_material_lines = fields.One2many('material.line', 'powerbank_id')
    finished_move_ids = fields.One2many('stock.move.line', 'powerbank_id')
    component_ids = fields.One2many('manufacturing.component', 'powerbank_id', string='Components', copy=False)
    lot_ids = fields.One2many('product.serial.number', 'powerbank_id', string='Serial Number')
    date_start = fields.Datetime()
    date_end = fields.Datetime()
    remaining_hours = fields.Float("Remaining Time", compute='_get_remaining_time')
    expected_date_end = fields.Datetime("Expected End", compute='compute_end_date')
    out_file_name = fields.Char("Serial File")
    out_file = fields.Binary("Serials")

    @api.onchange('location_src_id')
    def onchange_location_src(self):
        for rec in self.component_ids:
            rec.location_src_id = self.location_src_id
        for rec in self.input_material_lines:
            rec.location_src_id = self.location_src_id.id

    def action_view_quality(self):
        return {
            'res_model': 'mrp.quality',
            'type': 'ir.actions.act_window',
            'name': _("Quality Check"),
            'domain': [('powerbank_id', '=', self.id)],
            'view_mode': 'list,form',
        }

    def _get_remaining_time(self):
        for rec in self:
            hr_remain = 0
            if rec.date_start and rec.operation_id.process_duration:
                end = rec.date_start + timedelta(hours=rec.operation_id.process_duration)
                hr_remain = ((end - fields.Datetime.now()).seconds / 3600)
            rec.remaining_hours = hr_remain

    def action_remove_line(self):
        self.component_ids.unlink()

    def action_download_sample(self):
        return {
            "type": "ir.actions.act_url",
            "url": '/fnet_mrp/static/Sample Serials.xlsx',
            "target": "new",
        }

    def action_upload_serial(self):
        if not self.out_file:
            raise UserError("Please upload the serial number updated file.")
        file_data = base64.b64decode(self.out_file)
        wb = load_workbook(filename=io.BytesIO(file_data))
        sheet = wb.active
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            row_values = list(row)
            self.env['product.serial.number'].create({
                'product_id': self.product_id.id,
                'name': str(row_values[0]),
                'product_uom_id': self.product_uom_id.id,
                'powerbank_id': self.id
            })

    @api.depends('date_start', 'operation_id')
    def compute_end_date(self):
        for rec in self:
            if rec.date_start and rec.operation_id:
                rec.expected_date_end = rec.date_start + timedelta(hours=rec.operation_id.process_duration)
            else:
                rec.expected_date_end = False

    @api.constrains('date_end')
    def duration_constrain(self):
        for rec in self:
            if rec.date_end and rec.expected_date_end and rec.date_end < rec.expected_date_end:
                time = timedelta(hours=rec.operation_id.process_duration)
                dt = datetime(2000, 1, 1) + time
                raise UserError("Minimum duration to stop process is %s hours." % dt.strftime("%H:%M"))

    @api.onchange('operation_id')
    def _onchange_of_operation(self):
        if self.operation_id:
            self.product_id = self.operation_id.product_id.id
            self.location_src_id = self.operation_id.location_src_id.id
            self.location_dest_id = self.operation_id.location_dest_id.id
            self.bom_id = self.operation_id.bom_id.id

    @api.onchange('product_id')
    def _onchange_product_id(self):
        if self._origin:
            if self.product_id:
                self.bom_id = False
                self.component_ids = False

    @api.onchange('bom_id', 'product_qty')
    def _onchange_bom_id(self):
        self.input_material_lines = False
        self.component_ids = False
        if self.bom_id:
            child_records = []
            for line in self.bom_id.bom_line_ids:
                child_records.append((0, 0, {
                    'product_id': line.product_id.id,
                    'product_uom_category_id': line.product_uom_category_id.id,
                    'product_uom_id': line.product_uom_id.id,
                    'location_src_id': self.location_src_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'name': line.product_id.name,
                    'product_qty': (line.product_qty / self.bom_id.product_qty) * self.product_qty
                }))
            self.input_material_lines = child_records

    def get_serial_numbers(self):
        available_inputs = self.component_ids.filtered(lambda x: x.product_id.id == self.product_id.id and x.lot_id)
        child_records = []
        for line in available_inputs:
            child_records.append((0, 0, {
                'product_id': line.product_id.id,
                'product_uom_id': line.product_uom_id.id,
                'name': line.lot_id.name,
            }))
        self.lot_ids = child_records

    def clear_serial_numbers(self):
        self.lot_ids = False

    def check_available_stock(self):
        for line in self.input_material_lines:
            if line.product_qty != line.qty_done:
                raise UserError(_("Required materials and reserved materials are not same."))
        for line in self.component_ids:
            available_qty = line.product_id.get_available_quantity(line.location_src_id, line.lot_id)
            qty = line.product_uom_id._compute_quantity(line.product_qty, line.product_id.uom_id)
            line.available_qty = available_qty
            if qty > available_qty:
                raise UserError(
                    _("Required quantity is not available in the stock for %s. Please check on %s" % (
                    line.product_id.name, self.location_src_id.name)))
            line.check_available = True

    def action_start(self):
        for rec in self:
            rec.state = 'progress'
            self.date_start = fields.Datetime.now()

    def action_done_production(self):
        for rec in self:
            rec.check_available_stock()
            stock_moves = []
            for line in rec.component_ids:
                stock_move_vals = {
                    'inventory_name': rec.name,
                    'product_id': line.product_id.id,
                    'product_uom': line.product_uom_id.id,
                    'product_uom_qty': line.product_qty,
                    'location_id': line.location_src_id.id,
                    'location_dest_id': rec.production_location_id.id,
                    'powerbank_id': rec.id,
                    'move_line_ids': [(0, 0, {
                        'product_id': line.product_id.id,
                        'product_uom_id': line.product_id.uom_id.id,
                        'quantity': line.product_qty,
                        'location_id': line.location_src_id.id,
                        'location_dest_id': rec.production_location_id.id,
                        'company_id': rec.company_id.id,
                        'powerbank_id': rec.id,
                        'lot_id': line.lot_id.id if line.lot_id else False,
                    })],
                }
                stock_moves.append(stock_move_vals)
            # Batch create stock moves
            created_stock_moves = self.env['stock.move'].create(stock_moves)
            # Confirm and assign
            created_stock_moves._action_confirm()
            created_stock_moves._action_assign()
            # Write lot + qty + custom fields onto assigned move lines
            for move, line in zip(created_stock_moves, rec.component_ids):
                if move.move_line_ids:
                    move.move_line_ids.write({
                        'quantity': line.product_qty,
                        'lot_id': line.lot_id.id if line.lot_id else False,
                        'powerbank_id': rec.id,
                    })
                else:
                    self.env['stock.move.line'].create({
                        'move_id': move.id,
                        'product_id': line.product_id.id,
                        'product_uom_id': line.product_uom_id.id,
                        'quantity': line.product_qty,
                        'location_id': line.location_src_id.id,
                        'location_dest_id': rec.production_location_id.id,
                        'company_id': rec.company_id.id,
                        'powerbank_id': rec.id,
                        'lot_id': line.lot_id.id if line.lot_id else False,
                    })
                move.move_line_ids.picked = True
            created_stock_moves._action_done()
        if self.product_id.tracking != 'none' and len(self.lot_ids) != self.product_qty:
            raise UserError(_("Quantity to produce is %s but you have entered %s serial numbers." % (
                self.product_qty, len(self.lot_ids))))
        for lot in self.lot_ids:
            if lot.lot_id:
                available_qty = self.env['stock.quant']._get_available_quantity(self.product_id,
                                                                                self.production_location_id,
                                                                                lot_id=lot.lot_id)
                if available_qty == 0:
                    raise UserError(_("Lot not available in %s" % self.production_location_id.name))
        if self.product_id.tracking == 'none':
            stock_move = self.env['stock.move'].create({
                'inventory_name': self.name,
                'product_id': self.product_id.id,
                'product_uom': self.product_uom_id.id,
                'product_uom_qty': self.product_qty,
                'location_id': self.production_location_id.id,
                'location_dest_id': self.location_dest_id.id,
                'powerbank_id': self.id,
            })
            stock_move._action_confirm()
            stock_move._action_assign()
            existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
            if not existing_move_lines:
                self.env['stock.move.line'].create({
                    'move_id': stock_move.id,
                    'product_id': self.product_id.id,
                    'product_uom_id': self.product_uom_id.id,
                    'quantity': self.product_qty,
                    'location_id': self.production_location_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'company_id': self.company_id.id,
                    'powerbank_id': self.id,
                    'out_powerbank_id': self.id,
                })
                existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
            existing_move_lines.write({
                'powerbank_id': self.id,
                'out_powerbank_id': self.id,
            })
            stock_move.move_line_ids.picked = True
        else:
            for serial in self.lot_ids:
                lot_id = serial.lot_id
                if not serial.lot_id:
                    if not self.operation_id.allow_lot_create:
                        raise UserError(
                            _("The lot %s is not available and The current operation is not allowed to create new lot number.\n Please enable lot creation or check the inventory." % serial.name))
                    lot_id = self.env['stock.lot'].create({
                        'name': serial.name,
                        'ref': serial.name,
                        'product_id': self.product_id.id,
                        'company_id': self.env.company.id,
                    })
                    update_stock = self.env['stock.quant'].sudo().create({
                        'product_id': self.product_id.id,
                        'location_id': self.production_location_id.id,
                        'lot_id': lot_id.id,
                        'inventory_quantity': 1.0,
                    })
                    update_stock.action_apply_inventory()
                stock_move = self.env['stock.move'].create({
                    'inventory_name': self.name,
                    'product_id': self.product_id.id,
                    'product_uom': self.product_uom_id.id,
                    'product_uom_qty': 1,
                    'location_id': self.production_location_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'powerbank_id': self.id,
                })
                stock_move._action_confirm()
                stock_move._action_assign()
                existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
                if not existing_move_lines:
                    self.env['stock.move.line'].create({
                        'move_id': stock_move.id,
                        'product_id': self.product_id.id,
                        'product_uom_id': self.product_uom_id.id,
                        'quantity': 1,
                        'location_id': self.production_location_id.id,
                        'location_dest_id': self.location_dest_id.id,
                        'company_id': self.company_id.id,
                        'lot_id': lot_id.id,
                        'powerbank_id': self.id,
                        'out_powerbank_id': self.id,
                    })
                    existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
                existing_move_lines.write({
                    'powerbank_id': self.id,
                    'out_powerbank_id': self.id,
                    'lot_id': lot_id.id
                })
                for move_line in existing_move_lines:
                    move_line.write({
                        'powerbank_id': self.id,
                        'out_powerbank_id': self.id,
                        'lot_id': lot_id.id
                    })
                stock_move.move_line_ids.picked = True
        self.state = 'close'
        self.date_end = fields.Datetime.now()

    def action_close(self):
        stock_moves = self.finished_move_ids.mapped('move_id')
        for stock_move in stock_moves:
            stock_move._action_done()
        self.state = 'close'

    def action_view_product_move(self):
        return {
            'res_model': 'stock.move.line',
            'type': 'ir.actions.act_window',
            'name': _("Stock Move"),
            'domain': [('powerbank_id', '=', self.id)],
            'view_mode': 'list,form',
        }


    @api.model_create_multi
    def create(self, vals_list):
        seq = self.env['ir.sequence']
        for vals in vals_list:
            if not vals.get('name') or vals.get('name') == _('New'):
                vals['name'] = seq.next_by_code('mrp.powerbank') or _('New')
        return super().create(vals_list)

    def create_scrap_product(self):
        for rec in self:
            location_reject_id = self.env['stock.location'].search([('usage', '=', 'inventory')], limit=1)
            if rec.operation_id.location_reject_id:
                location_reject_id = rec.operation_id.location_reject_id
            lot_ids = self.finished_move_ids.filtered(lambda x: x.lot_id).mapped('lot_id')
            return {
                'name': _('Scrap Production?'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'manufacturing.scrap',
                'target': 'new',
                'context': {
                    'default_mrp_lot_ids': lot_ids.ids,
                    'default_name': rec.name,
                    'default_location_reject_id': location_reject_id.id,
                    'default_operation_id': rec.operation_id.id,
                    'default_product_id': rec.product_id.id,
                    'default_product_uom_id': rec.product_uom_id.id,
                    'default_location_src_id': rec.location_dest_id.id,
                    'default_powerbank_id': rec.id,
                },
            }