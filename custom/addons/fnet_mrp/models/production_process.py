from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import timedelta, datetime
import base64
# from xlrd import open_workbook
import io
from openpyxl import load_workbook
import logging
import re
from collections import Counter
from odoo import models, fields
_logger = logging.getLogger(__name__)
import json

from markupsafe import Markup
from odoo.exceptions import ValidationError


class ManufacturingStagesRevision(models.Model):
    _name = "manufacturing.stages.revision"
    _description = "Manufacturing Stage Revision"

    stage_id = fields.Many2one("manufacturing.stages", required=True)
    revision_number = fields.Integer()
    revision_type = fields.Selection([
        ('create', 'Created'),
        ('revision', 'Revision')
    ], default='revision')
    revision_date = fields.Datetime(default=fields.Datetime.now)
    user_id = fields.Many2one(
        'res.users',
        default=lambda self: self.env.user
    )
    html = fields.Html(
        string="Revision Details",
        sanitize=False
    )


class ManufacturingStagesRevisionLine(models.Model):
    _name = 'manufacturing.stages.revision.line'
    _description = 'Manufacturing Stage Revision Change'

    revision_id = fields.Many2one(
        'manufacturing.stages.revision', required=True, ondelete='cascade')
    field_label = fields.Char(string='Field')
    old_value = fields.Char()
    new_value = fields.Char()




class ProductionStages(models.Model):
    _name = 'manufacturing.stages'
    _description = 'Production Stages'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'sequence, id'

    sequence = fields.Integer(string='Sequence', default=10)
    name = fields.Char(string='Name', required=True ,tracking=True)
    operation_ids = fields.One2many(
        'manufacturing.operation', 'manufacturing_stages_id', copy=True,domain=[('model_id', '=', False)],)
    state = fields.Selection([
        ('draft', 'Draft'),
        ('in_revision', 'In Revision'),
        ('request', 'Requested'),
        ('approved', 'Approved'),
        ('cancel', 'Cancelled')], string='State',
        copy=False, index=True, default='draft',
        store=True, tracking=True)

    # ---- revision history ----
    revision_ids = fields.One2many(
        'manufacturing.stages.revision', 'stage_id')
    revision_count = fields.Integer(compute='_compute_revision_count')
    last_revision_snapshot = fields.Text(default='{}')
    current_revision_id = fields.Many2one(
        'manufacturing.stages.revision', copy=False)
    # scalar fields on this model to track
    REVISION_TRACKED_FIELDS = ['name']

    # operation fields we never snapshot
    OPERATION_SKIP_FIELDS = {
        'id', 'create_uid', 'create_date', 'write_uid', 'write_date',
        '__last_update', 'display_name',
        'manufacturing_stages_id',
        'bom_ids',
        'message_ids', 'message_follower_ids', 'activity_ids',
        'company_id',
    }

    @api.constrains('operation_ids')
    def constrains_operation_ids(self):
        for rec in self:
            operations = rec.operation_ids.sorted('sequence')
            if not operations:
                continue

            all_src_locations = operations.mapped('location_src_id')
            all_dest_locations = operations.mapped('location_dest_id')

            dest_counts = {}
            for op in operations:
                dest_counts[op.location_dest_id.id] = dest_counts.get(op.location_dest_id.id, 0) + 1

            # Collect all branch process types coming from any split operation
            branch_type_ids = self.env['manufacturing.process.type']
            for op in operations:
                if op.manufacturing_process_type_id.is_split_process:
                    branch_type_ids |= op.manufacturing_process_type_id.process_type1_id
                    branch_type_ids |= op.manufacturing_process_type_id.process_type2_id

            for index, line in enumerate(operations):
                is_first = index == 0
                is_last = index == len(operations) - 1
                is_branch = line.manufacturing_process_type_id in branch_type_ids

                if is_first and line.manufacturing_process_type_id.is_split_process:
                    raise UserError(_(
                        "Operation '%s' is a split process and cannot be the "
                        "first operation in the sequence."
                    ) % line.display_name)

                # if is_first and not line.allow_lot_create:
                #     raise UserError(_(
                #         "Operation '%s' must have 'Allow Lot Creation' enabled "
                #         "since it is the first operation in the sequence."
                #     ) % line.display_name)

                    # >>> PLACE IT HERE <
                    # Only the first operation may create lots
                if not is_first and line.allow_lot_create:
                    raise UserError(_(
                        "Operation '%s' has 'Allow Lot Creation' enabled, but "
                        "only the first operation in the sequence is allowed "
                        "to create lots."
                    ) % line.display_name)

                # Branch operations must come AFTER their split operation.
                # This runs at top level so it actually fires for split ops.
                if line.manufacturing_process_type_id.is_split_process:
                    split_index = index
                    branch_types = (
                            line.manufacturing_process_type_id.process_type1_id
                            | line.manufacturing_process_type_id.process_type2_id
                    )
                    for b_index, b_line in enumerate(operations):
                        if (b_line.manufacturing_process_type_id in branch_types
                                and b_index <= split_index):
                            raise UserError(_(
                                "Operation '%s' is a branch of split process "
                                "'%s' and must be placed after it in the "
                                "sequence."
                            ) % (b_line.display_name, line.display_name))

                # Middle operations (non-split, non-branch) must chain into the stage
                if (not is_first and not is_last
                        and not line.manufacturing_process_type_id.is_split_process
                        and not is_branch):
                    if line.location_src_id not in all_dest_locations:
                        raise UserError(_(
                            "Source location of operation '%s' must match the "
                            "destination location of some operation in this "
                            "stage."
                        ) % line.display_name)

                if is_last:
                    continue

                # A non-last operation's destination must be picked up by some
                # later operation. If nothing consumes it (e.g. it goes into
                # 'Consumed location'), the stock has nowhere to go -> invalid.
                if line.location_dest_id not in all_src_locations:
                    raise UserError(_(
                        "Operation '%s' sends stock to '%s', but no other "
                        "operation uses '%s' as its source location. Stock "
                        "sent there has nowhere to go."
                    ) % (
                                        line.display_name,
                                        line.location_dest_id.display_name,
                                        line.location_dest_id.display_name,
                                    ))

                # Convergence: multiple operations feed the same destination
                is_convergence = dest_counts.get(line.location_dest_id.id, 0) > 1
                if is_convergence:
                    if line.location_dest_id not in all_src_locations:
                        raise UserError(_(
                            "Destination location of operation '%s' must match "
                            "the source location of some operation in this "
                            "stage."
                        ) % line.display_name)
                    continue

                next_line = operations[index + 1]
                next_is_branch = next_line.manufacturing_process_type_id in branch_type_ids

                # A split operation fans out into its process_type1_id /
                # process_type2_id branches, each with its own src/dest, so the
                # plain "dest must equal next src" rule does not apply when either
                # side of the pair is a split or a branch of a split.
                if (line.manufacturing_process_type_id.is_split_process
                        or is_branch
                        or next_line.manufacturing_process_type_id.is_split_process
                        or next_is_branch):
                    continue

                if line.location_dest_id != next_line.location_src_id:
                    raise UserError(_(
                        "Operation '%s' destination location must match the "
                        "source location of the next operation '%s'."
                    ) % (line.display_name, next_line.display_name))

    def _generate_revision_html(self, old_values, new_values):
        self.ensure_one()

        html = """
        <style>
            table{
                width:100%;
                border-collapse:collapse;
                margin-top:15px;
                font-size:13px;
            }
            th{
                background:#0d6efd;
                color:#FFF;
                padding:8px;
                border:1px solid #DDD;
            }
            td{
                padding:8px;
                border:1px solid #DDD;
            }
            .section{
                background:#F2F2F2;
                font-weight:bold;
            }
        </style>

        <h2>Manufacturing Stage Revision</h2>

        <table>
            <tr>
                <th>Section</th>
                <th>Field</th>
                <th>Old Value</th>
                <th>New Value</th>
            </tr>
        """

        # ---------------------------
        # Manufacturing Stage Changes
        # ---------------------------

        for field in self.REVISION_TRACKED_FIELDS:
            old = old_values.get(field)
            new = new_values.get(field)

            if old != new:
                html += f"""
                <tr>
                    <td>Manufacturing Stage</td>
                    <td>{self._fields[field].string}</td>
                    <td>{self._format_revision_value(field, old)}</td>
                    <td>{self._format_revision_value(field, new)}</td>
                </tr>
                """

        # ---------------------------
        # Operation Changes
        # ---------------------------

        old_ops = old_values.get('_operations', {})
        new_ops = new_values.get('_operations', {})

        # --- NEW or EDITED operations ---
        for op_id, values in new_ops.items():
            op = self.env['manufacturing.operation'].browse(int(op_id)).exists()
            op_label = op.display_name if op else "Operation %s" % op_id

            is_new = op_id not in old_ops
            old_op = old_ops.get(op_id, {})

            if is_new:
                # just the header row — no field-level detail for new operations
                html += f"""
                <tr class="section">
                    <td colspan="4">🆕 New Process Created: {op_label}</td>
                </tr>
                """
            else:
                # existing operation: only changed fields
                changed_rows = ""
                for field_name, new_val in values.items():
                    if field_name == '_name':
                        continue
                    old_val = old_op.get(field_name)
                    if (old_val or False) == (new_val or False):
                        continue
                    changed_rows += f"""
                                    <tr>
                                        <td>{op_label}</td>
                                        <td>{op._fields[field_name].string}</td>
                                        <td>{self._format_op_value(field_name, old_val)}</td>
                                        <td>{self._format_op_value(field_name, new_val)}</td>
                                    </tr>
                                    """
                    if changed_rows:
                        html += f"""
                                        <tr class="section">
                                            <td colspan="4">✏️ Edited Process: {op_label}</td>
                                        </tr>
                                        """
                        html += changed_rows

        # --- DELETED operations ---
        for op_id, old_op in old_ops.items():
            if op_id not in new_ops:
                op_name = old_op.get('_name') or ("Operation %s" % op_id)
                html += f"""
                <tr class="section">
                    <td colspan="4">🗑️ This Process Deleted: {op_name}</td>
                </tr>
                """

        html += "</table>"

        return html


    def _compute_revision_count(self):
        for rec in self:
            rec.revision_count = len(rec.revision_ids)

    # ---- snapshot helpers ----
    def _get_operation_tracked_fields(self):
        Operation = self.env['manufacturing.operation']
        tracked = []
        for fname, field in Operation._fields.items():
            if fname in self.OPERATION_SKIP_FIELDS:
                continue
            if field.related or (field.compute and not field.store):
                continue
            if field.type in ('one2many', 'many2many'):
                continue
            tracked.append(fname)
        return tracked

    def _get_revision_values(self):
        self.ensure_one()
        vals = {}

        for fname in self.REVISION_TRACKED_FIELDS:
            field = self._fields[fname]
            value = self[fname]
            if field.type == 'many2one':
                vals[fname] = value.id or False
            elif field.type in ('one2many', 'many2many'):
                vals[fname] = value.ids
            else:
                vals[fname] = value or False

        op_snapshot = {}
        tracked_fields = self._get_operation_tracked_fields()
        for op in self.operation_ids:
            op_vals = {}
            for f in tracked_fields:
                fld = op._fields[f]
                v = op[f]
                op_vals[f] = v.id if fld.type == 'many2one' else (v or False)
            op_vals['_name'] = op.display_name          # store label for delete rows
            op_snapshot[str(op.id)] = op_vals
        vals['_operations'] = op_snapshot
        return vals

    def _format_revision_value(self, fname, value):
        if not value:
            return ''
        field = self._fields[fname]
        if field.type == 'many2one' and isinstance(value, int):
            rec = self.env[field.comodel_name].browse(value).exists()
            return rec.display_name if rec else str(value)
        return str(value)

    def _format_op_value(self, fname, value):
        if not value:
            return ''
        fld = self.env['manufacturing.operation']._fields[fname]
        if fld.type == 'many2one' and isinstance(value, int):
            rec = self.env[fld.comodel_name].browse(value).exists()
            return rec.display_name if rec else str(value)
        return str(value)



    def action_submit(self):
        for rec in self:
            rec.write({'state': 'request'})
            rec._notify_manufacturing_managers()

    def _notify_manufacturing_managers(self):
        self.ensure_one()
        group = self.env.ref('fnet_mrp.group_manufacturing_manager',
                             raise_if_not_found=False)
        if not group:
            raise ValidationError(_("Manufacturing Manager group is not configured."))

        recipient_emails = group.user_ids.filtered(lambda u: u.email).mapped('email')
        if not recipient_emails:
            raise ValidationError(_("No Manufacturing Manager has an email address configured."))

        requester = self.env.user
        stage_name = self.name or 'N/A'
        op_count = len(self.operation_ids)

        mail_body = f"""
            <table border="0" cellpadding="0" cellspacing="0" width="100%"
                   style="background-color:#ffffff; font-family: Helvetica, Arial, sans-serif; font-size:14px; color:#2d2d2d;">

                <!-- Header Banner -->
                <tr>
                    <td style="background-color:#1e7e34; padding:24px 32px;">
                        <h1 style="margin:0; color:#ffffff; font-size:20px; font-weight:600; letter-spacing:0.5px;">
                            Manufacturing Process Approval
                        </h1>
                    </td>
                </tr>

                <!-- Body -->
                <tr>
                    <td style="padding:32px;">

                        <p style="margin:0 0 16px 0;">Dear Manager,</p>

                        <p style="margin:0 0 16px 0; line-height:1.6;">
                            A manufacturing process has been submitted and is awaiting your approval.
                            Please review the details below and approve.
                        </p>

                        <!-- Info Box -->
                        <table border="0" cellpadding="0" cellspacing="0" width="100%"
                               style="background-color:#eef9f0; border-left:4px solid #1e7e34;
                                      border-radius:4px; margin:24px 0;">
                            <tr>
                                <td style="padding:16px 20px;">
                                    <table border="0" cellpadding="6" cellspacing="0" width="100%">
                                        <tr>
                                            <td style="color:#666666; font-size:13px; width:160px;">Process Name</td>
                                            <td style="color:#2d2d2d; font-weight:600;">{stage_name}</td>
                                        </tr>
                                        <tr>
                                            <td style="color:#666666; font-size:13px;">Number of Operations</td>
                                            <td style="color:#2d2d2d; font-weight:600;">{op_count}</td>
                                        </tr>
                                        <tr>
                                            <td style="color:#666666; font-size:13px;">Submitted By</td>
                                            <td style="color:#2d2d2d; font-weight:600;">{requester.name or ''}</td>
                                        </tr>
                                    </table>
                                </td>
                            </tr>
                        </table>

                        <!-- View Button -->
                        <p style="margin:24px 0;">
                            <a href="{self.get_base_url()}/web#id={self.id}&model=manufacturing.stages&view_type=form"
                               style="display: inline-block; background-color: #1e7e34; color: #ffffff;
                                      text-decoration: none; font-size: 13px; font-weight: 600;
                                      padding: 10px 24px; border-radius: 6px;">
                                Review &amp; Approve →
                            </a>
                        </p>

                        <p style="margin:24px 0 4px 0; line-height:1.6;">Thanks &amp; regards,</p>
                        <p style="margin:0; font-weight:600;">{requester.name or ''}</p>
                    </td>
                </tr>

                <!-- Footer -->
                <tr>
                    <td style="background-color:#f0f0f0; padding:16px 32px; border-top:1px solid #dddddd;">
                        <p style="margin:0; font-size:12px; color:#999999; text-align:center;">
                            This is an automated notification.
                        </p>
                    </td>
                </tr>

            </table>
            """

        mail_values = {
            'subject': f'Manufacturing Process Approval - {stage_name}',
            'email_to': ','.join(recipient_emails),
            'email_from': (requester.email or self.env.company.email or ''),
            'body_html': mail_body,
        }

        mail = self.env['mail.mail'].sudo().create(mail_values)
        mail.sudo().send()

    def check_operation_line(self):
        # self = manufacturing.stages record(s)
        # find every product.model that uses these stages
        models = self.env['product.model'].search([
            ('manufacturing_stages_ids', 'in', self.ids)
        ])
        for model in models:  # loop product.models, NOT self
            existing_lines = {
                line.manufacturing_operation_line_id.id: line
                for line in model.operation_ids
                if line.manufacturing_operation_line_id
            }
            desired_templates = self.env['manufacturing.operation']
            template_to_stage = {}
            for stage in model.manufacturing_stages_ids:
                for template in stage.operation_ids.filtered(
                        lambda o: not o.model_id and not o.bom_id):
                    desired_templates |= template
                    template_to_stage[template.id] = stage.id
            desired_ids = set(desired_templates.ids)
            existing_ids = set(existing_lines.keys())

            # DELETE — template no longer selected
            for tmpl_id in (existing_ids - desired_ids):
                existing_lines[tmpl_id].unlink()

            # DELETE — orphan lines
            model.operation_ids.filtered(
                lambda l: not l.manufacturing_operation_line_id).unlink()

            # CREATE — new templates (new record -> pass model_id + template link)
            for template in desired_templates.filtered(
                    lambda t: t.id in (desired_ids - existing_ids)):
                vals = template.copy_data()[0]
                vals.update({
                    'model_id': model.id,
                    'manufacturing_stages_id': template_to_stage[template.id],
                    'bom_id': False,
                    'manufacturing_operation_line_id': template.id,
                })
                self.env['manufacturing.operation'].create(vals)

            # UPDATE — kept lines (already have model_id + template link -> don't pass)
            for tmpl_id in (desired_ids & existing_ids):
                template = desired_templates.browse(tmpl_id)
                vals = template.copy_data()[0]
                vals.pop('model_id', None)  # already correct, don't touch
                vals.pop('manufacturing_operation_line_id', None)  # already correct, don't touch
                vals.update({
                    'manufacturing_stages_id': template_to_stage[tmpl_id],
                    'bom_id': False,
                })
                existing_lines[tmpl_id].write(vals)

    def action_approve(self):
        for rec in self:
            rec.write({'state': 'approved'})
            rec.check_operation_line()
            rec._notify_creator_approved()

    def _notify_creator_approved(self):
        self.ensure_one()

        creator = self.create_uid  # the user who created the record
        if not creator.email:
            # skip silently rather than blocking approval
            return

        approver = self.env.user
        stage_name = self.name or 'N/A'

        mail_body = f"""
            <table border="0" cellpadding="0" cellspacing="0" width="100%"
                   style="background-color:#ffffff; font-family: Helvetica, Arial, sans-serif; font-size:14px; color:#2d2d2d;">

                <tr>
                    <td style="background-color:#1e7e34; padding:24px 32px;">
                        <h1 style="margin:0; color:#ffffff; font-size:20px; font-weight:600; letter-spacing:0.5px;">
                            Manufacturing Process Approved
                        </h1>
                    </td>
                </tr>

                <tr>
                    <td style="padding:32px;">
                        <p style="margin:0 0 16px 0;">Dear {creator.name or ''},</p>

                        <p style="margin:0 0 16px 0; line-height:1.6;">
                            Your manufacturing process has been reviewed and
                            <strong style="color:#1e7e34;">approved</strong>.
                        </p>

                        <table border="0" cellpadding="0" cellspacing="0" width="100%"
                               style="background-color:#eef9f0; border-left:4px solid #1e7e34;
                                      border-radius:4px; margin:24px 0;">
                            <tr>
                                <td style="padding:16px 20px;">
                                    <table border="0" cellpadding="6" cellspacing="0" width="100%">
                                        <tr>
                                            <td style="color:#666666; font-size:13px; width:160px;">Process Name</td>
                                            <td style="color:#2d2d2d; font-weight:600;">{stage_name}</td>
                                        </tr>
                                        <tr>
                                            <td style="color:#666666; font-size:13px;">Status</td>
                                            <td style="color:#1e7e34; font-weight:600;">Approved</td>
                                        </tr>
                                        <tr>
                                            <td style="color:#666666; font-size:13px;">Approved By</td>
                                            <td style="color:#2d2d2d; font-weight:600;">{approver.name or ''}</td>
                                        </tr>
                                    </table>
                                </td>
                            </tr>
                        </table>

                        <p style="margin:24px 0;">
                            <a href="{self.get_base_url()}/web#id={self.id}&model=manufacturing.stages&view_type=form"
                               style="display: inline-block; background-color: #1e7e34; color: #ffffff;
                                      text-decoration: none; font-size: 13px; font-weight: 600;
                                      padding: 10px 24px; border-radius: 6px;">
                                View Process →
                            </a>
                        </p>

                        <p style="margin:24px 0 4px 0; line-height:1.6;">Thanks &amp; regards,</p>
                        <p style="margin:0; font-weight:600;">{approver.name or ''}</p>
                    </td>
                </tr>

                <tr>
                    <td style="background-color:#f0f0f0; padding:16px 32px; border-top:1px solid #dddddd;">
                        <p style="margin:0; font-size:12px; color:#999999; text-align:center;">
                            This is an automated notification.
                        </p>
                    </td>
                </tr>

            </table>
            """

        mail_values = {
            'subject': f'Manufacturing Process Approved - {stage_name}',
            'email_to': creator.email,
            'email_from': (approver.email or self.env.company.email or ''),
            'body_html': mail_body,
        }

        mail = self.env['mail.mail'].sudo().create(mail_values)
        mail.sudo().send()

    def action_open_revision(self):
        # take baseline + create the (empty) revision record
        for rec in self:
            rec.last_revision_snapshot = json.dumps(rec._get_revision_values())
            revision = self.env['manufacturing.stages.revision'].create({
                'stage_id': rec.id,
                'revision_number': len(rec.revision_ids) + 1,
                'revision_type': 'revision',
                'html': '',
            })
            rec.current_revision_id = revision.id
            rec.state = 'in_revision'
        return True

    def write(self, vals):
        res = super().write(vals)
        # regenerate the live revision html whenever something is edited
        # while we are inside a revision cycle
        for rec in self:
            if rec.state == 'in_revision' and rec.current_revision_id \
                    and not self.env.context.get('skip_revision_html'):
                old_values = json.loads(rec.last_revision_snapshot or "{}")
                new_values = rec._get_revision_values()
                html = rec._generate_revision_html(old_values, new_values)
                rec.current_revision_id.with_context(
                    skip_revision_html=True).html = html
        return res

    def action_save_revision(self):
        for rec in self:
            rec.last_revision_snapshot = json.dumps(rec._get_revision_values())
            rec.current_revision_id = False
            rec.state = 'request'
        return True

    @api.model_create_multi
    def create(self, vals_list):
        records = super().create(vals_list)
        for rec in records:
            rec.last_revision_snapshot = json.dumps(rec._get_revision_values())
        return records



class ProductionProcessType(models.Model):
    _name = 'manufacturing.process.type'
    _description = 'Production Process'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    name                     = fields.Char(required=True)
    is_power_bank            = fields.Boolean()
    # show_degas               = fields.Boolean("Degas")
    # show_packing             = fields.Boolean("Packing")
    # show_injection           = fields.Boolean("Injection")
    # show_cell_drying         = fields.Boolean("Cell Drying")
    # show_high_temperature    = fields.Boolean("High Temperature")
    # show_cell_clamp_baking   = fields.Boolean("Cell Clamp Baking")
    # show_capacity_test       = fields.Boolean("Capacity Test")
    # show_voltage_test        = fields.Boolean("Voltage Test")
    # show_drying_process      = fields.Boolean("Drying Process")
    # show_slitting_process    = fields.Boolean("Slitting Process")
    # show_pad_printing        = fields.Boolean("Pad Printing")
    # show_aged_formation_1    = fields.Boolean("Aged Formation Cell 1")
    # show_aged_formation_2    = fields.Boolean("Aged Formation Cell 2")
    # is_anode_slitting_process = fields.Boolean("Anode Slitting Process")
    # is_cathode_slitting_process = fields.Boolean("Anode Slitting Process")
    prefix = fields.Char("Prefix")


    is_split_process = fields.Boolean(copy=False)
    is_packing_process = fields.Boolean(copy=False)
    is_voltage_process = fields.Boolean(copy=False)
    is_capacity_process = fields.Boolean(copy=False)
    enable_ocv_ir = fields.Boolean(copy=False)
    enable_capacity_test = fields.Boolean(copy=False)
    is_next_packing_process = fields.Boolean(copy=False)
    process_type1_id = fields.Many2one(
        'manufacturing.process.type',
        string='Process 1',
        domain="[('id', '!=', id), ('id', '!=', process_type2_id)]",
    )
    process_type2_id = fields.Many2one(
        'manufacturing.process.type',
        string='Process 2',
        domain="[('id', '!=', id), ('id', '!=', process_type1_id)]",
    )
    process_type3_id = fields.Many2one(
        'manufacturing.process.type',
        string='Last Process',
        domain="[('id', '!=', id), ('id', '!=', process_type1_id), ('id', '!=', process_type2_id)]",
    )
    # @api.model_create_multi
    # def create(self, vals_list):
    #     for vals in vals_list:
    #         if vals.get('is_split_process'):
    #             vals.update({
    #                 'enable_ocv_capacity': True,
    #                 'enable_ocv_ir': True,
    #                 'enable_capacity_test': True,
    #             })
    #     return super().create(vals_list)
    #
    # def write(self, vals):
    #     if vals.get('is_split_process'):
    #         vals.update({
    #             'enable_ocv_capacity': True,
    #             'enable_ocv_ir': True,
    #             'enable_capacity_test': True,
    #         })
    #     return super().write(vals)


    @api.constrains('is_power_bank','is_packing_process','is_split_process','is_next_packing_process')
    def _check_power_bank(self):
        for rec in self:
            if rec.is_next_packing_process:
                if not rec.process_type3_id:
                    raise UserError("Process type 3 is required.")
            elif not rec.is_next_packing_process:
                rec.process_type3_id = False

            if rec.is_split_process and (not rec.process_type1_id or not rec.process_type2_id):
                raise UserError(_("Please select both Process 1 and Process 2."))
            elif not rec.is_split_process:
                rec.process_type1_id = False
                rec.process_type2_id = False

            if rec.is_power_bank:
                existing_record = self.search([
                    ('is_power_bank', '=', True),
                    ('id', '!=', rec.id)
                ], limit=1)

                if existing_record:
                    raise UserError("Only one record can be marked as Power Bank.")
            # if rec.is_packing_process:
            #     existing_record = self.search([
            #         ('is_packing_process', '=', True),
            #         ('id', '!=', rec.id)
            #     ], limit=1)

                # if existing_record:
                #     raise UserError("Only one record can be marked as Packing Process.")

#

class MachineData(models.Model):
    _inherit = 'machine.parameter'

    manufacturing_process_id = fields.Many2one('manufacturing.process')
    qr_code_printing_condition = fields.Selection([
        ('good', 'Good'),
        ('fail', 'Fail')
    ], compute='_compute_qr_code_printing_condition')

    def _compute_qr_code_printing_condition(self):
        for rec in self:
            rec.qr_code_printing_condition = 'fail'
            if rec.manufacturing_process_id:
                if rec.manufacturing_process_id.operation_id.dry_min_temperature <= rec.temperature <= rec.manufacturing_process_id.operation_id.dry_max_temperature and abs(
                        rec.manufacturing_process_id.operation_id.dry_min_vacuum) <= abs(rec.vacuum) <= abs(
                        rec.manufacturing_process_id.operation_id.dry_max_vacuum):
                    rec.qr_code_printing_condition = 'good'
                else:
                    rec.qr_code_printing_condition = 'fail'


class Utilities(models.Model):
    _inherit = 'utility.parameter'

    manufacturing_process_id = fields.Many2one('manufacturing.process')
    qr_code_printing_utility_condition = fields.Selection([
        ('good', 'Good'),
        ('fail', 'Fail')
    ], compute='_compute_qr_code_printing_utility')

    def _compute_qr_code_printing_utility(self):
        for rec in self:
            # rec.cell_drying_utility_condition = 'fail'
            if rec.manufacturing_process_id:
                if rec.manufacturing_process_id.operation_id.min_humidity <= rec.humidity <= rec.manufacturing_process_id.operation_id.max_humidity and rec.manufacturing_process_id.operation_id.min_temperature <= rec.temperature <= rec.manufacturing_process_id.operation_id.max_temperature:
                    rec.qr_code_printing_utility_condition = 'good'
                else:
                    rec.qr_code_printing_utility_condition = 'fail'


class QualityDetails(models.Model):
    _inherit = 'quality.parameter'

    manufacturing_process_id = fields.Many2one('manufacturing.process')


class StockProductionLOt(models.Model):
    _inherit = 'stock.lot'

    manufacturing_process_id = fields.Many2one('manufacturing.process')


class StockMove(models.Model):
    _inherit = 'stock.move'

    manufacturing_process_id = fields.Many2one('manufacturing.process')


class ProductSerialNumber(models.Model):
    _inherit = 'product.serial.number'

    manufacturing_process_id = fields.Many2one('manufacturing.process')
    sub_manufacturing_process_id = fields.Many2one('manufacturing.process')
    packing_manufacturing_process_id = fields.Many2one('manufacturing.process')
    voltage_manufacturing_process_id = fields.Many2one('manufacturing.process')
    capacity_manufacturing_process_id = fields.Many2one('manufacturing.process')
    batch_id = fields.Many2one('manufacturing.batch', string='Batch',related='manufacturing_process_id.batch_id')
    cell_weight = fields.Float(string='Cell Weight (g)', digits=(16, 4))  #  was Char
    # _sql_constraints = [
    #     ('unique_capacity_serial', 'unique(name, capacity_manufacturing_process_id)',
    #      'This serial number is already added to this Capacity process!'),
    #     ('unique_voltage_serial', 'unique(name, voltage_manufacturing_process_id)',
    #      'This serial number is already added to this Voltage process!'),
    # ]

class StockMoveLIne(models.Model):
    _inherit = 'stock.move.line'

    manufacturing_process_id = fields.Many2one('manufacturing.process')
    out_manufacturing_process_id = fields.Many2one('manufacturing.process')
    out_tray_id = fields.Many2one('product.tray')


class ManufacturingComponents(models.Model):
    _inherit = 'manufacturing.component'

    manufacturing_process_id = fields.Many2one('manufacturing.process')

    @api.onchange('product_id')
    def _onchange_product_manufacturing_process_id(self):
        if self.manufacturing_process_id:
            self.location_src_id = self.manufacturing_process_id.location_src_id.id
            self.location_dest_id = self.manufacturing_process_id.location_dest_id.id


class MaterialLine(models.Model):
    _inherit = 'material.line'

    manufacturing_process_id = fields.Many2one('manufacturing.process')
#
    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            # If manufacturing_process_id is set but production_plan_id is not, auto-fill it
            if vals.get('manufacturing_process_id') and not vals.get('production_plan_id'):
                process = self.env['manufacturing.process'].browse(vals['manufacturing_process_id'])
                if process.production_plan_id:
                    vals['production_plan_id'] = process.production_plan_id.id
        return super().create(vals_list)







class ProductionProcess(models.Model):
    _name = 'manufacturing.process'
    _description = 'Production Process Type'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = "id desc"



    @api.model
    def default_get(self, fields):
        defaults = super(ProductionProcess, self).default_get(fields)
        production_location = self.env['stock.location'].search([('usage', '=', 'production')], limit=1)
        defaults['production_location_id'] = production_location.id
        return defaults

    sequence = fields.Integer(index=True)
    def _get_default_product_uom_id(self):
        return self.env['uom.uom'].search([], limit=1, order='id').id
    company_id = fields.Many2one( 'res.company', 'Company', index=True, default=lambda self: self.env.company)
    location_src_id = fields.Many2one('stock.location')
    location_dest_id = fields.Many2one('stock.location')
    production_location_id = fields.Many2one('stock.location')


    lot_ids = fields.One2many('product.serial.number', 'manufacturing_process_id', string='Serial Number')
    sub_process_lot_ids = fields.One2many('product.serial.number', 'sub_manufacturing_process_id', string='Serial Number')
    packing_lot_ids = fields.One2many('product.serial.number', 'packing_manufacturing_process_id', string='Serial Number')
    voltage_lot_ids = fields.One2many('product.serial.number', 'voltage_manufacturing_process_id', string='Serial Number')
    capacity_lot_ids = fields.One2many('product.serial.number', 'capacity_manufacturing_process_id', string='Serial Number')

    name = fields.Char(
        'Reference', copy=False, readonly=True, default=lambda x: _('New'))
    product_id = fields.Many2one(
        'product.product', 'Product',
        domain="""[
               ('type', 'in', ['combo', 'consu']),
               '|',
                   ('company_id', '=', False),
                   ('company_id', '=', company_id)
           ]
           """,
        check_company=True,)
    bom_id = fields.Many2one(
        'manufacturing.bom', 'Bill of Material')
    component_ids = fields.One2many('manufacturing.component', 'manufacturing_process_id', string='Components', copy=False)
    finished_lines = fields.One2many('mrp.finished.line', 'manufacturing_process_id', string="Finished Lines", copy=False)
    state = fields.Selection([
        ('draft', 'Draft'),
        ('confirmed', 'Confirmed'),
        ('progress', 'In Progress'),
        ('hold', 'Hold'),
        ('done', 'Done'),
        ('close', 'Closed'),
        ('completed', 'Completed'),
        ('cancel', 'Cancelled')], string='State',
        copy=False, index=True, default='draft',
        store=True, tracking=True,
        help=" * Draft: The MO is not confirmed yet.\n"
             " * Confirmed: The MO is confirmed, the stock rules and the reordering of the components are triggerred.\n"
             " * In Progress: The production has started (on the MO or on the WO).\n"
             " * Close: The production is done, the MO has to be closed.\n"
             " * Done: The MO is closed, the stock moves are posted. \n"
             " * Cancelled: The MO has been cancelled, can't be confirmed anymore.")
    product_tracking = fields.Selection(related='product_id.tracking')
    product_tmpl_id = fields.Many2one('product.template', 'Product Template', related='product_id.product_tmpl_id')
    product_qty = fields.Float(
        'Quantity To Produce',
        default=1.0, digits='Product Unit of Measure',
        required=True, tracking=True,
       )
    product_uom_id = fields.Many2one(
        'uom.uom', 'Product Unit of Measure', default=_get_default_product_uom_id,
        readonly=True, required=True,
       domain="[('relative_uom_id', '=', product_uom_category_id)]")

    qty_producing = fields.Float(string="Quantity Producing", digits='Product Unit of Measure', copy=False)
    product_uom_category_id = fields.Many2one(related='product_id.uom_id.relative_uom_id')
    product_uom_qty = fields.Float(string='Total Quantity', store=True)
    user_id = fields.Many2one(
        'res.users', 'Responsible', default=lambda self: self.env.user,
        states={'done': [('readonly', True)], 'cancel': [('readonly', True)]})
    qty_produced = fields.Float(string="Quantity Produced")
    operation_id = fields.Many2one('manufacturing.operation')
    production_plan_id = fields.Many2one('production.plan')
    # injection_id= fields.Many2one('cell.injection')
    # cell_drying_id= fields.Many2one('cell.drying')

    product_model_id = fields.Many2one('product.model')

    machine_data_ids = fields.One2many(
        'machine.parameter', 'manufacturing_process_id', 'Machine Data',
        copy=False, states={'done': [('readonly', True)], 'cancel': [('readonly', True)]})
    utility_ids = fields.One2many(
        'utility.parameter', 'manufacturing_process_id', 'Utility Parameter',
        copy=False, states={'done': [('readonly', True)], 'cancel': [('readonly', True)]})
    quality_ids = fields.One2many(
        'quality.parameter', 'manufacturing_process_id', 'Quality',
        copy=False, states={'done': [('readonly', True)], 'cancel': [('readonly', True)]})
    machine_id = fields.Many2one('manufacturing.machine')
    start_time = fields.Datetime()
    end_time = fields.Datetime()
    # product_tray_id = fields.Many2one('product.tray')
    # assembly_id = fields.Many2one('assembly.cell')
    breakdown_ids = fields.One2many('production.breakdown', 'manufacturing_process_id')
    # injection_count = fields.Integer(compute='_compute_injection_count')
    finished_move_ids = fields.One2many('stock.move.line', 'out_manufacturing_process_id')
    expected_end_time = fields.Datetime("Expected End", compute='compute_end_date')
    out_file_name = fields.Char("Serial File")
    out_file = fields.Binary("Serials")
    input_material_lines = fields.One2many('material.line', 'manufacturing_process_id')
    remaining_hours = fields.Float("Remaining Time", compute='_get_remaining_time')
    process_status_check = fields.Boolean(compute='_compute_process_status_check')

    manufacturing_process_id = fields.Many2one('manufacturing.process')
    manufacturing_process_type_id = fields.Many2one('manufacturing.process.type',string="Operation Type")

    main_manufacturing_process_id = fields.Many2one('manufacturing.process')

    before_manufacturing_process_id = fields.Many2one('manufacturing.process')
    before_manufacturing_process_type_id = fields.Many2one('manufacturing.process.type')
    before_manufacturing_process_type_name = fields.Char(related='before_manufacturing_process_type_id.name')

    next_manufacturing_process_id = fields.Many2one('manufacturing.process')
    next_manufacturing_process_type_id = fields.Many2one('manufacturing.process.type')

    sub_process_manufacturing_process_id = fields.Many2one('manufacturing.process')


    is_first_process = fields.Boolean(copy=False)
    is_sub_process = fields.Boolean(copy=False)
    quality_count = fields.Integer(compute='_compute_quality_count')
    next_manufacturing_process_type_name = fields.Char(string="Next Process Type",compute='_compute_next_process_type_name', store=True )
    next_manufacturing_process_type_name_2 = fields.Char(string="Next Process Type",compute='_compute_next_process_type_name', store=True )
    is_next_process_created = fields.Boolean(copy=False)
    done_lot = fields.Boolean('')
    remaining_qty = fields.Integer(compute='_compute_remaining_qty',store=True)
    batch_id = fields.Many2one('manufacturing.batch', related='production_plan_id.batch_id', string='Batch',store=True)
    check_available = fields.Boolean(copy=False)

    is_capacity_created = fields.Boolean(copy=False)
    is_voltage_created = fields.Boolean(copy=False)
    remaining_qty_process_created = fields.Boolean(copy=False)
    name_1 = fields.Char(compute='_compute_next_process_type_name' ,store=True)
    name_2 = fields.Char(compute='_compute_next_process_type_name',  store=True)
    name_3 = fields.Char(compute='_compute_next_process_type_name', store=True )

    is_split_process = fields.Boolean(compute='_compute_process_flags', store=True)
    is_voltage_process = fields.Boolean(compute='_compute_process_flags', store=True)
    is_packing_process = fields.Boolean(compute='_compute_process_flags', store=True)
    is_capacity_process = fields.Boolean(compute='_compute_process_flags', store=True)
    enable_ocv_ir = fields.Boolean(compute='_compute_process_flags', store=True)
    enable_capacity_test = fields.Boolean(compute='_compute_process_flags', store=True)
    is_next_packing_process = fields.Boolean(compute='_compute_process_flags', store=True)
    is_packing_process_next = fields.Boolean(compute='_compute_process_flags', store=True)
    is_next_voltage_process = fields.Boolean(compute='_compute_process_flags', store=True)
    is_next_capacity_process = fields.Boolean(compute='_compute_process_flags', store=True)
    is_power_bank = fields.Boolean(copy=False)
    is_sub_process_created = fields.Boolean(copy=False)
    allow_lot_create = fields.Boolean(compute='_compute_allow_lot_create',store=True)
    capacity_lots = fields.Many2many(
        'product.serial.number',
        'manufacturing_process_capacity_lots_rel',
        'manufacturing_process_id',
        'serial_id',
        string='Capacity Lots',
    )
    voltage_lots = fields.Many2many(
        'product.serial.number',
        'manufacturing_process_voltage_lots_rel',
        'manufacturing_process_id',
        'serial_id',
        string='Voltage Lots',
    )

    rejection_reason_html = fields.Html(
        string="Rejection Summary",
        compute="_compute_rejection_summary",
        sanitize=False,
    )
    capacity_dest_location_id = fields.Many2one('stock.location', string='Capacity Destination Location')
    voltage_dest_location_id = fields.Many2one('stock.location', string='Voltage Destination Location')

    process_type1_id = fields.Many2one('manufacturing.process.type',compute='_compute_process_flags', store=True)
    process_type2_id = fields.Many2one('manufacturing.process.type',compute='_compute_process_flags', store=True)
    manufacturing_split_process1_id = fields.Many2one('manufacturing.process')
    manufacturing_split_process2_id = fields.Many2one('manufacturing.process')

    def _compute_quality_count(self):
        for rec in self:
            rec.quality_count = self.env['mrp.quality'].search([
                ('production_plan_id', '=', rec.production_plan_id.id),
                ('manufacturing_process_id', '=', rec.id),
                ('operation_id', '=', rec.operation_id.id),
            ])

    @api.onchange('capacity_lots')
    def _onchange_capacity_lots(self):
        # If a lot gets added to Capacity, automatically drop it from Voltage
        if self.capacity_lots and self.voltage_lots:
            overlap = self.voltage_lots & self.capacity_lots
            if overlap:
                self.voltage_lots = self.voltage_lots - overlap

    @api.onchange('voltage_lots')
    def _onchange_voltage_lots(self):
        # If a lot gets added to Voltage, automatically drop it from Capacity
        if self.voltage_lots and self.capacity_lots:
            overlap = self.capacity_lots & self.voltage_lots
            if overlap:
                self.capacity_lots = self.capacity_lots - overlap

    @api.model_create_multi
    def create(self, vals_list):
        records = super().create(vals_list)
        records._dedupe_capacity_voltage_lots()
        return records

    def write(self, vals):
        res = super().write(vals)
        if 'capacity_lots' in vals or 'voltage_lots' in vals:
            self._dedupe_capacity_voltage_lots()
        return res

    def _dedupe_capacity_voltage_lots(self):
        # Whatever is in capacity_lots always wins; silently remove overlap from voltage_lots
        for rec in self:
            overlap = rec.capacity_lots & rec.voltage_lots
            if overlap:
                rec.voltage_lots = rec.voltage_lots - overlap

    def _compute_rejection_summary(self):
        for rec in self:
            get_quality = self.env['mrp.quality'].search([
                ('manufacturing_process_id', '=', rec.id),
            ])
            if get_quality:
                reason_count = Counter(get_quality.mapped('reason'))

                html = """
                      <div style="font-family: Arial, sans-serif;">
                          <table style="border-collapse: collapse; width: 100%;">
                              <thead>
                                  <tr style="background-color:#f5f5f5;">
                                      <th style="border:1px solid #ddd;padding:8px;text-align:left;">Reason</th>
                                      <th style="border:1px solid #ddd;padding:8px;text-align:center;">Count</th>
                                  </tr>
                              </thead>
                              <tbody>
                  """
                for reason, count in reason_count.items():
                    html += """
                          <tr>
                              <td style="border:1px solid #ddd;padding:8px;">{}</td>
                              <td style="border:1px solid #ddd;padding:8px;text-align:center;">{}</td>
                          </tr>
                      """.format(reason or '-', count)
                html += """
                              </tbody>
                          </table>
                      </div>
                  """
                rec.rejection_reason_html = html
            else:
                rec.rejection_reason_html = ""

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if not vals.get('name') or vals['name'] == _('New'):
                manufacturing_process_type_id = vals.get('manufacturing_process_type_id')
                prefix = ''
                if manufacturing_process_type_id:
                    operation = self.env['manufacturing.process.type'].browse(manufacturing_process_type_id)
                    prefix = operation.prefix or ''
                number = self.env['ir.sequence'].next_by_code('manufacturing.process') or _('New')
                vals['name'] = f"{prefix}/{number}" if prefix else number
        records = super(ProductionProcess, self).create(vals_list)
        return records

    @api.depends('operation_id')
    def _compute_allow_lot_create(self):
        for rec in self:
            get_lot_create = self.env['manufacturing.process'].search([('is_first_process','=',True),('production_plan_id','=',self.production_plan_id.id)])
            rec.allow_lot_create = get_lot_create.operation_id.allow_lot_create


    @api.depends('manufacturing_process_type_id', 'is_capacity_created', 'is_voltage_created','state',
                 'remaining_qty_process_created', 'next_manufacturing_process_type_id')
    def _compute_process_flags(self):
        for rec in self:
            rec.is_split_process = rec.manufacturing_process_type_id.is_split_process
            # if self.manufacturing_process_type_id.is_split_process:
            rec.process_type1_id = rec.manufacturing_process_type_id.process_type1_id.id
            rec.process_type2_id = rec.manufacturing_process_type_id.process_type2_id.id
            all_lines = self.production_plan_id.operation_ids
            get_process_1 = all_lines.filtered(
                lambda x: x.manufacturing_process_type_id.id == self.process_type1_id.id
            )
            get_process_2 = all_lines.filtered(
                lambda x: x.manufacturing_process_type_id.id == self.process_type2_id.id
            )
            get_process_1_line = get_process_1[:1]
            get_process_2_line = get_process_2[:1]

            self.voltage_dest_location_id = get_process_1_line.operation_id.location_src_id.id
            self.capacity_dest_location_id = get_process_2_line.operation_id.location_src_id.id
            # rec.is_voltage_process = rec.manufacturing_process_type_id.is_voltage_process
            # rec.is_packing_process = rec.manufacturing_process_type_id.is_packing_process
            # rec.is_capacity_process = rec.manufacturing_process_type_id.is_capacity_process
            # rec.enable_ocv_ir = rec.manufacturing_process_type_id.enable_ocv_ir
            # rec.enable_capacity_test = rec.manufacturing_process_type_id.enable_capacity_test
            # rec.is_next_voltage_process = rec.next_manufacturing_process_type_id.is_voltage_process
            # rec.is_next_capacity_process = rec.next_manufacturing_process_type_id.is_capacity_process
            # rec.is_next_packing_process = rec.next_manufacturing_process_type_id.is_packing_process
            # rec.is_packing_process_next = rec.manufacturing_process_type_id.is_next_packing_process

    def action_open_voltage_process(self):
        self.ensure_one()
        process = self.env['manufacturing.process'].search([
            ('before_manufacturing_process_id', '=', self.id),
            ('manufacturing_process_type_id', '=', self.manufacturing_process_type_id.process_type1_id.id),
        ], limit=1)
        return {
            'name': self.name_1,
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'manufacturing.process',
            'res_id': process.id,
        }

    def action_open_capacity_process(self):
        process = self.env['manufacturing.process'].search([
            ('before_manufacturing_process_id', '=', self.id),
            ('manufacturing_process_type_id', '=', self.manufacturing_process_type_id.process_type2_id.id),
        ], limit=1)
        return {
            'name': self.name_2,
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'manufacturing.process',
            'res_id': process.id,
        }

    # def action_open_remain_qty_process(self):
    #     process = self.env['manufacturing.process'].search([
    #         ('before_manufacturing_process_id', '=', self.id),
    #         ('is_packing_process', '=', True),
    #     ], limit=1)
    #     return {
    #         'name': self.name_1,
    #         'type': 'ir.actions.act_window',
    #         'view_mode': 'form',
    #         'res_model': 'manufacturing.process',
    #         'res_id': process.id,
    #
    #     }

    @api.depends('next_manufacturing_process_type_id', 'state','is_split_process',
                 'is_voltage_created', 'is_capacity_created',
                 'remaining_qty_process_created')
    def _compute_next_process_type_name(self):
        for rec in self:
            # defaults — guarantees every computed field is always set
            rec.next_manufacturing_process_type_name = ''
            rec.next_manufacturing_process_type_name_2 = ''
            rec.name_1 = ''
            rec.name_2 = ''
            rec.name_3 = ''

            if rec.next_manufacturing_process_type_id:
                rec.next_manufacturing_process_type_name = f"Create {rec.next_manufacturing_process_type_id.name}"
                rec.next_manufacturing_process_type_name_2 = f"Open {rec.next_manufacturing_process_type_id.name}"

            if rec.is_split_process:
                pt1 = rec.manufacturing_process_type_id.process_type1_id
                pt2 = rec.manufacturing_process_type_id.process_type2_id

                rec.name_1 = f"{'Open' if rec.is_voltage_created else 'Create'} {pt1.name}"
                rec.name_2 = f"{'Open' if rec.is_capacity_created else 'Create'} {pt2.name}"
                # If you re-enable the third branch later:
                # pt3 = rec.manufacturing_process_type_id.process_type3_id
                # rec.name_3 = f"{'Open' if rec.remaining_qty_process_created else 'Create'} {pt3.name}"

    def action_create_split_process_type_1(self):
        # all_lines = self.production_plan_id.operation_ids.sorted('sequence')
        # current_line = all_lines.filtered(
        #     lambda x: x.manufacturing_process_type_id.is_voltage_process == True)
        # if not current_line:
        #     raise UserError(_("No packing process found in this Production Plan."))
        next_process = self.env['manufacturing.process'].create({
            'before_manufacturing_process_id': self.id,
            'before_manufacturing_process_type_id': self.manufacturing_process_type_id.id,
            'manufacturing_process_type_id': self.process_type1_id.id,
            'production_plan_id': self.production_plan_id.id,
            'product_model_id': self.product_model_id.id,
            'is_first_process': False,
            'done_lot': True,
        })

        self.write({'manufacturing_split_process1_id':next_process.id,'is_voltage_created':True})
        next_process._onchange_of_product_plan_id()
        next_process.write({
            'lot_ids': [(4, lot_id) for lot_id in self.voltage_lot_ids.ids],
        })
        return {
            'name': self.process_type1_id.name,
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'manufacturing.process',
            'res_id': next_process.id,
        }
    def action_create_split_process_type_2(self):
        # all_lines = self.production_plan_id.operation_ids.sorted('sequence')
        # current_line = all_lines.filtered(
        #     lambda x: x.manufacturing_process_type_id.is_capacity_process == True)
        # if not current_line:
        #     raise UserError(_("No packing process found in this Production Plan."))
        # Step 1: create WITHOUT lot_ids
        next_process = self.env['manufacturing.process'].create({
            'before_manufacturing_process_id': self.id,
            'before_manufacturing_process_type_id': self.manufacturing_process_type_id.id,
            'manufacturing_process_type_id': self.process_type2_id.id,
            'production_plan_id': self.production_plan_id.id,
            'product_model_id': self.product_model_id.id,
            'is_first_process': False,
            'done_lot': True,

        })
        self.write({'manufacturing_split_process2_id':next_process.id,'is_capacity_created':True})
        next_process._onchange_of_product_plan_id()
        next_process.write({
            'lot_ids': [(4, lot_id) for lot_id in self.capacity_lot_ids.ids],
        })
        return {
            'name': self.process_type2_id.name,
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'manufacturing.process',
            'res_id': next_process.id,
        }
    def action_remain_qty_process(self):
        # all_lines = self.production_plan_id.operation_ids.sorted('sequence')
        # current_line = all_lines.filtered(
        #     lambda x: x.manufacturing_process_type_id.is_packing_process == True)
        # if not current_line:
        #     raise UserError(_("No packing process found in this Production Plan."))
        # Step 1: create WITHOUT lot_ids
        next_process = self.env['manufacturing.process'].create({
            'before_manufacturing_process_id': self.id,
            'before_manufacturing_process_type_id': self.manufacturing_process_type_id.id,
            'manufacturing_process_type_id': self.manufacturing_process_type_id.process_type3_id.id,
            'production_plan_id': self.production_plan_id.id,
            'product_model_id': self.product_model_id.id,
            'is_first_process': False,
            'done_lot': True,
            'is_packing_process': True,

        })
        self.write({'remaining_qty_process_created':True})
        next_process._onchange_of_product_plan_id()
        next_process.write({
            'lot_ids': [(4, lot_id) for lot_id in self.packing_lot_ids.ids],
        })
        return {
            'name': current_line.manufacturing_process_type_id.name,
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'manufacturing.process',
            'res_id': next_process.id,
        }

    def action_create_capacity_lot(self):
        for rec in self:
            if not rec.capacity_lots:
                raise UserError(_("Please select the Capacity Lots."))

            existing_names = set(rec.capacity_lot_ids.mapped('name'))
            new_serials = rec.capacity_lots.filtered(lambda s: s.name not in existing_names)

            if not new_serials:
                raise UserError(_("Selected serials are already added to Capacity lot."))

            for serial in new_serials:
                self.env['product.serial.number'].create({
                    'product_id': serial.product_id.id,
                    'name': serial.name,
                    'product_uom_id': serial.product_uom_id.id,
                    'cell_weight': serial.cell_weight,
                    'lot_id': serial.lot_id.id,
                    'batch_id': serial.batch_id.id if serial.batch_id else False,
                    'capacity_manufacturing_process_id': rec.id,
                })
        self.action_create_packing_lot()

    def action_create_voltage_lot(self):
        for rec in self:
            if not rec.voltage_lots:
                raise UserError(_("Please select the Voltage Lots."))

            existing_names = set(rec.voltage_lot_ids.mapped('name'))
            new_serials = rec.voltage_lots.filtered(lambda s: s.name not in existing_names)

            if not new_serials:
                raise UserError(_("Selected serials are already added to Voltage lot."))

            for serial in new_serials:
                self.env['product.serial.number'].create({
                    'product_id': serial.product_id.id,
                    'name': serial.name,
                    'product_uom_id': serial.product_uom_id.id,
                    'cell_weight': serial.cell_weight,
                    'lot_id': serial.lot_id.id,
                    'batch_id': serial.batch_id.id if serial.batch_id else False,
                    'voltage_manufacturing_process_id': rec.id,
                })
        self.action_create_packing_lot()


    def action_create_packing_lot(self):
        for rec in self:
            assigned_names = set()

            if rec.manufacturing_process_type_id.process_type1_id and rec.manufacturing_process_type_id.process_type2_id:
                if not rec.voltage_lot_ids or not rec.capacity_lot_ids:
                    continue
                assigned_names = set(rec.voltage_lot_ids.mapped('name')) | set(rec.capacity_lot_ids.mapped('name'))
            elif rec.manufacturing_process_type_id.process_type1_id:
                assigned_names = set(rec.voltage_lot_ids.mapped('name'))
            elif rec.manufacturing_process_type_id.process_type2_id:
                assigned_names = set(rec.capacity_lot_ids.mapped('name'))

            all_serials = rec.lot_ids
            remaining_serials = all_serials.filtered(lambda s: s.name not in assigned_names)

            existing_packing_names = set(rec.packing_lot_ids.mapped('name'))
            new_serials = remaining_serials.filtered(lambda s: s.name not in existing_packing_names)

            for serial in new_serials:
                self.env['product.serial.number'].create({
                    'product_id': serial.product_id.id,
                    'name': serial.name,
                    'product_uom_id': serial.product_uom_id.id,
                    'cell_weight': serial.cell_weight,
                    'lot_id': serial.lot_id.id,
                    'batch_id': serial.batch_id.id if serial.batch_id else False,
                    'packing_manufacturing_process_id': rec.id,
                })

    def action_complete_sub_process(self):
        self.before_manufacturing_process_id.write({'state': 'close'})
        self.action_done_production()
        self.action_rejection_complete()

    def action_rejection_complete(self):
        get_quality = self.env['mrp.quality'].search([
            ('manufacturing_process_id', '=', self.id)
        ])

        for qc in get_quality:
            if qc.state != 'done':
                raise UserError(
                    _("Please complete all Quality Checks before proceeding. Complete them as Rework or Scrap.")
                )
        self.state = 'close'

    @api.depends('lot_ids', 'product_qty', 'state')
    def _compute_remaining_qty(self):
        for rec in self:
            rec.remaining_qty = len(rec.lot_ids) if rec.lot_ids else rec.product_qty

    # def action_view_lots(self):
    #     lot_ids = []
    #     production_plans = self.env['production.plan'].search([('state', '=', 'in_production')])
    #     lot_ids += self.env['stock.lot'].search([('production_plan_id', 'in', production_plans.ids),
    #                                              ('final_location_id', '=', self.locatio  n_src_id.id)]).mapped('id')
    #     return {
    #         'name': _('Available Lots'),
    #         'type': 'ir.actions.act_window',
    #         'view_mode': 'list,form',
    #         'res_model': 'stock.lot',
    #         'domain': [('id', 'in', lot_ids), ('location_id.usage', '!=', 'inventory')],
    #         'context': {'group_by': ['production_plan_id', 'final_location_id']},
    #     }

    def action_view_lots(self):
        self.ensure_one()

        move_lines = self.finished_move_ids if self.is_first_process else self.component_ids

        return {
            'name': _('Available Lots'),
            'type': 'ir.actions.act_window',
            'res_model': 'stock.lot',
            'view_mode': 'list,form',
            'domain': [('id', 'in', move_lines.mapped('lot_id').ids)],
        }
    def lot_creation(self):
        if self.product_id:
            self.product_id.write({
                'tracking': 'serial',
            })

        # for lot in self.lot_ids:
        #     if lot.lot_id:
        #         available_qty = self.env['stock.quant']._get_available_quantity(
        #             self.product_id,
        #             self.production_location_id,
        #             lot_id=lot.lot_id
        #         )
        #         if available_qty == 0:
        #             _logger.error(_("Lot not available in %s" % self.production_location_id.name))
                    # raise UserError(_("Lot not available in %s" % self.production_location_id.name))

        if self.product_id.tracking != 'none':
            for serial in self.lot_ids:
                if not serial.lot_id:
                    if not self.allow_lot_create:
                        _logger.error(
                            _("The lot %s is not available and The current operation is not allowed to create new lot number.\n Please enable lot creation or check the inventory." % serial.name))

                    lot_id = self.env['stock.lot'].create({
                        'name': serial.name,
                        'ref': serial.name,
                        'product_id': self.product_id.id,
                        'company_id': self.env.company.id,
                        'production_plan_id': self.production_plan_id.id,
                        'final_location_id': self.production_location_id.id,
                    })
                    update_stock = self.env['stock.quant'].sudo().create({
                        'product_id': self.product_id.id,
                        'location_id': self.production_location_id.id,
                        'lot_id': lot_id.id,
                        'inventory_quantity': 1.0,
                    })
                    update_stock.action_apply_inventory()
                    serial.write({'lot_id': lot_id.id})

        self.done_lot = True


    # @api.depends('pa_count')
    def _compute_process_status_check(self):
        for rec in self:
            process = self.env['first.article.inspection'].search(
                [('origin', '=', rec.name), ('state', '=', 'done'), ('inspection_type', '=', 'process')])
            if process:
                rec.process_status_check = True
            else:
                rec.process_status_check = False

    @api.onchange('location_src_id')
    def onchange_line(self):
        for rec in self.component_ids:
            rec.location_src_id = self.location_src_id


    def action_view_quality(self):
        return {
            'res_model': 'mrp.quality',
            'type': 'ir.actions.act_window',
            'name': _("Quality Check"),
            'domain': [('manufacturing_process_id', '=', self.id)],
            'view_mode': 'list,form',
        }

    def action_view_in_process_quality(self):
        in_process_records = self.env['process.quality.check'].search([('opt_id', '=', self.operation_id.id), ('origin', '=', self.name)])
        return {
            'res_model': 'process.quality.check',
            'type': 'ir.actions.act_window',
            'name': _("In Process Quality Check"),
            'domain': [('id', 'in', in_process_records.ids if in_process_records else [])],
            'view_mode': 'list,form',
        }

    def _get_remaining_time(self):
        for rec in self:
            hr_remain = 0
            if rec.start_time and rec.operation_id.process_duration:
                end = rec.start_time + timedelta(hours=rec.operation_id.process_duration)
                hr_remain = ((end - fields.Datetime.now()).seconds / 3600)
            rec.remaining_hours = hr_remain

    def action_remove_line(self):
        self.component_ids.unlink()

    def action_download_sample(self):
        return {
            "type": "ir.actions.act_url",
            "url": '/fnet_mrp/static/Sample_Serials.xlsx',
            "target": "new",
        }

    def action_upload_serial(self):
        if not self.out_file:
            raise UserError(_("Please upload the serial number updated file."))
        file_data = base64.b64decode(self.out_file)
        wb = load_workbook(filename=io.BytesIO(file_data))
        sheet = wb.active
        serial_data = []  # list of (serial_no, cell_weight)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_values = list(row)
            if not row_values or not row_values[0]:
                continue
            serial_no = str(row_values[0])
            # ── read cell_weight from column B (index 1) ─────────────────────
            cell_weight = 0.0
            if len(row_values) > 1 and row_values[1] is not None:
                try:
                    cell_weight = float(row_values[1])
                except (ValueError, TypeError):
                    cell_weight = 0.0
            serial_data.append((serial_no, cell_weight))

        if len(serial_data) != self.product_qty:
            raise UserError(_(
                "The uploaded file contains %s serial numbers, but the required quantity is %s."
            ) % (len(serial_data), self.product_qty))

        # NEW: check for serials that already exist before creating anything
        uploaded_names = [s[0] for s in serial_data]
        existing_lots = self.env['stock.lot'].search([
            ('name', 'in', uploaded_names),
        ])
        if existing_lots:
            raise UserError(_(
                "The following serial number(s) already exist and cannot be uploaded again: %s"
            ) % ", ".join(existing_lots.mapped('name')))

        for serial_no, cell_weight in serial_data:
            self.env['product.serial.number'].create({
                'product_id': self.product_id.id,
                'name': serial_no,
                'product_uom_id': self.product_uom_id.id,
                'cell_weight': cell_weight,  # ← from column B
                'batch_id': self.production_plan_id.batch_id.id if self.production_plan_id.batch_id else False,
                'manufacturing_process_id': self.id,
            })


    @api.depends('start_time', 'operation_id')
    def compute_end_date(self):
        for rec in self:
            if rec.start_time and rec.operation_id:
                rec.expected_end_time = rec.start_time + timedelta(hours=rec.operation_id.process_duration)
            else:
                rec.expected_end_time = False

    # @api.constrains('end_time', 'expected_end_time')
    def duration_constrain(self):
        for rec in self:
            if rec.end_time < rec.expected_end_time:
                time = timedelta(hours=rec.operation_id.process_duration)
                dt = datetime(2000, 1, 1) + time
                raise UserError("Minimum duration to stop process is %s hours." % dt.strftime("%H:%M"))

    def action_open_production_plan(self):
        self.ensure_one()

        return {
            'name': _('Production Plan'),
            'type': 'ir.actions.act_window',
            'res_model': 'production.plan',
            'view_mode': 'form',
            'res_id': self.production_plan_id.id,
            'target': 'current',
        }

    def action_open_before_cell_process(self):
        self.ensure_one()

        return {
            'name': self.before_manufacturing_process_type_name or _('Previous Process'),
            'type': 'ir.actions.act_window',
            'res_model': 'manufacturing.process',
            'view_mode': 'form',
            'res_id': self.before_manufacturing_process_id.id,
            'target': 'current',
        }


    def action_open_next_cell_process(self):
        return {
            'name': self.next_manufacturing_process_type_name_2,
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'manufacturing.process',
            'res_id': self.next_manufacturing_process_id.id,  # Assuming records is a single record
        }

    def action_create_remain_process(self):
        get_quality = self.env['mrp.quality'].search([
            ('manufacturing_process_id', '=', self.id)
        ])
        for qc in get_quality:
            if qc.state != 'done':
                raise UserError(
                    _("Please complete all Quality Checks before proceeding. Complete them as Rework or Scrap.")
                )
        next_process = self.env['manufacturing.process'].create({
            'before_manufacturing_process_id': self.id,
            'before_manufacturing_process_type_id': self.manufacturing_process_type_id.id,
            'manufacturing_process_type_id': self.next_manufacturing_process_type_id.id,
            'production_plan_id': self.production_plan_id.id,
            'product_model_id': self.product_model_id.id,
            'is_first_process': False,
            'done_lot': False,
        })
        self.write({'next_manufacturing_process_id': next_process.id, 'is_next_process_created': True})
        next_process._onchange_of_product_plan_id()
        next_process.write({
            'lot_ids': [(4, lot_id) for lot_id in self.packing_lot_ids.ids],
        })

        if self.allow_lot_create:
            next_process.done_lot = True
            next_process.lot_creation()

        return {
            'name': self.next_manufacturing_process_id.manufacturing_process_type_id.name,
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'manufacturing.process',
            'res_id': next_process.id,

        }
    def action_create_next_cell_process(self):
        get_quality = self.env['mrp.quality'].search([
            ('manufacturing_process_id', '=', self.id)
        ])
        for qc in get_quality:
            if qc.state != 'done':
                raise UserError(
                    _("Please complete all Quality Checks before proceeding. Complete them as Rework or Scrap.")
                )

        next_process = self.env['manufacturing.process'].create({
            'before_manufacturing_process_id': self.id,
            'before_manufacturing_process_type_id': self.manufacturing_process_type_id.id,
            'manufacturing_process_type_id': self.next_manufacturing_process_type_id.id,
            'production_plan_id': self.production_plan_id.id,
            'product_model_id': self.product_model_id.id,
            'is_first_process': False,
            'done_lot': False,
        })
        self.write({'next_manufacturing_process_id': next_process.id, 'is_next_process_created': True})
        next_process._onchange_of_product_plan_id()
        next_process.write({
            'lot_ids': [(4, lot_id) for lot_id in self.lot_ids.ids],
        })
        if self.allow_lot_create:
            next_process.done_lot = True
            next_process.lot_creation()

        return {
            'name': self.next_manufacturing_process_id.manufacturing_process_type_id.name,
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'manufacturing.process',
            'res_id': next_process.id,
        }



    def action_create_sub_process(self):
        # Step 1: create WITHOUT lot_ids
        sub_process = self.env['manufacturing.process'].create({
            'main_manufacturing_process_id': self.id,
            'before_manufacturing_process_id': self.id,
            'manufacturing_process_type_id': self.manufacturing_process_type_id.id,
            'production_plan_id': self.production_plan_id.id,
            'product_model_id': self.product_model_id.id,
            'is_first_process': False,
            'is_sub_process': True,
            'done_lot': True,
        })
        self.write({'sub_process_manufacturing_process_id': sub_process.id})
        sub_process._onchange_of_product_plan_id()

        if sub_process.product_id:
            sub_process.product_id.write({'tracking': 'serial'})
        sub_process.write({
            'lot_ids': [(4, lot_id) for lot_id in self.sub_process_lot_ids.ids],
        })
        for line in sub_process.input_material_lines:
            line.action_upload_serial()
        sub_process.check_available_stock()

        # self.state = 'close'
        return {
            'name': 'Sub Process',
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'manufacturing.process',
            'res_id': sub_process.id,
        }

    def action_open_sub_process(self):
        return {
            'name': "Sub Process",
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'manufacturing.process',
            'res_id': self.sub_process_manufacturing_process_id.id,  # Assuming records is a single record
        }

    def action_open_main_process(self):
        return {
            'name': 'Main Process',
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'manufacturing.process',
            'res_id': self.main_manufacturing_process_id.id,  # Assuming records is a single record
        }
    @api.onchange('state')
    def _onchange_of_state(self):
        if self.state and self.component_ids:
            for line in self.component_ids:
                line.state = self.state

    def _resolve_bom(self):
        """Common lookup shared by every process branch."""
        production_operation = self.env['production.operation'].search(
            [
                ('manufacturing_process_type_id', '=', self.manufacturing_process_type_id.id),
                ('production_plan_id', '=', self.production_plan_id.id),
            ],
            order='sequence asc',
            limit=1,
        )
        self.operation_id = production_operation.operation_id
        self.product_model_id = self.production_plan_id.model_id
        self.bom_id = production_operation.operation_id.bom_id
        return production_operation

    @api.onchange('production_plan_id')
    def _onchange_of_product_plan_id(self):
        if not self.production_plan_id:
            return

        self.product_id = self.production_plan_id.product_id
        before = self.before_manufacturing_process_id

        if self.is_first_process:
            # first process compares only against other first processes
            existing_qty = sum(
                self.search([
                    ('production_plan_id', '=', self.production_plan_id.id),
                    ('is_first_process', '=', True),
                    ('id', '!=', self.id),
                ]).mapped('remaining_qty')
            )
            self.sequence = 1
            self.manufacturing_process_type_id = self.production_plan_id.first_process_type_id

            production_operation = self.env['production.operation'].search(
                [('manufacturing_process_type_id', '=', self.manufacturing_process_type_id.id),
                 ('production_plan_id', '=', self.production_plan_id.id)], limit=1)
            self.operation_id = production_operation.operation_id
            self.product_model_id = self.production_plan_id.model_id
            self.bom_id = production_operation.operation_id.bom_id
            self.product_qty = self.production_plan_id.expected_production_qty - existing_qty
            if not self.bom_id:
                self.component_ids = False

        elif self.is_sub_process:
            production_operation = self._resolve_bom()
            self.sequence = 1 + self.sequence
            before.is_sub_process_created = True
            self.manufacturing_process_type_id = production_operation.manufacturing_process_type_id
            self.operation_id = production_operation.operation_id  # kept: refreshed after type change
            self.product_qty = len(self.main_manufacturing_process_id.sub_process_lot_ids)

        else:
            # every remaining branch shares the same skeleton; only product_qty differs
            production_operation = self._resolve_bom()
            self.sequence = 1 + self.sequence
            self.manufacturing_process_type_id = production_operation.manufacturing_process_type_id
            self.operation_id = production_operation.operation_id

            if before.is_split_process and before.manufacturing_split_process1_id.id == self._origin.id:
                self.product_qty = len(before.voltage_lot_ids)
            elif before.is_split_process and before.manufacturing_split_process2_id.id == self._origin.id:
                self.product_qty = len(before.capacity_lot_ids)
            elif before.is_split_process and before.next_manufacturing_process_type_id == self.manufacturing_process_type_id:
                self.product_qty = len(before.packing_lot_ids)
            else:
                self.product_qty = before.remaining_qty

        self._onchange_of_operation()

    @api.onchange('manufacturing_process_type_id')
    def _onchange_of_operation_type(self):
        if self.manufacturing_process_type_id:
            machine = self.env['manufacturing.machine'].search([('manufacturing_process_type_id', '=', self.manufacturing_process_type_id)], limit=1)
            if machine:
                self.machine_id = machine.id

    @api.onchange('operation_id')
    def _onchange_of_operation(self):
        if self.operation_id and self.is_sub_process:
            production_location = self.env['stock.location'].search([('usage', '=', 'production')], limit=1)
            # self.product_id = self.operation_id.product_id.id
            self.location_src_id = self.before_manufacturing_process_id.location_src_id.id
            self.location_dest_id = self.operation_id.location_dest_id.id
            self.production_location_id = production_location.id
            # self.bom_id = self.operation_id.bom_id.id
            if self.allow_lot_create:
                self.product_id.write({
                    'tracking': 'serial',
                })
            else:
                self.product_id.write({
                    'tracking': 'none',
                })

        elif self.operation_id:
            production_location = self.env['stock.location'].search([('usage', '=', 'production')], limit=1)
            # self.product_id = self.operation_id.product_id.id
            self.location_src_id = self.operation_id.location_src_id.id
            self.location_dest_id = self.operation_id.location_dest_id.id
            self.production_location_id = production_location.id
            # self.bom_id = self.operation_id.bom_id.id
            if self.allow_lot_create:
                self.product_id.write({
                    'tracking': 'serial',
                })
            else:
                self.product_id.write({
                    'tracking': 'none',
                })
        else:
            self.bom_id = False
            self.component_ids =False
        if self.component_ids:
            for line in self.component_ids:
                line.location_src_id = self.location_src_id.id
                line.location_dest_id = self.production_location_id.id
        # self.bom_id = self.operation_id.bom_id.id
        self._onchange_bom_id()
        self._onchange_of_operation_type()

    @api.onchange('product_id')
    def _onchange_product_id(self):
        if self._origin:
            if self.product_id:
                self.bom_id = False  # ← remove the trailing comma
                self.component_ids = False

    @api.onchange('bom_id', 'product_qty')
    def _onchange_bom_id(self):
        self.input_material_lines = False
        self.component_ids = False

        if not self.bom_id:
            return

        child_records = []
        before = self.before_manufacturing_process_id

        if self.is_first_process:
            for line in self.bom_id.bom_line_ids:
                child_records.append((0, 0, {
                    'product_id': line.product_id.id,
                    'product_uom_category_id': line.product_uom_category_id.id,
                    'product_uom_id': line.product_uom_id.id,
                    'location_src_id': self.location_src_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'name': line.product_id.name,
                    'product_qty': (line.product_qty / self.bom_id.product_qty) * self.product_qty,
                    'production_plan_id': self.production_plan_id.id,
                    'manufacturing_process_id': self._origin.id or False,
                }))

        elif self.is_sub_process:
            for line in self.bom_id.bom_line_ids:
                child_records.append((0, 0, {
                    'product_id': line.product_id.id,
                    'product_uom_category_id': line.product_uom_category_id.id,
                    'product_uom_id': line.product_uom_id.id,
                    'location_src_id': self.location_src_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'name': line.product_id.name,
                    'product_qty': self.product_qty,
                    'production_plan_id': self.production_plan_id.id,
                    'manufacturing_process_id': self._origin.id or False,
                }))

        elif before.is_split_process and before.process_type1_id == self.manufacturing_process_type_id:
            for line in self.bom_id.bom_line_ids:
                child_records.append((0, 0, {
                    'product_id': line.product_id.id,
                    'product_uom_category_id': line.product_uom_category_id.id,
                    'product_uom_id': line.product_uom_id.id,
                    'location_src_id': self.location_src_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'name': line.product_id.name,
                    'product_qty': len(before.voltage_lot_ids),
                    'production_plan_id': self.production_plan_id.id,
                    'manufacturing_process_id': self._origin.id or False,
                }))

        elif before.is_split_process and before.process_type2_id == self.manufacturing_process_type_id:
            for line in self.bom_id.bom_line_ids:
                child_records.append((0, 0, {
                    'product_id': line.product_id.id,
                    'product_uom_category_id': line.product_uom_category_id.id,
                    'product_uom_id': line.product_uom_id.id,
                    'location_src_id': self.location_src_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'name': line.product_id.name,
                    'product_qty': len(before.capacity_lot_ids),
                    'production_plan_id': self.production_plan_id.id,
                    'manufacturing_process_id': self._origin.id or False,
                }))

        elif before.is_split_process and before.next_manufacturing_process_type_id == self.manufacturing_process_type_id:
            for line in self.bom_id.bom_line_ids:
                child_records.append((0, 0, {
                    'product_id': line.product_id.id,
                    'product_uom_category_id': line.product_uom_category_id.id,
                    'product_uom_id': line.product_uom_id.id,
                    'location_src_id': self.location_src_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'name': line.product_id.name,
                    'product_qty': len(before.packing_lot_ids),
                    'production_plan_id': self.production_plan_id.id,
                    'manufacturing_process_id': self._origin.id or False,
                }))

        else:
            remaining_qty = before.remaining_qty
            for line in self.bom_id.bom_line_ids:
                child_records.append((0, 0, {
                    'product_id': line.product_id.id,
                    'product_uom_category_id': line.product_uom_category_id.id,
                    'product_uom_id': line.product_uom_id.id,
                    'location_src_id': self.location_src_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'name': line.product_id.name,
                    'product_qty': remaining_qty,
                    'production_plan_id': self.production_plan_id.id,
                    'manufacturing_process_id': self._origin.id or False,
                }))

        self.input_material_lines = child_records

    @api.onchange('location_src_id')
    def onchange_location_src_id(self):
        for rec in self.input_material_lines:
            rec.location_src_id = self.location_src_id.id

    def get_serial_numbers(self):
        available_inputs = self.component_ids.filtered(lambda x: x.product_id.id == self.product_id.id and x.lot_id)
        child_records = []
        for line in available_inputs:
            child_records.append((0, 0, {
                'product_id': line.product_id.id,
                'product_uom_id': line.product_uom_id.id,
                'name': line.lot_id.name,
            }))
        self.lot_ids = child_records

    def clear_serial_numbers(self):
        self.lot_ids = False

    def check_available_stock(self):
        for rec in self:
            if rec.allow_lot_create:
                rec.product_id.write({
                    'tracking': 'serial',
                })
            else:
                rec.product_id.write({
                    'tracking': 'none',
                })
            rec.need_material_request = False
            rec.check_available = False
            for line in rec.input_material_lines:
                available_qty = line.product_id.get_available_quantity(line.location_src_id)
                if available_qty < line.product_qty:
                    rec.need_material_request = True
                elif available_qty >= line.product_qty:
                    rec.check_available = True
            for line in rec.component_ids:
                available_qty = line.product_id.get_available_quantity(line.location_src_id, line.lot_id)
                qty = line.product_uom_id._compute_quantity(line.product_qty, line.product_id.uom_id)
                line.available_qty = float(available_qty)
                if qty > available_qty:
                    line.check_available = True
                    rec.check_available = True
            # self.env.cr.commit()
            # return {
            #     'type': 'ir.actions.client',
            #     'tag': 'reload',
            # }

        # for line in self.input_material_lines:
            # if line.product_qty != line.qty_done:
            #     raise UserError(_("Required materials and reserved materials are not same."))
        # for line in self.component_ids:
        #     available_qty = line.product_id.get_available_quantity(line.location_src_id, line.lot_id)
        #     qty = line.product_uom_id._compute_quantity(line.product_qty, line.product_id.uom_id)
        #     line.available_qty = available_qty
        #     if qty > available_qty:
        #         raise UserError(
        #             _("Required quantity is not available in the stock for %s. Please check on %s" % (
        #             line.product_id.name, self.location_src_id.name)))
        #     line.check_available = True

    def action_start(self):
        for line in self.component_ids:
            available_qty = line.product_id.get_available_quantity(line.location_src_id, line.lot_id)
            qty = line.product_uom_id._compute_quantity(line.product_qty, line.product_id.uom_id)
            line.available_qty = float(available_qty)
            if qty > available_qty:
                raise UserError(
                    _("Required quantity is not available in the stock for %s. Please check on %s" % (
                        line.product_id.name, self.location_src_id.name)))
            line.check_available = True
            self.check_available = True
        for line in self.input_material_lines:
            if line.product_qty != line.qty_done:
                raise UserError(_("Required materials and reserved materials are not same."))
        for rec in self:
            if rec.is_first_process:
                existing_qty = sum(self.search(
                    [('production_plan_id', '=', rec.production_plan_id.id), ('id', '!=', rec.id),('is_first_process','=',True) ]).mapped(
                    'remaining_qty'))
                total_qty = existing_qty + rec.product_qty
                if existing_qty > rec.production_plan_id.expected_production_qty:
                    raise UserError(
                        _("Total Quantity (%s) cannot be greater than the Expected Production Quantity (%s).") % (
                            existing_qty, rec.production_plan_id.expected_production_qty))
            elif not self.is_sub_process and rec.production_plan_id:
                before_process = rec.before_manufacturing_process_id
                if rec.product_qty > before_process.remaining_qty:  #  rec not self
                    raise UserError(
                        _("Total Quantity (%s) cannot be greater than the remaining quantity of the previous process (%s).") % (
                            rec.product_qty,  #  rec not self
                            before_process.remaining_qty,
                        )
                    )
            # if rec.operation_id.allow_lot_create and not rec.lot_ids:
            #     raise UserError(
            #         _("This is a lot-enabled product. Please upload the Lot/Serial Number in the Serial Number tab.")
            #     )

            # for lot in rec.lot_ids:
            #     if not lot.lot_id and not lot.is_available:
            #         raise UserError(
            #             _("This product requires a Lot/Serial Number. Please create or select a Lot/Serial Number.")
            #         )

            rec.check_available_stock()

            rec.write({
                'state': 'progress',
                'start_time': fields.Datetime.now(),
            })

    # def action_done_production(self):
    #     self.with_delay(eta=10)._action_done_production()

    def _process_serial_moves(self, serials, dest_location):
        """
        Per-serial lot creation + stock move to a given destination location.
        Extracted from action_done_production so it can be reused for both
        the split-process case (3 destinations) and the normal case (1 destination),
        with identical logic in both cases.
        """
        for serial in serials:
            lot_id = serial.lot_id

            if not serial.lot_id:
                # check production_location_id first before creating a duplicate lot.
                # If lot_creation() already produced this serial's lot there, reuse it.
                existing_lot = self.env['stock.lot'].sudo().search([
                    ('name', '=', serial.name),
                    ('product_id', '=', self.product_id.id),
                ], limit=1)

                available_qty = 0
                if existing_lot:
                    available_qty = self.env['stock.quant']._get_available_quantity(
                        self.product_id,
                        self.production_location_id,
                        lot_id=existing_lot,
                    )

                if existing_lot and available_qty > 0:
                    # lot already exists & available in production_location_id -> reuse, don't recreate
                    lot_id = existing_lot
                    serial.write({'lot_id': lot_id.id})
                else:
                    if not self.allow_lot_create:
                        _logger.error(_(
                            "The lot %s is not available and The current operation "
                            "is not allowed to create new lot number.\n "
                            "Please enable lot creation or check the inventory." % serial.name
                        ))
                    lot_id = self.env['stock.lot'].sudo().create({
                        'name': serial.name,
                        'ref': serial.name,
                        'product_id': self.product_id.id,
                        'company_id': self.env.company.id,
                        'production_plan_id': self.production_plan_id.id,
                    })
                    #  Create quant with cell_weight + batch_id
                    update_stock = self.env['stock.quant'].sudo().create({
                        'product_id': self.product_id.id,
                        'location_id': self.production_location_id.id,
                        'lot_id': lot_id.id,
                        'inventory_quantity': 1.0,
                        'cell_weight': serial.cell_weight,
                        'batch_id': serial.batch_id.id if serial.batch_id else False,
                    })
                    update_stock.action_apply_inventory()
                    serial.write({'lot_id': lot_id.id})

            else:
                #  Lot exists — update existing quant cell_weight + batch_id
                existing_quant = self.env['stock.quant'].sudo().search([
                    ('product_id', '=', self.product_id.id),
                    ('location_id', '=', self.production_location_id.id),
                    ('lot_id', '=', lot_id.id),
                ], limit=1)
                if existing_quant:
                    existing_quant.write({
                        'cell_weight': serial.cell_weight,
                        'batch_id': serial.batch_id.id if serial.batch_id else False,
                    })
                else:
                    self.env['stock.quant'].sudo().create({
                        'product_id': self.product_id.id,
                        'location_id': self.production_location_id.id,
                        'lot_id': lot_id.id,
                        'inventory_quantity': 1.0,
                        'cell_weight': serial.cell_weight,
                        'batch_id': serial.batch_id.id if serial.batch_id else False,
                    })

            stock_move = self.env['stock.move'].sudo().create({
                'inventory_name': self.name,
                'product_id': self.product_id.id,
                'product_uom': self.product_uom_id.id,
                'product_uom_qty': 1,
                'location_id': self.production_location_id.id,
                'location_dest_id': dest_location.id,
                'manufacturing_process_id': self.id,
            })
            stock_move._action_confirm()
            stock_move._action_assign()
            existing_move_lines = self.env['stock.move.line'].search([
                ('move_id', '=', stock_move.id)
            ])
            if not existing_move_lines:
                self.env['stock.move.line'].sudo().create({
                    'move_id': stock_move.id,
                    'product_id': self.product_id.id,
                    'product_uom_id': self.product_uom_id.id,
                    'quantity': 1,
                    'location_id': self.production_location_id.id,
                    'location_dest_id': dest_location.id,
                    'company_id': self.company_id.id,
                    'lot_id': lot_id.id,
                    'manufacturing_process_id': self.id,
                    'out_manufacturing_process_id': self.id,
                })
                existing_move_lines = self.env['stock.move.line'].search([
                    ('move_id', '=', stock_move.id)
                ])
            existing_move_lines.write({
                'manufacturing_process_id': self.id,
                'out_manufacturing_process_id': self.id,
                'lot_id': lot_id.id,
            })
            stock_move.move_line_ids.picked = True
            stock_move._action_done()  # Moved here so dest quant exists before search below

            #  Update dest location quant cell_weight + batch_id after move done
            dest_quant = self.env['stock.quant'].sudo().search([
                ('product_id', '=', self.product_id.id),
                ('location_id', '=', dest_location.id),
                ('lot_id', '=', lot_id.id),
            ], limit=1)
            if dest_quant:
                dest_quant.write({
                    'cell_weight': serial.cell_weight,
                    'batch_id': serial.batch_id.id if serial.batch_id else False,
                })
            else:
                #  Dest quant missing — create it with batch_id
                self.env['stock.quant'].sudo().create({
                    'product_id': self.product_id.id,
                    'location_id': dest_location.id,
                    'lot_id': lot_id.id,
                    'inventory_quantity': 1.0,
                    'cell_weight': serial.cell_weight,
                    'batch_id': serial.batch_id.id if serial.batch_id else False,
                })

    def action_done_production(self):
        for rec in self:

            if self.allow_lot_create and not self.lot_ids:
                raise UserError("This is lot enabled product.Please upload the Lot/Serial Number in the Serial Number tab.")
            rec.lot_creation()
        # NEW: if this is a split process, trigger Capacity/Voltage lot creation first.
        # These already call action_create_packing_lot() internally, so packing_lot_ids
        # gets populated too as the "remaining" serials.
        for rec in self:
            if rec.is_split_process:
                if rec.process_type1_id:
                    rec.action_create_capacity_lot()
                if rec.process_type2_id:
                    rec.action_create_voltage_lot()

        for rec in self:
            rec.check_available_stock()
            stock_moves = []
            for line in rec.component_ids:
                stock_move_vals = {
                    'inventory_name': rec.name,
                    'product_id': line.product_id.id,
                    'product_uom': line.product_uom_id.id,
                    'product_uom_qty': line.product_qty,
                    'location_id': line.location_src_id.id,
                    'location_dest_id': rec.production_location_id.id,
                    'manufacturing_process_id': rec.id,
                    'move_line_ids': [(0, 0, {
                        'product_id': line.product_id.id,
                        'product_uom_id': line.product_id.uom_id.id,
                        'quantity': line.product_qty,
                        'location_id': line.location_src_id.id,
                        'location_dest_id': rec.production_location_id.id,
                        'company_id': rec.company_id.id,
                        'lot_id': line.lot_id.id if line.lot_id else False,
                        'manufacturing_process_id': rec.id,
                    })],
                }
                stock_moves.append(stock_move_vals)
            created_stock_moves = self.env['stock.move'].create(stock_moves)
            created_stock_moves._action_confirm()
            created_stock_moves._action_assign()
            for move, line in zip(created_stock_moves, rec.component_ids):
                if move.move_line_ids:
                    move.move_line_ids.write({
                        'quantity': line.product_qty,
                        'lot_id': line.lot_id.id if line.lot_id else False,
                        'manufacturing_process_id': rec.id,
                    })
                else:
                    self.env['stock.move.line'].create({
                        'move_id': move.id,
                        'product_id': line.product_id.id,
                        'product_uom_id': line.product_uom_id.id,
                        'quantity': line.product_qty,
                        'location_id': line.location_src_id.id,
                        'location_dest_id': rec.production_location_id.id,
                        'company_id': rec.company_id.id,
                        'lot_id': line.lot_id.id if line.lot_id else False,
                        'manufacturing_process_id': rec.id,
                    })
                move.move_line_ids.picked = True
            created_stock_moves._action_done()

        for lot in self.lot_ids:
            if lot.lot_id:
                available_qty = self.env['stock.quant']._get_available_quantity(
                    self.product_id,
                    self.production_location_id,
                    lot_id=lot.lot_id
                )
                if available_qty == 0:
                    _logger.error(_("Lot not available in %s" % self.production_location_id.name))

        if self.product_id.tracking == 'none':
            stock_move = self.env['stock.move'].create({
                'inventory_name': self.name,
                'product_id': self.product_id.id,
                'product_uom': self.product_uom_id.id,
                'product_uom_qty': self.product_qty,
                'location_id': self.production_location_id.id,
                'location_dest_id': self.location_dest_id.id,
                'manufacturing_process_id': self.id,
            })
            stock_move._action_confirm()
            stock_move._action_assign()
            existing_move_lines = self.env['stock.move.line'].search([
                ('move_id', '=', stock_move.id)
            ])
            if not existing_move_lines:
                self.env['stock.move.line'].create({
                    'move_id': stock_move.id,
                    'product_id': self.product_id.id,
                    'product_uom_id': self.product_uom_id.id,
                    'quantity': self.product_qty,
                    'location_id': self.production_location_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'company_id': self.company_id.id,
                    'manufacturing_process_id': self.id,
                    'out_manufacturing_process_id': self.id,
                })
                existing_move_lines = self.env['stock.move.line'].search([
                    ('move_id', '=', stock_move.id)
                ])
            existing_move_lines.write({
                'manufacturing_process_id': self.id,
                'out_manufacturing_process_id': self.id,
            })
            stock_move.move_line_ids.picked = True

            dest_quant = self.env['stock.quant'].sudo().search([
                ('product_id', '=', self.product_id.id),
                ('location_id', '=', self.location_dest_id.id),
            ], limit=1)
            if dest_quant:
                dest_quant.write({
                    'batch_id': self.batch_id.id if self.batch_id else False,
                })
            else:
                self.env['stock.quant'].sudo().create({
                    'product_id': self.product_id.id,
                    'location_id': self.location_dest_id.id,
                    'inventory_quantity': self.product_qty,
                    'batch_id': self.batch_id.id if self.batch_id else False,
                })

        else:
            # NEW: route serials to different destinations when is_split_process
            if self.is_split_process:
                if self.packing_lot_ids:
                    self._process_serial_moves(self.packing_lot_ids, self.location_dest_id)
                if self.capacity_lot_ids:
                    self._process_serial_moves(self.capacity_lot_ids, self.capacity_dest_location_id)
                if self.voltage_lot_ids:
                    self._process_serial_moves(self.voltage_lot_ids, self.voltage_dest_location_id)
            else:
                # UNCHANGED: exact same behavior as before this refactor
                self._process_serial_moves(self.lot_ids, self.location_dest_id)

        if self.allow_lot_create:
            self.state = 'done'
        else:
            self.state = 'close'
        self.end_time = fields.Datetime.now()
        self.action_close()

    def action_close(self):
        for move_line in self.finished_move_ids:
            exist_serial = self.lot_ids.filtered(
                lambda x: x.name == move_line.lot_id.name
            )
            if exist_serial:
                exist_serial[0].write({'tray_id': move_line.out_tray_id.id})
            else:
                if move_line.out_tray_id and move_line.lot_id:
                    self.env['product.serial.number'].create({
                        'product_id': move_line.product_id.id,
                        'product_uom_id': move_line.product_uom_id.id,
                        'lot_id': move_line.lot_id.id,
                        'name': move_line.lot_id.name,
                        'tray_id': move_line.out_tray_id.id,
                    })

        stock_moves = self.finished_move_ids.mapped('move_id')
        stock_moves._action_done()

        for rec in self:
            all_lines = rec.production_plan_id.operation_ids.sorted('sequence')
            # Next process = the operation line whose source location matches
            # this process's destination location. Same rule for every process
            # (split, branch, or normal).
            next_op_line = all_lines.filtered(
                lambda x: x.operation_id.location_src_id.id == rec.location_dest_id.id
            )[:1]

            if next_op_line:
                rec.next_manufacturing_process_type_id = next_op_line.manufacturing_process_type_id.id
            else:
                rec.next_manufacturing_process_id = False
                rec.next_manufacturing_process_type_id = False
            for serial in rec.lot_ids:
                if serial.lot_id:
                    quants = self.env['stock.quant'].sudo().search([
                        ('product_id', '=', rec.product_id.id),
                        ('lot_id', '=', serial.lot_id.id),
                    ])
                    if quants:
                        quants.write({
                            'cell_weight': serial.cell_weight,
                            'batch_id': serial.batch_id.id if serial.batch_id else False,  #  batch_id
                        })
                    else:
                        #  No quants at all — create with batch_id
                        self.env['stock.quant'].sudo().create({
                            'product_id': rec.product_id.id,
                            'location_id': rec.location_dest_id.id,
                            'lot_id': serial.lot_id.id,
                            'inventory_quantity': 1.0,
                            'cell_weight': serial.cell_weight,
                            'batch_id': serial.batch_id.id if serial.batch_id else False,  #  batch_id
                        })


    def action_view_product_move(self):
        # Get move lines directly linked to this process
        direct_lines = self.env['stock.move.line'].search([
            ('manufacturing_process_id', '=', self.id),
        ]).ids

        # Get move lines created from quality checks of this process
        quality_line_ids = self.env['stock.move.line'].search([
            ('quality_id.manufacturing_process_id', '=', self.id),
        ]).ids

        all_line_ids = list(set(direct_lines + quality_line_ids))

        return {
            'res_model': 'stock.move.line',
            'type': 'ir.actions.act_window',
            'name': _("Stock Move"),
            'domain': [('id', 'in', all_line_ids)],
            'view_mode': 'list,form',
        }

    @api.constrains('product_qty', 'production_plan_id')
    def _check_product_qty(self):
        for rec in self:
            if not rec.production_plan_id:
                continue

            if rec.is_first_process:
                if rec.product_qty <= 0:
                    raise UserError(
                        _("Zero Quantity (%s) cannot be processed.") % rec.product_qty
                    )

                existing_qty = sum(
                    self.search([
                        ('production_plan_id', '=', rec.production_plan_id.id),
                        ('is_first_process', '=', True),
                        ('id', '!=', rec.id),
                    ]).mapped('remaining_qty')
                )
                total_qty = existing_qty + rec.product_qty

                if total_qty > rec.production_plan_id.expected_production_qty:
                    raise UserError(
                        _("Total Quantity (%s) cannot be greater than the Expected Production Quantity (%s).") % (
                            total_qty,
                            rec.production_plan_id.expected_production_qty
                        )
                    )
            else:
                if not rec.before_manufacturing_process_id:
                    continue
                before_process = rec.before_manufacturing_process_id
                if rec.product_qty > before_process.remaining_qty:
                    raise UserError(
                        _("Total Quantity (%s) cannot be greater than the remaining quantity of the previous process (%s).") % (
                            rec.product_qty,
                            before_process.remaining_qty,
                        )
                    )


    def create_scrap_product(self):
        for rec in self:
            location_reject_id = self.env['stock.location'].search([('usage', '=', 'inventory')], limit=1)
            if rec.operation_id.location_reject_id:
                location_reject_id = rec.operation_id.location_reject_id
            lot_ids = self.finished_move_ids.filtered(lambda x: x.lot_id).mapped('lot_id')
            return {
                'name': _('Scrap Production?'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'manufacturing.scrap',
                'target': 'new',
                'context': {
                    'default_mrp_lot_ids': lot_ids.ids,
                    'default_name': rec.name,
                    'default_location_reject_id': location_reject_id.id,
                    'default_production_plan_id': self.production_plan_id.id,
                    'default_operation_id': rec.operation_id.id,
                    'default_product_id': rec.product_id.id,
                    'default_product_uom_id': rec.product_uom_id.id,
                    'default_location_src_id': rec.location_dest_id.id,
                    'default_manufacturing_process_id': rec.id,
                },
            }


    def action_update_lot_product(self):
        for rec in self:
            return {
                'name': _('Update Lot'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'tray.product.lot',
                'target': 'new',
                'context': {
                    'default_manufacturing_process_id': rec.id,
                    'default_dest_location_id': rec.location_dest_id.id,
                    'default_location_src_id':rec.location_src_id.id
                },
            }

    def action_break_production(self):
        for rec in self:
            view = self.env.ref('fnet_mrp.production_breakdown_form_view1')
            return {
                'name': _('Production  Breakdown'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'views': [(view.id, 'form')],
                'view_id': view.id,
                'res_model': 'production.breakdown',
                'target': 'new',
                'context': {
                    'default_manufacturing_process_id': rec.id,
                    'default_type': 'HOLD',
                },
            }

    def action_restart_production(self):
        for rec in self:
            view = self.env.ref('fnet_mrp.production_breakdown_form_view2')
            return {
                'name': _('Production  Breakdown'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'views': [(view.id, 'form')],
                'view_id': view.id,
                'res_model': 'production.breakdown',
                'target': 'new',
                'context': {
                    'default_manufacturing_process_id': rec.id,
                    'default_type': 'RESTART',
                },
            }

class ProductionBreakdown(models.Model):
    _inherit = 'production.breakdown'

    manufacturing_process_id = fields.Many2one('manufacturing.process')
    end_time = fields.Datetime(string='End Time')

    def action_done_breakdown(self):
        for rec in self:
            if rec.manufacturing_process_id:
                if rec.type == 'HOLD':
                    rec.manufacturing_process_id.state = 'hold'
                else:
                    rec.manufacturing_process_id.state = 'progress'

                rec._send_breakdown_email()

        return super(ProductionBreakdown, self).action_done_breakdown()

    def _send_breakdown_email(self):
        self.ensure_one()
        process = self.manufacturing_process_id

        manager_group = self.env.ref('fnet_mrp.group_manufacturing_manager', raise_if_not_found=False)
        if not manager_group:
            return

        managers = self.env['res.users'].sudo().search([
            ('group_ids', 'in', [manager_group.id]),
            ('active', '=', True),
        ])

        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        base_url += '/web#id=%d&view_type=form&model=%s' % (process.id, process._name)

        # Dynamic color scheme based on breakdown type
        # HOLD -> red, RESTART -> green
        if self.type == 'RESTART':
            main_color = '#2e7d32'  # green (text/accent)
            header_color = '#66bb6a'  # mild green for header/footer
            light_bg = '#eaf7ea'  # light green background
            border_light = '#c4e6c4'  # light green border
            header_title = 'Production Restarted'
        else:
            main_color = '#c62828'  # red (text/accent)
            header_color = '#e57373'  # mild red for header/footer
            light_bg = '#fdeaea'  # light red background
            border_light = '#f0c4c4'  # light red border
            header_title = 'Production %s Reported' % (self.type or '')

        if self.type == 'RESTART':
            intro_text = (
                    'Production has been restarted. Root Cause: <b style="color:%s;">%s</b>'
                    % (main_color, self.root_cause or '-')
            )
        else:
            intro_text = (
                    'Production breakdown Reason : <b style="color:%s;">%s</b>'
                    % (main_color, self.reason or '-')
            )

        for user in managers:
            if not user.partner_id.email:
                continue

            body_html = """
            <div style="font-family: Arial, sans-serif; margin: 0 auto;
                        border: 1px solid #c0c0c0; border-radius: 8px; overflow: hidden;">

                <div style="background-color: %s; padding: 12px 32px;">
                    <h2 style="color: #ffffff; margin: 0; font-size: 16px;">
                        %s
                    </h2>
                </div>

                <div style="padding: 32px; background-color: #ffffff;">

                    <p style="color: #1a1a1a; font-size: 15px;">
                        Dear <strong>%s</strong>,
                    </p>

                    <p style="color: #444444; font-size: 14px; line-height: 1.6;">
                        Please review the details below.
                    </p>

                    <div style="background-color: %s;
                                border-left: 4px solid %s;
                                border-radius: 4px;
                                padding: 18px 20px;
                                margin: 20px 0;
                                color: %s;
                                font-size: 14px;
                                font-weight: 600;">

                        %s

                    </div>

                    <div style="margin:24px 0;">
                        <a href="%s"
                           style="display:inline-block;
                                  background-color:%s;
                                  color:#ffffff;
                                  text-decoration:none;
                                  padding:10px 24px;
                                  border-radius:6px;
                                  font-weight:600;">
                            View Manufacturing Process &#8594;
                        </a>
                    </div>

                    <p style="color:#444444;font-size:14px;">
                        Thanks &amp; Regards,
                    </p>

                    <p style="color:%s;font-size:15px;font-weight:bold;">
                        %s
                    </p>

                </div>

                <div style="background-color:%s;
                            padding:8px 32px;
                            text-align:center;">

                    <p style="color:#ffffff;font-size:11px;margin:0;">
                        This is an automated notification from
                        <strong style="color:#ffffff;">%s</strong>.
                    </p>

                </div>

            </div>
            """ % (
                header_color,
                header_title,
                user.name,
                light_bg, main_color, main_color,
                intro_text,
                base_url,
                main_color,
                main_color,
                self.env.user.name,
                header_color,
                process.company_id.name or 'Odoo ERP',
            )

            if self.type == 'RESTART':
                subject = 'Production Restarted - %s' % process.display_name
            else:
                subject = 'Production Breakdown - %s' % process.display_name


            self.env['mail.mail'].sudo().create({
                'auto_delete': False,
                'author_id': self.env.user.partner_id.id,
                'email_from': (
                        process.company_id.partner_id.email_formatted
                        or self.env.user.email_formatted
                        or self.env.ref('base.user_root').email_formatted
                ),
                'email_to': user.partner_id.email,
                'subject': subject,
                'body_html': body_html,
            }).send()


class StockQuant(models.Model):
    _inherit = 'stock.quant'

    cell_weight = fields.Float(string='Cell Weight (g)', digits=(16, 4))
    manufacturing_process_id = fields.Many2one('manufacturing.process')
    batch_id = fields.Many2one('manufacturing.batch')