from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import timedelta, datetime
import base64
# from xlrd import open_workbook
import io
from openpyxl import load_workbook
import logging

_logger = logging.getLogger(__name__)


class MachineData(models.Model):
    _inherit = 'machine.parameter'

    qr_code_printing_id = fields.Many2one('qr.code.printing')
    qr_code_printing_condition = fields.Selection([
        ('good', 'Good'),
        ('fail', 'Fail')
    ], compute='_compute_qr_code_printing_condition')
    manufacturing_process_id = fields.Many2one('manufacturing.process', string="Manufacturing Process")

    def _compute_qr_code_printing_condition(self):
        for rec in self:
            rec.qr_code_printing_condition = 'fail'
            if rec.qr_code_printing_id:
                if rec.qr_code_printing_id.operation_id.dry_min_temperature <= rec.temperature <= rec.qr_code_printing_id.operation_id.dry_max_temperature and abs(
                        rec.qr_code_printing_id.operation_id.dry_min_vacuum) <= abs(rec.vacuum) <= abs(
                        rec.qr_code_printing_id.operation_id.dry_max_vacuum):
                    rec.qr_code_printing_condition = 'good'
                else:
                    rec.qr_code_printing_condition = 'fail'


class Utilities(models.Model):
    _inherit = 'utility.parameter'

    qr_code_printing_id = fields.Many2one('qr.code.printing')
    qr_code_printing_utility_condition = fields.Selection([
        ('good', 'Good'),
        ('fail', 'Fail')
    ], compute='_compute_qr_code_printing_utility')
    manufacturing_process_id = fields.Many2one('manufacturing.process', string="Manufacturing Process")

    def _compute_qr_code_printing_utility(self):
        for rec in self:
            # rec.cell_drying_utility_condition = 'fail'
            if rec.qr_code_printing_id:
                if rec.qr_code_printing_id.operation_id.min_humidity <= rec.humidity <= rec.qr_code_printing_id.operation_id.max_humidity and rec.qr_code_printing_id.operation_id.min_temperature <= rec.temperature <= rec.qr_code_printing_id.operation_id.max_temperature:
                    rec.qr_code_printing_utility_condition = 'good'
                else:
                    rec.qr_code_printing_utility_condition = 'fail'


class QualityDetails(models.Model):
    _inherit = 'quality.parameter'

    qr_code_printing_id = fields.Many2one('qr.code.printing')
    manufacturing_process_id = fields.Many2one('manufacturing.process', string="Manufacturing Process")


class StockProductionLOt(models.Model):
    _inherit = 'stock.lot'

    qr_code_printing_id = fields.Many2one('qr.code.printing')
    manufacturing_process_id = fields.Many2one('manufacturing.process', string="Manufacturing Process")


class StockMove(models.Model):
    _inherit = 'stock.move'

    qr_code_printing_id = fields.Many2one('qr.code.printing')
    manufacturing_process_id = fields.Many2one('manufacturing.process', string="Manufacturing Process")


class ProductSerialNumber(models.Model):
    _inherit = 'product.serial.number'

    qr_code_printing_id = fields.Many2one('qr.code.printing')
    manufacturing_process_id = fields.Many2one('manufacturing.process', string="Manufacturing Process")


class StockMoveLIne(models.Model):
    _inherit = 'stock.move.line'

    qr_code_printing_id = fields.Many2one('qr.code.printing')
    out_qr_code_printing_id = fields.Many2one('qr.code.printing')
    out_tray_id = fields.Many2one('product.tray')
    manufacturing_process_id = fields.Many2one('manufacturing.process', string="Manufacturing Process")
    manufacturing_process_out_id = fields.Many2one('manufacturing.process', string="Manufacturing Process")


class ManufacturingComponents(models.Model):
    _inherit = 'manufacturing.component'

    qr_code_printing_id = fields.Many2one('qr.code.printing')
    manufacturing_process_id = fields.Many2one('manufacturing.process', string="Manufacturing Process")

    @api.onchange('product_id')
    def _onchange_product_qr_code_printing_id(self):
        if self.qr_code_printing_id:
            self.location_src_id = self.qr_code_printing_id.location_src_id.id
            self.location_dest_id = self.qr_code_printing_id.location_dest_id.id


class MaterialLine(models.Model):
    _inherit = 'material.line'

    qr_code_printing_id = fields.Many2one('qr.code.printing')
    manufacturing_process_id = fields.Many2one('manufacturing.process', string="Manufacturing Process")


class QrCodePrinting(models.Model):
    _name = 'qr.code.printing'
    _description = 'QR Code Printing'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = "id desc"

    @api.model
    def default_get(self, fields):
        defaults = super(QrCodePrinting, self).default_get(fields)
        production_location = self.env['stock.location'].search([('usage', '=', 'production')], limit=1)
        defaults['production_location_id'] = production_location.id
        return defaults

    def _get_default_product_uom_id(self):
        return self.env['uom.uom'].search([], limit=1, order='id').id

    type = fields.Selection([
        ('anode_slitting', 'Anode Slitting '),
        ('cathode_slitting', 'Cathode Slitting '),
        ('anode_drying', 'Anode Drying'),
        ('cathode_drying', 'Cathode Drying'),
        ('diaphragm_drying', 'Diaphragm Drying'),
        ('anode_electrode_making', 'Anode Electrode Making'),
        ('cathode_electrode_making', 'Cathode Electrode Making'),
        ('winding', 'Winding'),
        ('hot_press_jelly', 'Hot Press Jelly'),
        ('assembly', 'Assembly'),
        ('qr_code_print', 'QR Code Printing'),
        ('cell_drying', 'Cell Drying'),
        ('injection', 'Injection'),
        ('high_temperature', 'High Temperature'),
        ('cell_clamp_baking', 'Cell Clamp Baking'),
        ('ht_clamp_baking', 'HT + Cell Clamp Baking'),
        ('aged_formation_cell', 'Aged Formation Cell'),
        ('degas', 'Degas'),
        ('dsf', 'Double side Folding'),
        ('pad_printing', 'Pad Printing'),
        ('capacity_test', 'Capacity Test'),
        ('voltage_test', 'Voltage Test'),
        ('aged_formation_cell_2', 'Aged Formation Cell 2'),
        ('voltage_test_2', 'Voltage Test 2'),
        ('packing', 'Packing'),
        ('powerbank', 'Power Bank')
    ], default='qr_code_print')
    company_id = fields.Many2one(
        'res.company', 'Company', index=True,
        default=lambda self: self.env.company)
    location_src_id = fields.Many2one('stock.location')
    location_dest_id = fields.Many2one('stock.location')
    production_location_id = fields.Many2one('stock.location')
    lot_ids = fields.One2many('product.serial.number', 'manufacturing_process_id', string='Serial Number')

    name = fields.Char(
        'Reference', copy=False, readonly=True, default=lambda x: _('New'))
    product_id = fields.Many2one(
        'product.product', 'Product',
        domain="""[
               ('type', 'in', ['combo', 'consu']),
               '|',
                   ('company_id', '=', False),
                   ('company_id', '=', company_id)
           ]
           """,
        check_company=True,)
    bom_id = fields.Many2one(
        'manufacturing.bom', 'Bill of Material')
    component_ids = fields.One2many('manufacturing.component', 'manufacturing_process_id', string='Components', copy=False)
    finished_lines = fields.One2many('mrp.finished.line', 'manufacturing_process_id', string="Finished Lines", copy=False)
    state = fields.Selection([
        ('draft', 'Draft'),
        ('confirmed', 'Confirmed'),
        ('progress', 'In Progress'),
        ('hold', 'Hold'),
        ('done', 'Done'),
        ('close', 'Closed'),
        ('cancel', 'Cancelled')], string='State',
        copy=False, index=True, default='draft',
        store=True, tracking=True,
        help=" * Draft: The MO is not confirmed yet.\n"
             " * Confirmed: The MO is confirmed, the stock rules and the reordering of the components are trigerred.\n"
             " * In Progress: The production has started (on the MO or on the WO).\n"
             " * Close: The production is done, the MO has to be closed.\n"
             " * Done: The MO is closed, the stock moves are posted. \n"
             " * Cancelled: The MO has been cancelled, can't be confirmed anymore.")
    product_tracking = fields.Selection(related='product_id.tracking')
    product_tmpl_id = fields.Many2one('product.template', 'Product Template', related='product_id.product_tmpl_id')
    product_qty = fields.Float(
        'Quantity To Produce',
        default=1.0, digits='Product Unit of Measure',
        required=True, tracking=True,
       )
    product_uom_id = fields.Many2one(
        'uom.uom', 'Product Unit of Measure', default=_get_default_product_uom_id,
        readonly=True, required=True,
       domain="[('relative_uom_id', '=', product_uom_category_id)]")

    qty_producing = fields.Float(string="Quantity Producing", digits='Product Unit of Measure', copy=False)
    product_uom_category_id = fields.Many2one(related='product_id.uom_id.relative_uom_id')
    product_uom_qty = fields.Float(string='Total Quantity', store=True)
    user_id = fields.Many2one(
        'res.users', 'Responsible', default=lambda self: self.env.user,
        states={'done': [('readonly', True)], 'cancel': [('readonly', True)]})
    qty_produced = fields.Float(string="Quantity Produced")
    operation_id = fields.Many2one('manufacturing.operation')
    production_plan_id = fields.Many2one('production.plan')
    injection_id= fields.Many2one('cell.injection')
    cell_drying_id= fields.Many2one('cell.drying')

    product_model_id = fields.Many2one('product.model')

    machine_data_ids = fields.One2many(
        'machine.parameter', 'manufacturing_process_id', 'Machine Data',
        copy=False, states={'done': [('readonly', True)], 'cancel': [('readonly', True)]})
    utility_ids = fields.One2many(
        'utility.parameter', 'manufacturing_process_id', 'Utility Parameter',
        copy=False, states={'done': [('readonly', True)], 'cancel': [('readonly', True)]})
    quality_ids = fields.One2many(
        'quality.parameter', 'manufacturing_process_id', 'Quality',
        copy=False, states={'done': [('readonly', True)], 'cancel': [('readonly', True)]})
    machine_id = fields.Many2one('manufacturing.machine')
    start_time = fields.Datetime()
    end_time = fields.Datetime()
    # product_tray_id = fields.Many2one('product.tray')
    assembly_id = fields.Many2one('assembly.cell')
    breakdown_ids = fields.One2many('production.breakdown', 'manufacturing_process_id')
    assembly_count = fields.Integer(compute='_compute_assembly_count')
    cell_drying_count = fields.Integer(compute='_compute_cell_drying_count')
    # injection_count = fields.Integer(compute='_compute_injection_count')
    finished_move_ids = fields.One2many('stock.move.line', 'manufacturing_process_out_id')
    expected_end_time = fields.Datetime("Expected End", compute='compute_end_date')
    out_file_name = fields.Char("Serial File")
    out_file = fields.Binary("Serials")
    input_material_lines = fields.One2many('material.line', 'manufacturing_process_id')
    remaining_hours = fields.Float("Remaining Time", compute='_get_remaining_time')
    process_status_check = fields.Boolean(compute='_compute_process_status_check')


    # @api.depends('pa_count')
    def _compute_process_status_check(self):
        for rec in self:
            process = self.env['first.article.inspection'].search(
                [('origin', '=', rec.name), ('state', '=', 'done'), ('inspection_type', '=', 'process')])
            if process:
                rec.process_status_check = True
            else:
                rec.process_status_check = False

    @api.onchange('location_src_id')
    def onchange_line(self):
        for rec in self.component_ids:
            rec.location_src_id = self.location_src_id

    def action_view_quality(self):
        return {
            'res_model': 'mrp.quality',
            'type': 'ir.actions.act_window',
            'name': _("Quality Check"),
            'domain': [('manufacturing_process_id', '=', self.id)],
            'view_mode': 'list,form',
        }

    def action_view_in_process_quality(self):
        in_process_records = self.env['process.quality.check'].search([('opt_id', '=', self.operation_id.id), ('origin', '=', self.name)])
        return {
            'res_model': 'process.quality.check',
            'type': 'ir.actions.act_window',
            'name': _("In Process Quality Check"),
            'domain': [('id', 'in', in_process_records.ids if in_process_records else [])],
            'view_mode': 'list,form',
        }

    def _get_remaining_time(self):
        for rec in self:
            hr_remain = 0
            if rec.start_time and rec.operation_id.process_duration:
                end = rec.start_time + timedelta(hours=rec.operation_id.process_duration)
                hr_remain = ((end - fields.Datetime.now()).seconds / 3600)
            rec.remaining_hours = hr_remain

    def action_remove_line(self):
        self.component_ids.unlink()

    def action_download_sample(self):
        return {
            "type": "ir.actions.act_url",
            "url": '/fnet_mrp/static/Sample Serials.xlsx',
            "target": "new",
        }

    def action_upload_serial(self):
        if not self.out_file:
            raise UserError("Please upload the serial number updated file.")
        file_data = base64.b64decode(self.out_file)
        wb = load_workbook(filename=io.BytesIO(file_data))
        sheet = wb.active
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            row_values = list(row)
            self.env['product.serial.number'].create({
                'product_id': self.product_id.id,
                'name': str(row_values[0]),
                'product_uom_id': self.product_uom_id.id,
                'manufacturing_process_id': self.id
            })

    @api.depends('start_time', 'operation_id')
    def compute_end_date(self):
        for rec in self:
            if rec.start_time and rec.operation_id:
                rec.expected_end_time = rec.start_time + timedelta(hours=rec.operation_id.process_duration)
            else:
                rec.expected_end_time = False

    # @api.constrains('end_time', 'expected_end_time')
    def duration_constrain(self):
        for rec in self:
            if rec.end_time < rec.expected_end_time:
                time = timedelta(hours=rec.operation_id.process_duration)
                dt = datetime(2000, 1, 1) + time
                raise UserError("Minimum duration to stop process is %s hours." % dt.strftime("%H:%M"))

    def _compute_assembly_count(self):
        for rec in self:
            rec.assembly_count = self.env['assembly.cell'].search_count([('id', '=', rec.assembly_id.id)])

    def _compute_cell_drying_count(self):
        for rec in self:
            rec.cell_drying_count = self.env['cell.drying'].search_count([('qr_code_printing_id', '=', rec.id)])

    # def _compute_injection_count(self):
    #     for rec in self:
    #         rec.injection_count = self.env['cell.injection'].search_count([('cell_drying_id', '=', rec.id)])

    def action_assembly_record_record(self):
        records = self.env['assembly.cell'].search([('id', '=', self.assembly_id.id)])
        if len(records) == 1:
            return {
                'name': _('Assembly Cell'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'assembly.cell',
                'res_id': records.id,  # Assuming records is a single record
            }
        else:
            return {
                'name': _('Assembly Cell'),
                'type': 'ir.actions.act_window',
                'view_mode': 'list,form',
                'res_model': 'assembly.cell',
                'domain': [('id', '=', self.assembly_id.id)]
            }

    def show_cell_drying_record(self):
        records = self.env['cell.drying'].search([('qr_code_printing_id', '=', self.id)])
        if len(records) == 1:
            # If there is only one record, return the form view directly
            return {
                'name': _('Cell Drying'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'cell.drying',
                'res_id': records.id,  # Assuming records is a single record
            }
        else:
            # If there are multiple records, return the list and form view
            return {
                'name': _('Cell Drying'),
                'type': 'ir.actions.act_window',
                'view_mode': 'list,form',
                'res_model': 'cell.drying',
                'domain': [('qr_code_printing_id', '=', self.id)],
            }

    def show_cell_injection_record(self):
        records = self.env['cell.injection'].search([('cell_drying_id', '=', self.id)])
        if len(records) == 1:
            # If there is only one record, return the form view directly
            return {
                'name': _('Cell Injection'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'cell.injection',
                'res_id': records.id,  # Assuming records is a single record
            }
        else:
            # If there are multiple records, return the list and form view
            return {
                'name': _('Cell  Injection'),
                'type': 'ir.actions.act_window',
                'view_mode': 'list,form',
                'res_model': 'cell.injection',
                'domain': [('cell_drying_id', '=', self.id)],
            }

    def action_create_cell_drying(self):
        cell_drying = self.env['cell.drying'].create({
            'qr_code_printing_id': self.id,
            'production_plan_id': self.production_plan_id.id,
            'product_model_id': self.product_model_id.id
        })
        cell_drying._onchange_of_product_plan_id()
        return {
            'name': _('Cell Drying'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'cell.drying',
            'res_id': cell_drying.id
        }

    def action_create_injection_cell(self):
        injection_cell = self.env['cell.injection'].create({
            'cell_drying_id': self.id,
            'production_plan_id': self.production_plan_id.id,
            'product_model_id': self.product_model_id.id
        })
        injection_cell._onchange_of_product_plan_id()
        return {
            'name': _('Cell Injection'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'cell.injection',
            'res_id': injection_cell.id
        }

    @api.onchange('state')
    def _onchange_of_state(self):
        if self.state and self.component_ids:
            for line in self.component_ids:
                line.state = self.state

    @api.onchange('production_plan_id')
    def _onchange_of_product_plan_id(self):
        if self.production_plan_id:
            production_operation = self.env['production.operation'].search(
                [('operation_type', '=', self.type), ('production_plan_id', '=', self.production_plan_id.id)], limit=1)
            self.operation_id = production_operation.operation_id.id
            self.product_model_id = self.production_plan_id.model_id.id
            self.bom_id = self.operation_id.bom_id.id
            self.product_qty = self.production_plan_id.expected_production_qty
            if not self.bom_id:
                self.component_ids = False
            self._onchange_of_operation()

    @api.onchange('type')
    def _onchange_of_operation_type(self):
        if self.type:
            machine = self.env['manufacturing.machine'].search([('type', '=', self.type)], limit=1)
            if machine:
                self.machine_id = machine.id

    @api.onchange('operation_id')
    def _onchange_of_operation(self):
        if self.operation_id:
            production_location = self.env['stock.location'].search([('usage', '=', 'production')], limit=1)
            self.product_id = self.operation_id.product_id.id
            self.location_src_id = self.operation_id.location_src_id.id
            self.location_dest_id = self.operation_id.location_dest_id.id
            self.production_location_id = production_location.id
            self.bom_id = self.operation_id.bom_id.id
            self.product_id.write({
                'tracking': 'none',
            })
        else:
            self.bom_id = False
            self.component_ids =False
        if self.component_ids:
            for line in self.component_ids:
                line.location_src_id = self.location_src_id.id
                line.location_dest_id = self.production_location_id.id
        self.bom_id = self.operation_id.bom_id.id
        self._onchange_bom_id()
        self._onchange_of_operation_type()

    @api.onchange('product_id')
    def _onchange_product_id(self):
        if self._origin:
            if self.product_id:
                self.bom_id = False  # ← remove the trailing comma
                self.component_ids = False

    @api.onchange('bom_id', 'product_qty')
    def _onchange_bom_id(self):
        self.input_material_lines = False
        self.component_ids = False
        if self.bom_id:
            child_records = []
            for line in self.bom_id.bom_line_ids:
                child_records.append((0, 0, {
                    'product_id': line.product_id.id,
                    'product_uom_category_id': line.product_uom_category_id.id,
                    'product_uom_id': line.product_uom_id.id,
                    'location_src_id': self.location_src_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'name': line.product_id.name,
                    'product_qty': (line.product_qty / self.bom_id.product_qty) * self.product_qty
                }))
            self.input_material_lines = child_records

    @api.onchange('location_src_id')
    def onchange_location_src_id(self):
        for rec in self.input_material_lines:
            rec.location_src_id = self.location_src_id.id

    def get_serial_numbers(self):
        available_inputs = self.component_ids.filtered(lambda x: x.product_id.id == self.product_id.id and x.lot_id)
        child_records = []
        for line in available_inputs:
            child_records.append((0, 0, {
                'product_id': line.product_id.id,
                'product_uom_id': line.product_uom_id.id,
                'name': line.lot_id.name,
            }))
        self.lot_ids = child_records

    def clear_serial_numbers(self):
        self.lot_ids = False

    def check_available_stock(self):
        for line in self.component_ids:
            available_qty = line.product_id.get_available_quantity(line.location_src_id, line.lot_id)
            qty = line.product_uom_id._compute_quantity(line.product_qty, line.product_id.uom_id)
            line.available_qty = float(available_qty)
            if qty > available_qty:
                raise UserError(
                    _("Required quantity is not available in the stock for %s. Please check on %s" % (
                    line.product_id.name, self.location_src_id.name)))
            line.check_available = True
        for line in self.input_material_lines:
            if line.product_qty != line.qty_done:
                raise UserError(_("Required materials and reserved materials are not same."))
        # for line in self.component_ids:
        #     available_qty = line.product_id.get_available_quantity(line.location_src_id, line.lot_id)
        #     qty = line.product_uom_id._compute_quantity(line.product_qty, line.product_id.uom_id)
        #     line.available_qty = available_qty
        #     if qty > available_qty:
        #         raise UserError(
        #             _("Required quantity is not available in the stock for %s. Please check on %s" % (
        #             line.product_id.name, self.location_src_id.name)))
        #     line.check_available = True

    def action_start(self):
        for rec in self:
            rec.state = 'progress'
            self.start_time = fields.Datetime.now()

    # def action_done_production(self):
    #     self.with_delay(eta=10)._action_done_production()

    def action_done_production(self):
        for rec in self:
            rec.check_available_stock()
            stock_moves = []
            for line in rec.component_ids:
                stock_move_vals = {
                    'inventory_name': rec.name,
                    'product_id': line.product_id.id,
                    'product_uom': line.product_uom_id.id,
                    'product_uom_qty': line.product_qty,
                    'location_id': line.location_src_id.id,
                    'location_dest_id': rec.production_location_id.id,
                    'manufacturing_process_id': rec.id,
                    'move_line_ids': [(0, 0, {
                        'product_id': line.product_id.id,
                        'product_uom_id': line.product_id.uom_id.id,
                        'quantity': line.product_qty,
                        'location_id': line.location_src_id.id,
                        'location_dest_id': rec.production_location_id.id,
                        'company_id': rec.company_id.id,
                        'lot_id': line.lot_id.id if line.lot_id else False,
                        'manufacturing_process_id': rec.id,
                    })],
                }
                stock_moves.append(stock_move_vals)
            # Batch create stock moves
            created_stock_moves = self.env['stock.move'].create(stock_moves)
            # Confirm and assign
            created_stock_moves._action_confirm()
            created_stock_moves._action_assign()
            # Write lot + qty + custom fields onto assigned move lines
            for move, line in zip(created_stock_moves, rec.component_ids):
                if move.move_line_ids:
                    move.move_line_ids.write({
                        'quantity': line.product_qty,
                        'lot_id': line.lot_id.id if line.lot_id else False,
                        'manufacturing_process_id': rec.id,
                    })
                else:
                    self.env['stock.move.line'].create({
                        'move_id': move.id,
                        'product_id': line.product_id.id,
                        'product_uom_id': line.product_uom_id.id,
                        'quantity': line.product_qty,
                        'location_id': line.location_src_id.id,
                        'location_dest_id': rec.production_location_id.id,
                        'company_id': rec.company_id.id,
                        'lot_id': line.lot_id.id if line.lot_id else False,
                        'manufacturing_process_id': rec.id,
                    })
                move.move_line_ids.picked = True
            created_stock_moves._action_done()
        # if self.product_id.tracking != 'none' and len(self.lot_ids) != self.product_qty:
        #     raise UserError(_("Quantity to produce is %s but you have entered %s serial numbers." % (
        #     self.product_qty, len(self.lot_ids))))
        for lot in self.lot_ids:
            if lot.lot_id:
                available_qty = self.env['stock.quant']._get_available_quantity(self.product_id,
                                                                                self.production_location_id,
                                                                                lot_id=lot.lot_id)
                if available_qty == 0:
                    _logger.error(_("Lot not available in %s" % self.production_location_id.name))
                    # raise UserError(_("Lot not available in %s" % self.production_location_id.name))
        if self.product_id.tracking == 'none':
            stock_move = self.env['stock.move'].create({
                'inventory_name': self.name,
                'product_id': self.product_id.id,
                'product_uom': self.product_uom_id.id,
                'product_uom_qty': self.product_qty,
                'location_id': self.production_location_id.id,
                'location_dest_id': self.location_dest_id.id,
                'manufacturing_process_id': self.id,
            })
            stock_move._action_confirm()
            stock_move._action_assign()
            existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
            if not existing_move_lines:
                self.env['stock.move.line'].create({
                    'move_id': stock_move.id,
                    'product_id': self.product_id.id,
                    'product_uom_id': self.product_uom_id.id,
                    'quantity': self.product_qty,
                    'location_id': self.production_location_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'company_id': self.company_id.id,
                    'manufacturing_process_id': self.id,
                    'manufacturing_process_out_id': self.id,
                })
                existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
            existing_move_lines.write({
                'manufacturing_process_id': self.id,
                'manufacturing_process_out_id': self.id,
            })
            stock_move.move_line_ids.picked = True
        else:
            for serial in self.lot_ids:
                lot_id = serial.lot_id
                if not serial.lot_id:
                    if not self.operation_id.allow_lot_create:
                        _logger.error(
                            _("The lot %s is not available and The current operation is not allowed to create new lot number.\n Please enable lot creation or check the inventory." % serial.name))
                        # raise UserError(_("The lot %s is not available and The current operation is not allowed to create new lot number.\n Please enable lot creation or check the inventory." %serial.name))
                    lot_id = self.env['stock.lot'].create({
                        'name': serial.name,
                        'ref': serial.name,
                        'product_id': self.product_id.id,
                        'company_id': self.env.company.id,
                        'production_plan_id': self.production_plan_id.id
                    })
                    update_stock = self.env['stock.quant'].sudo().create({
                        'product_id': self.product_id.id,
                        'location_id': self.production_location_id.id,
                        'lot_id': lot_id.id,
                        'inventory_quantity': 1.0,
                    })
                    update_stock.action_apply_inventory()
                stock_move = self.env['stock.move'].create({
                    'inventory_name': self.name,
                    'product_id': self.product_id.id,
                    'product_uom': self.product_uom_id.id,
                    'product_uom_qty': 1,
                    'location_id': self.production_location_id.id,
                    'location_dest_id': self.location_dest_id.id,
                    'manufacturing_process_id': self.id,
                })
                stock_move._action_confirm()
                stock_move._action_assign()
                existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
                if not existing_move_lines:
                    self.env['stock.move.line'].create({
                        'move_id': stock_move.id,
                        'product_id': self.product_id.id,
                        'product_uom_id': self.product_uom_id.id,
                        'quantity': 1,
                        'location_id': self.production_location_id.id,
                        'location_dest_id': self.location_dest_id.id,
                        'company_id': self.company_id.id,
                        'lot_id': lot_id.id,
                        'manufacturing_process_id': self.id,
                        'manufacturing_process_out_id': self.id,
                    })
                    existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
                existing_move_lines.write({
                    'manufacturing_process_id': self.id,
                    'manufacturing_process_out_id': self.id,
                    'lot_id': lot_id.id
                })
                for move_line in existing_move_lines:
                    move_line.write({
                        'manufacturing_process_id': self.id,
                        'manufacturing_process_out_id': self.id,
                        'lot_id': lot_id.id
                    })
                stock_move.move_line_ids.picked = True
        self.state = 'done'
        self.end_time = fields.Datetime.now()

    def action_close(self):
        for move_line in self.finished_move_ids:
            exist_serial = self.lot_ids.filtered(lambda x: x.name == move_line.lot_id.name)
            if exist_serial:
                exist_serial[0].write({'tray_id': move_line.out_tray_id.id})
            else:
                if move_line.out_tray_id and move_line.lot_id:
                    self.env['product.serial.number'].create({
                        'product_id': move_line.product_id.id,
                        'product_uom_id': move_line.product_uom_id.id,
                        'lot_id': move_line.lot_id.id,
                        'name': move_line.lot_id.name,
                        'tray_id': move_line.out_tray_id.id,
                    })
        stock_moves = self.finished_move_ids.mapped('move_id')
        stock_moves._action_done()
        # for stock_move in stock_moves:
        #     # stock_move._action_confirm()  # Confirm the move
        #     stock_move._action_done()
        self.state = 'close'
        
    def action_view_product_move(self):
        return {
            'res_model': 'stock.move.line',
            'type': 'ir.actions.act_window',
            'name': _("Stock Move"),
            'domain': [('manufacturing_process_id', '=', self.id)],
            'view_mode': 'list,form',
        }

    def open_cell_drying_processing(self):
        return {
            'name': _('Cell Drying'),
            'type': 'ir.actions.act_window',
            'view_mode': 'list,form',
            'res_model': 'cell.drying',
            'domain': [('id', '=', self.cell_drying_id.id)]
        }

    def open_injection_processing(self):
        return {
            'name': _('Injection'),
            'type': 'ir.actions.act_window',
            'view_mode': 'list,form',
            'res_model': 'cell.injection',
            'domain': [('id', '=', self.injection_id.id)]
        }

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if not vals.get('name') or vals['name'] == _('New'):
                vals['name'] = self.env['ir.sequence'].next_by_code('qr.code.printing') or _('New')
        return super().create(vals_list)

    def create_scrap_product(self):
        for rec in self:
            location_reject_id = self.env['stock.location'].search([('usage', '=', 'inventory')], limit=1)
            if rec.operation_id.location_reject_id:
                location_reject_id = rec.operation_id.location_reject_id
            lot_ids = self.finished_move_ids.filtered(lambda x: x.lot_id).mapped('lot_id')
            return {
                'name': _('Scrap Production?'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'manufacturing.scrap',
                'target': 'new',
                'context': {
                    'default_mrp_lot_ids': lot_ids.ids,
                    'default_name': rec.name,
                    'default_location_reject_id': location_reject_id.id,
                    'default_operation_id': rec.operation_id.id,
                    'default_product_id': rec.product_id.id,
                    'default_product_uom_id': rec.product_uom_id.id,
                    'default_location_src_id': rec.location_dest_id.id,
                    'default_qr_code_printing_id': rec.id,
                },
            }


    def action_update_lot_product(self):
        for rec in self:
            return {
                'name': _('Update Lot'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'tray.product.lot',
                'target': 'new',
                'context': {
                    'default_qr_code_printing_id': rec.id,
                    'default_dest_location_id': rec.location_dest_id.id,
                    'default_location_src_id':rec.location_src_id.id
                },
            }

    def action_break_production(self):
        for rec in self:
            view = self.env.ref('fnet_mrp.production_breakdown_form_view2')
            return {
                'name': _('Production  Breakdown'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'views': [(view.id, 'form')],
                'view_id': view.id,
                'res_model': 'production.breakdown',
                'target': 'new',
                'context': {
                    'default_qr_code_printing_id': rec.id,
                    'default_type': 'HOLD',
                },
            }

    def action_restart_production(self):
        for rec in self:
            view = self.env.ref('fnet_mrp.production_breakdown_form_view2')
            return {
                'name': _('Production  Breakdown'),
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'views': [(view.id, 'form')],
                'view_id': view.id,
                'res_model': 'production.breakdown',
                'target': 'new',
                'context': {
                    'default_qr_code_printing_id': rec.id,
                    'default_type': 'RESTART',
                },
            }


class ProductionBreakdown(models.Model):
    _inherit = 'production.breakdown'

    qr_code_printing_id = fields.Many2one('cell.drying')
    end_time = fields.Datetime(string = 'End Time')

    manufacturing_process_id = fields.Many2one('manufacturing.process', string="Manufacturing Process")

    def action_done_breakdown(self):
        for rec in self:
            if rec.qr_code_printing_id:
                if rec.type == 'HOLD':
                    rec.qr_code_printing_id.state = 'hold'
                else:
                    rec.qr_code_printing_id.state = 'progress'

        return super(ProductionBreakdown, self).action_done_breakdown()

