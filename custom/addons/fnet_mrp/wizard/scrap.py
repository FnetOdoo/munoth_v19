from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
from openpyxl import load_workbook


class StockMoveLIne(models.Model):
    _inherit = 'stock.move.line'
    manufacturing_scrap_id = fields.Many2one('manufacturing.scrap')


class StockMove(models.Model):
    _inherit = 'stock.move'
    manufacturing_scrap_id = fields.Many2one('manufacturing.scrap')


class ManufacturingScrap(models.Model):
    _name = 'manufacturing.scrap'
    _description = "Manufacturing Scrp"
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = "id desc"

    name = fields.Char('Reference')
    product_id = fields.Many2one('product.product')
    mrp_lot_ids = fields.Many2many('stock.lot', 'mrp_lot_rel', string="MRP Lots")
    lot_ids = fields.Many2many('stock.lot', 'lot_select_rel', string='Lots')
    lot_id = fields.Many2one('stock.lot', string='Lot')
    scrap_location_id = fields.Many2one('stock.location')
    location_src_id = fields.Many2one('stock.location')
    location_reject_id = fields.Many2one('stock.location', string="Rejected Location")
    company_id = fields.Many2one(
        'res.company', 'Company', index=True,
        default=lambda self: self.env.company)
    tracking = fields.Selection([
        ('serial', 'Serial'),
        ('lot', 'Lot'),
        ('none', 'No Tracking')
    ], related='product_id.tracking')
    owner_id = fields.Many2one('res.partner', 'Owner', states={'done': [('readonly', True)]}, check_company=True)
    move_id = fields.Many2one('stock.move', 'Scrap Move', readonly=True, check_company=True, copy=False)
    picking_id = fields.Many2one('stock.picking', 'Picking', states={'done': [('readonly', True)]}, check_company=True)
    scrap_qty = fields.Float('Quantity', default=1.0)
    state = fields.Selection([
        ('draft', 'Draft'),
        ('done', 'Done')],
        string='Status', default="draft", readonly=True, tracking=True)
    date_done = fields.Datetime('Date', readonly=True)
    product_uom_id = fields.Many2one(
        'uom.uom', 'Unit of Measure',
        required=True, domain="[('relative_uom_id', '=', product_uom_category_id)]")
    product_uom_category_id = fields.Many2one(related='product_id.uom_id.relative_uom_id')
    reason = fields.Text(required=True)
    production_plan_id = fields.Many2one('production.plan', string='Production Plan', required=True)

    manufacturing_process_id = fields.Many2one('manufacturing.process')

    operation_id = fields.Many2one('manufacturing.operation', string="Mrp Operation")
    cell_drying_id = fields.Many2one('cell.drying', string="Cell Drying")
    injection_id = fields.Many2one('cell.injection', string="Injection")
    ht_cell_id = fields.Many2one('high.temperature.cell', string="High Temperature")
    clamp_baking_id = fields.Many2one('cell.clamp.baking')
    degas_id = fields.Many2one('degas.cell')
    capacity_id = fields.Many2one('capacity.test')
    voltage_test_id = fields.Many2one('voltage.test')
    lot_file_name = fields.Char(string='File Name')
    lot_file = fields.Binary(string='Lot File')

    def action_upload_serial(self):
        if not self.lot_file:
            raise UserError("Please upload the serial number file.")
        file_data = base64.b64decode(self.lot_file)
        wb = load_workbook(filename=io.BytesIO(file_data))
        sheet = wb.active
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            row_values = list(row)
            lot = self.env['stock.lot'].search([('name', '=', row_values)])
            # self.mrp_lot_ids = [(4, lot.id)]
        # return {
        #     'name': _('Scrap Production?'),
        #     'type': 'ir.actions.act_window',
        #     'view_mode': 'form',
        #     'res_model': 'manufacturing.scrap',
        #     'res_id': self.id,
        # }

    @api.onchange('product_id')
    def _onchange_of_product_id(self):
        if self.product_id:
            self.product_uom_id = self.product_id.uom_id.id

    def action_view_product_move(self):
        return {
            'res_model': 'stock.move.line',
            'type': 'ir.actions.act_window',
            'name': _("Stock Move"),
            'domain': [('manufacturing_scrap_id', '=', self.id)],
            'view_mode': 'list,form',
        }

    def action_create_scrap_product(self):
        for rec in self:
            if self.lot_file:
                file_data = base64.b64decode(self.lot_file)
                wb = load_workbook(filename=io.BytesIO(file_data))
                sheet = wb.active
                all_rows = list(sheet.iter_rows(min_row=2, values_only=True))
                for row in all_rows:
                    row_values = list(row)
                    lot = self.env['stock.lot'].search([('name', 'in', row_values)])
                    if not lot:
                        raise UserError(_("No Lot found for the number %s" % rec.location_src_id.name))
                for row in all_rows:
                    row_values = list(row)
                    lot = self.env['stock.lot'].search([('name', 'in', row_values)])
                    available_qty = rec.product_id.get_available_quantity(rec.location_src_id, lot)
                    qty = rec.product_uom_id._compute_quantity(rec.scrap_qty, rec.product_id.uom_id)
                    if qty > available_qty or qty == 0:
                        raise UserError(_("Required quantity is not available in the stock for %s. Please check on %s" % (rec.product_id.name, rec.location_src_id.name)))
                    existing_quality_id = self.env['mrp.quality'].search([('production_plan_id', '=', self.production_plan_id.id), ('operation_id', '=', self.operation_id.id), ('lot_id', '=', lot.id)])
                    if existing_quality_id:
                        raise UserError(_("Quality check is already created for the lot %s") % lot.name)
                    if not self.production_plan_id:
                        raise UserError('Kindly select the Production Plan')
                    quality_id = self.env['mrp.quality'].create({
                        'state': 'request',
                        'production_plan_id': self.production_plan_id.id,
                        'operation_id': self.operation_id.id,
                        'product_id': self.product_id.id,
                        'product_uom_id': self.product_uom_id.id,
                        'location_src_id': self.location_reject_id.id,
                        'location_dest_id': self.location_src_id.id,
                        'lot_id': lot.id,
                        'reason': self.reason,
                        'quantity': self.scrap_qty,
                        'manufacturing_process_id': self.manufacturing_process_id.id,
                    })
                    # """Creating moves"""
                    stock_move = self.env['stock.move'].create({
                        'inventory_name': rec.name,
                        'product_id': rec.product_id.id,
                        'product_uom': rec.product_uom_id.id,
                        'product_uom_qty': self.scrap_qty,
                        'location_id': rec.location_src_id.id,
                        'location_dest_id': rec.location_reject_id.id,
                        'manufacturing_process_id': self.manufacturing_process_id.id,

                    })
                    stock_move._action_confirm()
                    stock_move._action_assign()
                    existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
                    if not existing_move_lines:
                        self.env['stock.move.line'].create({
                            'move_id': stock_move.id,
                            'product_id': rec.product_id.id,
                            'product_uom_id': rec.product_id.uom_id.id,
                            'quantity': self.scrap_qty,
                            'location_id': rec.location_src_id.id,
                            'location_dest_id': rec.location_reject_id.id,
                            'company_id': rec.company_id.id,
                            'lot_id': lot.id,
                            'manufacturing_process_id': self.manufacturing_process_id.id,

                        })
                        existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
                    for move_line in existing_move_lines:
                        move_line.write({
                            'lot_id': lot.id or False,
                            'quality_id': quality_id.id,
                        })
                    stock_move.move_line_ids.picked = True
                    stock_move._action_done()
            elif rec.lot_ids:
                for lot in rec.lot_ids:
                    available_qty = rec.product_id.get_available_quantity(rec.location_src_id, lot)
                    qty = rec.product_uom_id._compute_quantity(rec.scrap_qty, rec.product_id.uom_id)
                    if qty > available_qty or qty == 0:
                        raise UserError(_("Required quantity is not available in the stock for %s. Please check on %s" % (
                            rec.product_id.name, rec.location_src_id.name)))
                    existing_quality_id = self.env['mrp.quality'].search([('operation_id', '=', self.operation_id.id), ('lot_id', '=', lot.id)])
                    if existing_quality_id:
                        raise UserError(_("Quality check is already created for the lot %s") % lot.name)
                    quality_id = self.env['mrp.quality'].create({
                        'state': 'request',
                        'operation_id': self.operation_id.id,
                        'product_id': self.product_id.id,
                        'product_uom_id': self.product_uom_id.id,
                        'location_src_id': self.location_reject_id.id,
                        'location_dest_id': self.location_src_id.id,
                        'lot_id': lot.id,
                        'reason': self.reason,
                        'quantity': self.scrap_qty,
                        'manufacturing_process_id': self.manufacturing_process_id.id,
                    })
                    # # """Creating moves"""
                    stock_move = self.env['stock.move'].create({
                        'inventory_name': rec.name,
                        'product_id': rec.product_id.id,
                        'product_uom': rec.product_uom_id.id,
                        'product_uom_qty': self.scrap_qty,
                        'location_id': rec.location_src_id.id,
                        'location_dest_id': rec.location_reject_id.id,
                        'manufacturing_process_id': self.manufacturing_process_id.id,

                    })
                    stock_move._action_confirm()
                    stock_move._action_assign()
                    existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
                    if not existing_move_lines:
                        self.env['stock.move.line'].create({
                            'move_id': stock_move.id,
                            'product_id': rec.product_id.id,
                            'product_uom_id': rec.product_id.uom_id.id,
                            'quantity': self.scrap_qty,
                            'location_id': rec.location_src_id.id,
                            'location_dest_id': rec.location_reject_id.id,
                            'company_id': rec.company_id.id,
                            'lot_id': lot.id,
                            'manufacturing_process_id': self.manufacturing_process_id.id,

                        })
                        existing_move_lines = self.env['stock.move.line'].search([('move_id', '=', stock_move.id)])
                    for move_line in existing_move_lines:
                        move_line.write({
                            'lot_id': lot.id or False,
                            'quality_id': quality_id.id,
                        })
                    stock_move.move_line_ids.picked = True
                    stock_move._action_done()
            else:
                raise UserError('Upload the lot file or select the lot')