from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from openpyxl import load_workbook
import xlrd
import base64
import io
import logging

_logger = logging.getLogger(__name__)


class UpdateBoxesWizard(models.TransientModel):
    _name = "update.boxes.wizard"

    upload_file = fields.Binary(string="Upload file", tracking=True)
    file_name = fields.Char('File Name', tracking=True)
    stock_picking_id = fields.Many2one('stock.picking', string='Stock Picking')
    picking_type_code = fields.Selection(related='stock_picking_id.picking_type_code', readonly=True)

    def action_update_boxes_receipt(self):
        if not self.upload_file:
            raise ValidationError(_('The uploaded file is empty. Please upload a valid file.'))

        try:
            file_data = base64.b64decode(self.upload_file)
        except Exception as e:
            _logger.error("Base64 decode failed: %s", repr(e))
            raise ValidationError(_('The uploaded file could not be decoded. Please re-upload.'))

        rows_data = []

        try:
            wb = load_workbook(filename=io.BytesIO(file_data))
            sheet = wb.active
            rows_data = [
                row for row in sheet.iter_rows(min_row=2, values_only=True)
                if any(cell is not None and str(cell).strip() != '' for cell in row)
            ]
        except Exception as e:
            _logger.error("openpyxl failed: %s", repr(e))
            try:
                wb = xlrd.open_workbook(file_contents=file_data)
                sheet = wb.sheet_by_index(0)
                rows_data = [
                    sheet.row_values(r) for r in range(1, sheet.nrows)
                    if any(str(v).strip() != '' for v in sheet.row_values(r))
                ]
            except Exception as e2:
                _logger.error("xlrd failed: %s", repr(e2))
                raise ValidationError(_('Upload a valid .xls or .xlsx file.'))

        if not rows_data:
            raise ValidationError(_('No data rows found in the uploaded file.'))

        picking = self.stock_picking_id
        if not picking:
            raise ValidationError(_('No linked Receipt found.'))

        move_lines = picking.move_line_ids.sorted('id')
        if not move_lines:
            raise ValidationError(_('No move lines found on this Receipt. Please check the picking first.'))

        if len(rows_data) > len(move_lines):
            raise ValidationError(_(
                'More rows in Excel (%s) than move lines (%s).'
            ) % (len(rows_data), len(move_lines)))

        picking_type = picking.picking_type_id
        allow_create = picking_type.use_create_lots
        allow_existing = picking_type.use_existing_lots

        not_found_serials = []
        blocked_serials = []

        for idx, row in enumerate(rows_data):
            lot_serial = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            raw_box_value = row[2] if len(row) > 2 else None
            boxes_value = str(raw_box_value).strip() if raw_box_value else ''

            if idx >= len(move_lines):
                continue

            move_line = move_lines[idx]
            product = move_line.product_id

            if boxes_value:
                move_line.boxes = boxes_value

            if not lot_serial:
                _logger.info("Row %s -> %s: no serial given, lot untouched", idx + 2, product.display_name)
                continue

            if not allow_create and not allow_existing:
                blocked_serials.append(lot_serial)
                continue

            existing_lot = self.env['stock.lot'].search([
                ('name', '=', lot_serial),
                ('product_id', '=', product.id),
            ], limit=1)

            if existing_lot:
                # Found an existing lot — always safe to assign, in any mode.
                # Write BOTH fields so the grid displays correctly no matter
                # which widget (char vs many2one) this operation type is using.
                move_line.write({
                    'lot_id': existing_lot.id,
                    'lot_name': lot_serial,
                })
                _logger.info("Row %s -> %s: assigned EXISTING lot '%s'", idx + 2, product.display_name, lot_serial)

            else:
                if allow_create:
                    lot = self.env['stock.lot'].create({
                        'name': lot_serial,
                        'product_id': product.id,
                        'company_id': self.env.company.id,
                    })
                    move_line.write({
                        'lot_id': lot.id,
                        'lot_name': lot_serial,
                    })
                    _logger.info("Row %s -> %s: CREATED new lot '%s'", idx + 2, product.display_name, lot_serial)
                else:
                    not_found_serials.append(lot_serial)

        errors = []
        if not_found_serials:
            errors.append(_(
                'This operation type only allows selecting EXISTING lots. '
                'The following serials do not exist and could not be assigned: %s'
            ) % ', '.join(not_found_serials))
        if blocked_serials:
            errors.append(_(
                'This operation type does not allow entering Lot/Serial Numbers. '
                'The following values in the file were ignored: %s'
            ) % ', '.join(blocked_serials))

        if errors:
            raise ValidationError('\n'.join(errors))

        return {'type': 'ir.actions.client', 'tag': 'reload'}


    def action_update_boxes_delivery(self):
        if not self.upload_file:
            raise ValidationError(_('The uploaded file is empty. Please upload a valid file.'))

        # --- Decode uploaded file ---
        try:
            file_data = base64.b64decode(self.upload_file)
        except Exception as e:
            _logger.error("Base64 decode failed: %s", repr(e))
            raise ValidationError(_('The uploaded file could not be decoded. Please re-upload.'))

        _logger.info("Decoded file size: %s bytes", len(file_data))
        _logger.info("First 8 bytes (hex): %s", file_data[:8].hex())

        rows_data = []

        # --- Try xlsx (openpyxl) first ---
        try:
            wb = load_workbook(filename=io.BytesIO(file_data))
            sheet = wb.active
            rows_data = [
                row for row in sheet.iter_rows(min_row=2, values_only=True)
                if any(cell is not None and str(cell).strip() != '' for cell in row)
            ]
        except Exception as e:
            _logger.error("openpyxl failed: %s", repr(e))
            # --- Fall back to legacy .xls (xlrd) ---
            try:
                wb = xlrd.open_workbook(file_contents=file_data)
                sheet = wb.sheet_by_index(0)
                rows_data = [
                    sheet.row_values(r) for r in range(1, sheet.nrows)
                    if any(str(v).strip() != '' for v in sheet.row_values(r))
                ]
            except Exception as e2:
                _logger.error("xlrd failed: %s", repr(e2))
                raise ValidationError(_(
                    'Upload a valid .xls or .xlsx file. '
                    'If the problem persists, please contact support with the file.'
                ))

        if not rows_data:
            raise ValidationError(_('No data rows found in the uploaded file.'))

        picking = self.stock_picking_id
        if not picking:
            raise ValidationError(_('No linked Delivery Order found.'))

        move_lines = picking.move_line_ids.sorted('id')
        if not move_lines:
            raise ValidationError(_('No move lines found on this Delivery Order. Please check availability first.'))

        if len(rows_data) > len(move_lines):
            raise ValidationError(_(
                'More rows in Excel (%s) than move lines (%s).'
            ) % (len(rows_data), len(move_lines)))

        # Clear old box values first
        move_lines.write({'boxes': False})

        missing_serials = []  # serial not found as an existing lot
        unavailable_serials = []  # lot exists but no stock at this location
        unmatched_serials = []  # lot exists & available, but no free move line found
        used_move_line_ids = set()
        box_serial_map = {}  # box_number -> [serials], for cross-check/summary

        for idx, row in enumerate(rows_data):
            # Column A = S.No (ignored), Column B = Lot/Serial Number, Column C = Boxes
            lot_serial = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            raw_box_value = row[2] if len(row) > 2 else None
            boxes_value = str(raw_box_value).strip() if raw_box_value else ''

            if not lot_serial:
                continue

            ref_product = move_lines[:1].product_id

            lot = self.env['stock.lot'].search([
                ('name', '=', lot_serial),
                ('product_id', '=', ref_product.id),
            ], limit=1)

            if not lot:
                missing_serials.append(lot_serial)
                continue

            available_qty = self.env['stock.quant']._get_available_quantity(
                ref_product,
                picking.location_id,
                lot_id=lot,
            )
            if available_qty <= 0:
                unavailable_serials.append(lot_serial)
                continue

            move_line = move_lines.filtered(
                lambda ml: not ml.lot_id and ml.id not in used_move_line_ids
            )[:1]

            if not move_line:
                unmatched_serials.append(lot_serial)
                continue

            move_line.lot_id = lot.id
            move_line.boxes = boxes_value
            used_move_line_ids.add(move_line.id)

            box_serial_map.setdefault(boxes_value or '(no box)', []).append(lot_serial)

            _logger.info(
                "Lot: %s -> Box: %s, Move Line ID: %s, Available Qty: %s",
                lot_serial, boxes_value, move_line.id, available_qty
            )

        # Consolidated error reporting
        errors = []
        if missing_serials:
            errors.append(_('Serials not found as existing lots: %s') % ', '.join(missing_serials))
        if unavailable_serials:
            errors.append(_('Serials found but not available at location "%s": %s') % (
                picking.location_id.display_name, ', '.join(unavailable_serials)
            ))
        if unmatched_serials:
            errors.append(
                _('Serials valid & available, but no free move line to assign: %s') % ', '.join(unmatched_serials))

        if errors:
            raise ValidationError('\n'.join(errors))

        for box, serials in box_serial_map.items():
            _logger.info("Box '%s' contains %s serial(s): %s", box, len(serials), serials)

        return {'type': 'ir.actions.client', 'tag': 'reload'}
